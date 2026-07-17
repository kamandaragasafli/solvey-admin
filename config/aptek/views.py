from calendar import monthrange
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Max, Prefetch, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from medicine.models import Medical

from .models import AnbarHereket, Aptek, Depo, DrugPrice, Qaime
from .pdf_import import QaimeParseError, _clean_aptek_name
from .services import import_qaime_pdf

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZ_MONTHS = {
    1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
    5: 'May', 6: 'İyun', 7: 'İyul', 8: 'Avqust',
    9: 'Sentyabr', 10: 'Oktyabr', 11: 'Noyabr', 12: 'Dekabr',
}


def _parse_date(value, fallback):
    if not value:
        return fallback
    try:
        return date.fromisoformat(value)
    except ValueError:
        return fallback


def _excel_qty(value):
    num = Decimal(str(value))
    if num == num.to_integral_value():
        return int(num), '0'
    return float(num), '0.########'


def _default_date_range(today=None):
    today = today or timezone.localdate()
    date_from = date(today.year, today.month, 1)
    _, last_day = monthrange(today.year, today.month)
    date_to = date(today.year, today.month, last_day)
    return date_from, date_to


EXCLUDE_SESSION_KEY = 'aptek_exclude_ids'
ACTIVE_DEPO_SESSION_KEY = 'aptek_active_depo_id'


def _ensure_default_depo():
    depo = Depo.objects.filter(is_default=True).first()
    if depo:
        return depo
    depo = Depo.objects.order_by('id').first()
    if depo:
        if not depo.is_default:
            depo.is_default = True
            depo.save(update_fields=['is_default'])
        return depo
    return Depo.objects.create(name='Əsas depo', is_default=True)


def _get_active_depo(request):
    depo_id = request.session.get(ACTIVE_DEPO_SESSION_KEY)
    depo = Depo.objects.filter(pk=depo_id).first() if depo_id else None
    if not depo:
        depo = _ensure_default_depo()
        request.session[ACTIVE_DEPO_SESSION_KEY] = depo.id
        request.session.modified = True
    return depo


def _set_active_depo(request, depo):
    request.session[ACTIVE_DEPO_SESSION_KEY] = depo.id
    request.session.modified = True


def _parse_exclude_ids(raw_values):
    exclude_ids = []
    for raw in raw_values:
        for part in str(raw).split(','):
            part = part.strip()
            if part.isdigit():
                exclude_ids.append(int(part))
    return sorted(set(exclude_ids))


def _get_session_exclude_ids(request):
    raw = request.session.get(EXCLUDE_SESSION_KEY) or []
    return _parse_exclude_ids(raw)


def _ledger_filters(request):
    today = timezone.localdate()
    default_from, default_to = _default_date_range(today)

    date_from = _parse_date(request.GET.get('from'), default_from)
    date_to = _parse_date(request.GET.get('to'), default_to)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    aptek_id = request.GET.get('aptek') or None
    if aptek_id in (None, '', 'all'):
        aptek_id = None

    status_filter = request.GET.get('status') or None
    if status_filter in (None, '', 'all'):
        status_filter = None

    # İstisnalar sidebar səhifəsindən session-da saxlanılır
    exclude_ids = _get_session_exclude_ids(request)
    if aptek_id and int(aptek_id) in exclude_ids:
        exclude_ids = [x for x in exclude_ids if x != int(aptek_id)]

    return date_from, date_to, aptek_id, status_filter, exclude_ids


def _exclude_query(exclude_ids):
    if not exclude_ids:
        return ''
    return '&'.join(f'exclude={eid}' for eid in exclude_ids)


def _return_filters(request):
    today = timezone.localdate()
    default_from, default_to = _default_date_range(today)

    date_from = _parse_date(request.GET.get('from'), default_from)
    date_to = _parse_date(request.GET.get('to'), default_to)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    aptek_id = request.GET.get('aptek') or None
    if aptek_id in (None, '', 'all'):
        aptek_id = None

    search = (request.GET.get('q') or '').strip()

    return date_from, date_to, aptek_id, search


def _aptekler_filters(request):
    """Apteklər səhifəsi üçün daha geniş default aralıq."""
    today = timezone.localdate()
    default_from = date(today.year, today.month, 1)
    _, last_day = monthrange(today.year, today.month)
    default_to = date(today.year, today.month, last_day)

    date_from = _parse_date(request.GET.get('from'), default_from)
    date_to = _parse_date(request.GET.get('to'), default_to)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    aptek_id = request.GET.get('aptek') or None
    if aptek_id in (None, '', 'all'):
        aptek_id = None

    search = (request.GET.get('q') or '').strip()
    return date_from, date_to, aptek_id, search


def _date_range_label(date_from, date_to):
    if date_from.month == date_to.month and date_from.year == date_to.year:
        return AZ_MONTHS.get(date_from.month, '')
    return f'{date_from.strftime("%d.%m.%Y")} — {date_to.strftime("%d.%m.%Y")}'


def _document_movement_context(request, *, doc_type, movement_type, note_prefix):
    date_from, date_to, aptek_id, search = _return_filters(request)
    depo = _get_active_depo(request)

    movements = (
        AnbarHereket.objects.filter(
            depo=depo,
            movement_type=movement_type,
            date__gte=date_from,
            date__lte=date_to,
        )
        .filter(
            Q(qaime__document_type=doc_type)
            | Q(note__istartswith=note_prefix)
            | Q(note__istartswith='Qaime')
        )
        .exclude(note=EVVEL_NOTE)
        .select_related('drug', 'aptek', 'qaime')
    )

    if aptek_id:
        movements = movements.filter(aptek_id=aptek_id)
    if search:
        movements = movements.filter(
            Q(drug__med_name__icontains=search)
            | Q(drug__med_full_name__icontains=search)
        )

    movements = movements.order_by('-date', 'aptek__name', 'drug__med_name')

    rows = []
    total_qty = Decimal('0')
    qaime_ids = set()
    for movement in movements:
        total_qty += movement.quantity
        if movement.qaime_id:
            qaime_ids.add(movement.qaime_id)
        rows.append({
            'date': movement.date,
            'aptek': movement.aptek.name if movement.aptek else '—',
            'qaime_number': movement.qaime.number if movement.qaime else '—',
            'drug': movement.drug.med_name,
            'drug_full': movement.drug.med_full_name or '',
            'quantity': movement.quantity,
            'pdf_url': movement.qaime.pdf.url if movement.qaime and movement.qaime.pdf else '',
        })

    return {
        'rows': rows,
        'aptekler': Aptek.objects.filter(depo=depo),
        'selected_aptek': str(aptek_id) if aptek_id else '',
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'search': search,
        'period_label': _date_range_label(date_from, date_to),
        'record_count': len(rows),
        'doc_count': len(qaime_ids),
        'total_qty': total_qty,
    }


def _qaimeler_context(request):
    date_from, date_to, aptek_id, search = _return_filters(request)
    depo = _get_active_depo(request)

    hereket_qs = (
        AnbarHereket.objects.filter(depo=depo, movement_type=AnbarHereket.MOVEMENT_OUT)
        .select_related('drug')
        .order_by('id')
    )
    qaime_qs = (
        Qaime.objects.filter(
            depo=depo,
            document_type=Qaime.DOC_QAIME,
            doc_date__gte=date_from,
            doc_date__lte=date_to,
        )
        .select_related('aptek')
        .prefetch_related(Prefetch('hereketler', queryset=hereket_qs))
        .order_by('-doc_date', '-id')
    )

    if aptek_id:
        qaime_qs = qaime_qs.filter(aptek_id=aptek_id)
    if search:
        qaime_qs = qaime_qs.filter(
            Q(hereketler__drug__med_name__icontains=search)
            | Q(hereketler__drug__med_full_name__icontains=search)
        ).distinct()

    rows = []
    total_qty = Decimal('0')
    for qaime in qaime_qs:
        hereketler = list(qaime.hereketler.all())
        qty = sum((h.quantity for h in hereketler), Decimal('0'))
        total_qty += qty

        names = []
        for hereket in hereketler:
            name = hereket.drug.med_name
            if name not in names:
                names.append(name)
        if not names:
            drug_label = '—'
        elif len(names) <= 2:
            drug_label = ', '.join(names)
        else:
            drug_label = f"{', '.join(names[:2])} +{len(names) - 2}"

        rows.append({
            'id': qaime.id,
            'date': qaime.doc_date,
            'aptek': qaime.aptek.name,
            'qaime_number': qaime.number,
            'drug_label': drug_label,
            'drug_count': len(names),
            'quantity': qty,
            'pdf_url': qaime.pdf.url if qaime.pdf else '',
        })

    return {
        'rows': rows,
        'aptekler': Aptek.objects.filter(depo=depo),
        'selected_aptek': str(aptek_id) if aptek_id else '',
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'search': search,
        'period_label': _date_range_label(date_from, date_to),
        'record_count': len(rows),
        'doc_count': len(rows),
        'total_qty': total_qty,
    }


def _anbara_elave_context(request):
    date_from, date_to, aptek_id, search = _return_filters(request)
    depo = _get_active_depo(request)

    movements = (
        AnbarHereket.objects.filter(
            depo=depo,
            movement_type=AnbarHereket.MOVEMENT_IN,
            date__gte=date_from,
            date__lte=date_to,
        )
        .exclude(
            Q(qaime__document_type=Qaime.DOC_RETURN)
            | Q(note__istartswith='Geri qaytarma')
        )
        .select_related('drug', 'aptek', 'qaime')
    )

    if aptek_id:
        movements = movements.filter(aptek_id=aptek_id)
    if search:
        movements = movements.filter(
            Q(drug__med_name__icontains=search)
            | Q(drug__med_full_name__icontains=search)
            | Q(note__icontains=search)
        )

    movements = movements.order_by('-date', 'note', 'drug__med_name')

    rows = []
    total_qty = Decimal('0')
    sources = set()
    for movement in movements:
        total_qty += movement.quantity
        source = movement.note or 'Anbara əlavə'
        sources.add(source)
        rows.append({
            'date': movement.date,
            'source': source,
            'aptek': movement.aptek.name if movement.aptek else '—',
            'drug': movement.drug.med_name,
            'drug_full': movement.drug.med_full_name or '',
            'quantity': movement.quantity,
        })

    return {
        'rows': rows,
        'aptekler': Aptek.objects.filter(depo=depo),
        'selected_aptek': str(aptek_id) if aptek_id else '',
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'search': search,
        'period_label': _date_range_label(date_from, date_to),
        'record_count': len(rows),
        'source_count': len(sources),
        'total_qty': total_qty,
    }


def _sum_qty(qs):
    return qs.aggregate(total=Sum('quantity'))['total'] or Decimal('0')


def _build_ledger(date_from, date_to, aptek_id=None, status_filter=None, exclude_ids=None, depo=None):
    exclude_ids = exclude_ids or []
    drugs = Medical.objects.filter(status=True).order_by('position', 'med_name')
    rows = []
    totals = {
        'evvel': Decimal('0'),
        'gelen': Decimal('0'),
        'cixan': Decimal('0'),
        'qalan': Decimal('0'),
    }

    for drug in drugs:
        base_qs = AnbarHereket.objects.filter(drug=drug)
        if depo is not None:
            base_qs = base_qs.filter(depo=depo)

        # Əvvələ qalıq: istisna apteklərin keçmiş çıxışları da çıxarılır
        in_before = base_qs.filter(
            movement_type=AnbarHereket.MOVEMENT_IN, date__lt=date_from
        )
        out_before = base_qs.filter(
            movement_type=AnbarHereket.MOVEMENT_OUT, date__lt=date_from
        )
        if exclude_ids:
            out_before = out_before.exclude(aptek_id__in=exclude_ids)
        evvel = _sum_qty(in_before) - _sum_qty(out_before)

        gelen = _sum_qty(
            base_qs.filter(
                movement_type=AnbarHereket.MOVEMENT_IN,
                date__gte=date_from,
                date__lte=date_to,
            )
        )

        out_qs = base_qs.filter(
            movement_type=AnbarHereket.MOVEMENT_OUT,
            date__gte=date_from,
            date__lte=date_to,
        )
        if aptek_id:
            out_qs = out_qs.filter(aptek_id=aptek_id)
        if exclude_ids:
            out_qs = out_qs.exclude(aptek_id__in=exclude_ids)
        cixan = _sum_qty(out_qs)

        qalan = evvel + gelen - cixan
        low = qalan <= 5

        if status_filter == 'low' and not low:
            continue
        if status_filter == 'ok' and low:
            continue

        rows.append({
            'ad': drug.med_name,
            'evvel': evvel,
            'gelen': gelen,
            'cixan': cixan,
            'qalan': qalan,
            'low': low,
        })

        totals['evvel'] += evvel
        totals['gelen'] += gelen
        totals['cixan'] += cixan
        totals['qalan'] += qalan

    return rows, totals


EVVEL_NOTE = 'Əvvələ qalıq'
ANBARA_ELAVE_NOTE = 'Anbara əlavə'


def _opening_date_for_month(year: int, month: int) -> date:
    return date(year, month, 1) - timedelta(days=1)


def _parse_month(value, fallback_year, fallback_month):
    if not value:
        return fallback_year, fallback_month
    try:
        parts = value.split('-')
        return int(parts[0]), int(parts[1])
    except (ValueError, IndexError):
        return fallback_year, fallback_month


def _user_context(request):
    today = timezone.localdate()
    user = request.user
    user_label = user.get_full_name() or user.username
    if user.email:
        user_label = user.email
    active_depo = _get_active_depo(request)
    return {
        'today': today.strftime('%d.%m.%Y'),
        'user_label': user_label,
        'active_depo': active_depo,
        'depolar': Depo.objects.all(),
    }


@login_required
def evvele_qaliq(request):
    depo = _get_active_depo(request)
    today = timezone.localdate()
    year, month = _parse_month(request.GET.get('month') or request.POST.get('month'), today.year, today.month)
    opening_date = _opening_date_for_month(year, month)
    selected_month = f'{year:04d}-{month:02d}'
    month_label = AZ_MONTHS.get(month, '')

    if request.method == 'POST':
        with transaction.atomic():
            saved = 0
            for drug in Medical.objects.filter(status=True):
                raw = request.POST.get(f'qty_{drug.id}', '').strip().replace(',', '.')
                if not raw:
                    continue
                try:
                    qty = Decimal(raw)
                except Exception:
                    continue
                if qty <= 0:
                    continue
                # Əvvəlki qeydlər silinmir — yalnız yeni giriş əlavə olunur
                AnbarHereket.objects.create(
                    depo=depo,
                    drug=drug,
                    movement_type=AnbarHereket.MOVEMENT_IN,
                    quantity=qty,
                    date=opening_date,
                    note=EVVEL_NOTE,
                )
                saved += 1

        messages.success(
            request,
            f'{month_label} {year} — qalıq düzəlişi əlavə olundu ({saved} dərman). Əvvəlki qeydlər saxlanıldı.',
        )
        return redirect(f"{reverse('aptek:evvele_qaliq')}?month={selected_month}")

    existing = {}
    for row in (
        AnbarHereket.objects.filter(
            depo=depo,
            note=EVVEL_NOTE,
            date=opening_date,
            movement_type=AnbarHereket.MOVEMENT_IN,
        )
        .values('drug_id')
        .annotate(total=Sum('quantity'))
    ):
        existing[row['drug_id']] = row['total'] or Decimal('0')

    drugs = []
    for drug in Medical.objects.filter(status=True).order_by('position', 'med_name'):
        drug.current_qty = existing.get(drug.id, Decimal('0'))
        drugs.append(drug)

    context = {
        'drugs': drugs,
        'selected_month': selected_month,
        'month_label': month_label,
        'opening_date': opening_date,
        **_user_context(request),
    }
    return render(request, 'evvele_qaliq.html', context)


@login_required
def anbara_elave_form(request):
    depo = _get_active_depo(request)
    today = timezone.localdate()
    selected_date = _parse_date(
        request.GET.get('date') or request.POST.get('date'),
        today,
    )
    note = (request.GET.get('note') or request.POST.get('note') or ANBARA_ELAVE_NOTE).strip()
    if not note:
        note = ANBARA_ELAVE_NOTE

    if request.method == 'POST':
        with transaction.atomic():
            AnbarHereket.objects.filter(
                depo=depo,
                movement_type=AnbarHereket.MOVEMENT_IN,
                date=selected_date,
                note=note,
                qaime__isnull=True,
            ).delete()

            saved = 0
            for drug in Medical.objects.filter(status=True):
                raw = request.POST.get(f'qty_{drug.id}', '').strip().replace(',', '.')
                if not raw:
                    continue
                try:
                    qty = Decimal(raw)
                except Exception:
                    continue
                if qty <= 0:
                    continue
                AnbarHereket.objects.create(
                    depo=depo,
                    drug=drug,
                    movement_type=AnbarHereket.MOVEMENT_IN,
                    quantity=qty,
                    date=selected_date,
                    note=note,
                )
                saved += 1

        messages.success(
            request,
            f'{selected_date.strftime("%d.%m.%Y")} — anbara əlavə yadda saxlanıldı ({saved} dərman).',
        )
        params = urlencode({'date': selected_date.isoformat(), 'note': note})
        return redirect(f"{reverse('aptek:anbara_elave_form')}?{params}")

    existing = {
        row['drug_id']: row['quantity']
        for row in AnbarHereket.objects.filter(
            depo=depo,
            movement_type=AnbarHereket.MOVEMENT_IN,
            date=selected_date,
            note=note,
            qaime__isnull=True,
        ).values('drug_id', 'quantity')
    }

    drugs = []
    for drug in Medical.objects.filter(status=True).order_by('position', 'med_name'):
        drug.qty = existing.get(drug.id, Decimal('0'))
        drugs.append(drug)

    context = {
        'drugs': drugs,
        'selected_date': selected_date.isoformat(),
        'note': note,
        **_user_context(request),
    }
    return render(request, 'anbara_elave_form.html', context)


@login_required
def geri_qaytarma(request):
    context = _document_movement_context(
        request,
        doc_type=Qaime.DOC_RETURN,
        movement_type=AnbarHereket.MOVEMENT_IN,
        note_prefix='Geri qaytarma',
    )
    context.update(_user_context(request))
    return render(request, 'geri_qaytarma.html', context)


@login_required
def anbara_elave_siyahi(request):
    context = _anbara_elave_context(request)
    context.update(_user_context(request))
    return render(request, 'anbara_elave.html', context)


@login_required
def qaimeler(request):
    context = _qaimeler_context(request)
    context.update(_user_context(request))
    return render(request, 'qaimeler.html', context)


@login_required
def qaime_delete(request, pk):
    if request.method != 'POST':
        return redirect('aptek:qaimeler')

    depo = _get_active_depo(request)
    qaime = Qaime.objects.filter(pk=pk, depo=depo).select_related('aptek').first()
    if not qaime:
        messages.error(request, 'Qaimə tapılmadı.')
        return redirect('aptek:qaimeler')

    aptek_name = qaime.aptek.name
    number = qaime.number
    doc_type = qaime.document_type
    doc_label = 'Geri qaytarma' if doc_type == Qaime.DOC_RETURN else 'Qaimə'
    redirect_name = (
        'aptek:geri_qaytarma' if doc_type == Qaime.DOC_RETURN else 'aptek:qaimeler'
    )

    with transaction.atomic():
        movement_count = AnbarHereket.objects.filter(qaime=qaime).count()
        if qaime.pdf:
            qaime.pdf.delete(save=False)
        qaime.delete()

    messages.success(
        request,
        f'{doc_label} №{number} ({aptek_name}) silindi — '
        f'{movement_count} anbar hərəkəti ləğv olundu.',
    )
    return redirect(redirect_name)


@login_required
def istisnalar(request):
    depo = _get_active_depo(request)
    aptekler = Aptek.objects.filter(depo=depo).order_by('name')
    for aptek in aptekler:
        clean_name = _clean_aptek_name(aptek.name)
        if aptek.name != clean_name:
            aptek.name = clean_name
            aptek.save(update_fields=['name'])

    if request.method == 'POST':
        exclude_ids = _parse_exclude_ids(request.POST.getlist('exclude'))
        valid_ids = set(
            Aptek.objects.filter(depo=depo, id__in=exclude_ids).values_list('id', flat=True)
        )
        exclude_ids = sorted(valid_ids)
        request.session[EXCLUDE_SESSION_KEY] = exclude_ids
        request.session.modified = True

        if exclude_ids:
            names = list(
                Aptek.objects.filter(id__in=exclude_ids).order_by('name').values_list('name', flat=True)
            )
            messages.success(
                request,
                'İstisnalar yadda saxlanıldı: ' + ', '.join(names),
            )
        else:
            messages.success(request, 'İstisna seçilməyib — bütün apteklər hesaba daxil edilir.')
        return redirect('aptek:istisnalar')

    exclude_ids = _get_session_exclude_ids(request)
    context = {
        'aptekler': aptekler,
        'exclude_ids': [str(x) for x in exclude_ids],
        'excluded_apteks': list(Aptek.objects.filter(depo=depo, id__in=exclude_ids).order_by('name')),
        **_user_context(request),
    }
    return render(request, 'istisnalar.html', context)


LOW_STOCK_QTY = Decimal('5')
_DEPO_COLORS = ('cyan', 'purple', 'green', 'pink', 'blue')


def _format_money(value):
    num = Decimal(str(value or 0))
    quantized = num.quantize(Decimal('0.01'))
    text = f'{quantized:,.2f}'.replace(',', ' ')
    return f'{text} ₼'


def _depo_stock_rows(depo):
    """Dərman üzrə cari qalıq (giriş − çıxış) seçilmiş depo üçün."""
    rows = (
        AnbarHereket.objects.filter(depo=depo)
        .values('drug_id', 'drug__med_name', 'drug__med_price')
        .annotate(
            qty_in=Coalesce(
                Sum('quantity', filter=Q(movement_type=AnbarHereket.MOVEMENT_IN)),
                Decimal('0'),
            ),
            qty_out=Coalesce(
                Sum('quantity', filter=Q(movement_type=AnbarHereket.MOVEMENT_OUT)),
                Decimal('0'),
            ),
        )
        .order_by('drug__med_name')
    )
    stock = []
    for row in rows:
        qty = (row['qty_in'] or Decimal('0')) - (row['qty_out'] or Decimal('0'))
        if qty == 0:
            continue
        price = row['drug__med_price'] or Decimal('0')
        stock.append({
            'name': row['drug__med_name'],
            'sku': f'DRM-{row["drug_id"]:04d}',
            'qty': qty,
            'min': LOW_STOCK_QTY,
            'unit': 'əd',
            'value': qty * price,
            'low': qty <= LOW_STOCK_QTY,
        })
    return stock


@login_required
def depolar(request):
    if request.method == 'POST':
        action = request.POST.get('action') or 'create'
        if action == 'switch':
            depo = Depo.objects.filter(pk=request.POST.get('depo_id')).first()
            if not depo:
                messages.error(request, 'Depo tapılmadı.')
            else:
                _set_active_depo(request, depo)
                messages.success(request, f'Aktiv depo: {depo.name}')
            next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('aptek:anbar_dashboard')
            return redirect(next_url)

        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Depo adı daxil edin.')
            return redirect('aptek:depolar')
        if Depo.objects.filter(name__iexact=name).exists():
            messages.error(request, 'Bu adda depo artıq var.')
            return redirect('aptek:depolar')

        depo = Depo.objects.create(name=name, is_default=False)
        _set_active_depo(request, depo)
        messages.success(
            request,
            f'"{depo.name}" əlavə olundu. Bu depo boş başlayır (bütün miqdarlar 0).',
        )
        return redirect('aptek:depolar')

    active = _get_active_depo(request)
    warehouses = []
    total_sku = 0
    total_low = 0
    total_value = Decimal('0')

    for idx, depo in enumerate(Depo.objects.all()):
        stock = _depo_stock_rows(depo)
        sku_count = len(stock)
        low_count = sum(1 for item in stock if item['low'])
        ok_count = sku_count - low_count
        value = sum((item['value'] for item in stock), Decimal('0'))
        fill_pct = int(round((ok_count / sku_count) * 100)) if sku_count else 0

        if depo.id == active.id:
            status = 'aktiv'
        elif depo.is_default:
            status = 'əsas'
        elif sku_count == 0:
            status = 'boş'
        else:
            status = 'hazır'

        aptek_count = Aptek.objects.filter(depo=depo).count()
        qaime_count = Qaime.objects.filter(depo=depo).count()

        warehouses.append({
            'id': depo.id,
            'name': depo.name,
            'code': f'DEP-{depo.id:03d}',
            'color': _DEPO_COLORS[idx % len(_DEPO_COLORS)],
            'status': status,
            'is_default': depo.is_default,
            'is_active': depo.id == active.id,
            'address': f'{aptek_count} aptek · {qaime_count} qaimə',
            'sku_count': sku_count,
            'low_count': low_count,
            'ok_count': ok_count,
            'value': _format_money(value),
            'fill_pct': fill_pct,
            'stock': stock,
        })

        total_sku += sku_count
        total_low += low_count
        total_value += value

    stats = {
        'total': len(warehouses),
        'sku': total_sku,
        'low': total_low,
        'value': _format_money(total_value),
    }

    context = {
        'warehouses': warehouses,
        'stats': stats,
        **_user_context(request),
    }
    return render(request, 'depolar.html', context)


def _qty_map_for_depo(depo, drug_ids=None):
    qs = AnbarHereket.objects.filter(depo=depo)
    if drug_ids is not None:
        qs = qs.filter(drug_id__in=drug_ids)
    rows = qs.values('drug_id').annotate(
        qty_in=Coalesce(
            Sum('quantity', filter=Q(movement_type=AnbarHereket.MOVEMENT_IN)),
            Decimal('0'),
        ),
        qty_out=Coalesce(
            Sum('quantity', filter=Q(movement_type=AnbarHereket.MOVEMENT_OUT)),
            Decimal('0'),
        ),
    )
    return {
        row['drug_id']: (row['qty_in'] or Decimal('0')) - (row['qty_out'] or Decimal('0'))
        for row in rows
    }


def _drug_status(qty, expiry_date, today):
    if expiry_date and expiry_date < today:
        return 'expired'
    if qty <= 0:
        return 'expired'
    if expiry_date and expiry_date <= today + timedelta(days=90):
        return 'soon'
    if qty <= LOW_STOCK_QTY:
        return 'low'
    return 'ok'


@login_required
def dermanlar(request):
    depo = _get_active_depo(request)
    today = timezone.localdate()
    _THUMB = ('cyan', 'purple', 'green', 'pink', 'blue')

    if request.method == 'POST':
        drug_id = request.POST.get('drug_id')
        drug = Medical.objects.filter(pk=drug_id, status=True).first()
        if not drug:
            messages.error(request, 'Dərman seçin.')
            return redirect('aptek:dermanlar')

        try:
            price = Decimal(str(request.POST.get('price') or '').replace(',', '.'))
        except Exception:
            messages.error(request, 'Qiymət düzgün deyil.')
            return redirect('aptek:dermanlar')

        expiry_raw = (request.POST.get('expiry_date') or '').strip()
        expiry_date = None
        if expiry_raw:
            try:
                expiry_date = date.fromisoformat(expiry_raw)
            except ValueError:
                messages.error(request, 'SKT tarixi düzgün deyil.')
                return redirect('aptek:dermanlar')

        DrugPrice.objects.update_or_create(
            depo=depo,
            drug=drug,
            defaults={
                'price': price,
                'expiry_date': expiry_date,
            },
        )
        messages.success(request, f'{drug.med_name} qiyməti yadda saxlanıldı.')
        return redirect('aptek:dermanlar')

    default_depo = Depo.objects.filter(is_default=True).first() or depo
    DrugPrice.objects.filter(depo__isnull=True).update(depo=default_depo)

    drugs = Medical.objects.filter(status=True).order_by('position', 'med_name')
    price_by_drug = {
        dp.drug_id: dp
        for dp in DrugPrice.objects.filter(depo=depo).select_related('drug')
    }
    qty_map = _qty_map_for_depo(depo, list(drugs.values_list('id', flat=True)))

    medicines = []
    total_value = Decimal('0')
    low_count = 0
    expiring_count = 0

    for idx, drug in enumerate(drugs):
        dp = price_by_drug.get(drug.id)
        qty = qty_map.get(drug.id, Decimal('0'))
        price = dp.price if dp else None
        expiry = dp.expiry_date if dp else None
        status = _drug_status(qty, expiry, today)
        if status == 'low':
            low_count += 1
        if status == 'soon':
            expiring_count += 1
        if qty > 0 and price is not None:
            total_value += qty * price

        parts = drug.med_name.split()
        initials = (parts[0][:1] + (parts[1][:1] if len(parts) > 1 else '')).upper() or '?'

        medicines.append({
            'id': drug.id,
            'name': drug.med_name,
            'sku': f'DRM-{drug.id:04d}',
            'category': '—',
            'price': _format_money(price) if price is not None else '—',
            'quantity': qty,
            'expiry': expiry.strftime('%d.%m.%Y') if expiry else '—',
            'warehouse': depo.name,
            'status': status,
            'color': _THUMB[idx % len(_THUMB)],
            'initials': initials,
            'has_price': price is not None,
        })

    priced_ids = set(price_by_drug.keys())
    available_drugs = drugs.exclude(pk__in=priced_ids)

    context = {
        'medicines': medicines,
        'stats': {
            'total': len(medicines),
            'low': low_count,
            'expiring': expiring_count,
            'value': _format_money(total_value),
        },
        'available_drugs': available_drugs,
        **_user_context(request),
    }
    return render(request, 'dermanlar.html', context)


@login_required
def derman_detail(request, pk):
    depo = _get_active_depo(request)
    drug = Medical.objects.filter(pk=pk, status=True).first()
    if not drug:
        messages.error(request, 'Dərman tapılmadı.')
        return redirect('aptek:dermanlar')

    today = timezone.localdate()
    dp = DrugPrice.objects.filter(depo=depo, drug=drug).first()
    qty = _qty_map_for_depo(depo, [drug.id]).get(drug.id, Decimal('0'))
    expiry = dp.expiry_date if dp else None
    status = _drug_status(qty, expiry, today)

    movements = (
        AnbarHereket.objects.filter(depo=depo, drug=drug)
        .select_related('aptek', 'qaime')
        .order_by('-date', '-id')[:100]
    )
    out_total = _sum_qty(
        AnbarHereket.objects.filter(
            depo=depo, drug=drug, movement_type=AnbarHereket.MOVEMENT_OUT
        )
    )

    context = {
        'item': dp,
        'medicine': {
            'name': drug.med_name,
            'sku': f'DRM-{drug.id:04d}',
            'expiry_date': expiry.strftime('%d.%m.%Y') if expiry else None,
            'form': None,
        },
        'stats': {
            'stock': qty,
            'out': out_total,
            'warehouses': 1,
            'status': status,
            'price': _format_money(dp.price) if dp else '—',
        },
        'movements': movements,
        **_user_context(request),
    }
    return render(request, 'derman_detail.html', context)


@login_required
def aptekler(request):
    depo = _get_active_depo(request)
    date_from, date_to, _, search_from_filters = _aptekler_filters(request)
    search = (request.GET.get('q') or search_from_filters or '').strip()

    date_filter = Q(
        anbar_hereketleri__date__gte=date_from,
        anbar_hereketleri__date__lte=date_to,
    )

    aptek_qs = (
        Aptek.objects.filter(depo=depo).annotate(
            qaime_count=Count(
                'anbar_hereketleri__qaime',
                filter=date_filter
                & Q(
                    anbar_hereketleri__movement_type=AnbarHereket.MOVEMENT_OUT,
                    anbar_hereketleri__qaime__document_type=Qaime.DOC_QAIME,
                ),
                distinct=True,
            ),
            return_count=Count(
                'anbar_hereketleri__qaime',
                filter=date_filter
                & Q(
                    anbar_hereketleri__movement_type=AnbarHereket.MOVEMENT_IN,
                    anbar_hereketleri__qaime__document_type=Qaime.DOC_RETURN,
                ),
                distinct=True,
            ),
            last_activity=Max(
                'anbar_hereketleri__date',
                filter=date_filter,
            ),
            out_qty=Sum(
                'anbar_hereketleri__quantity',
                filter=date_filter
                & Q(anbar_hereketleri__movement_type=AnbarHereket.MOVEMENT_OUT),
            ),
            in_return_qty=Sum(
                'anbar_hereketleri__quantity',
                filter=date_filter
                & Q(
                    anbar_hereketleri__movement_type=AnbarHereket.MOVEMENT_IN,
                    anbar_hereketleri__qaime__document_type=Qaime.DOC_RETURN,
                ),
            ),
        )
        .order_by('name')
    )

    if search:
        aptek_qs = aptek_qs.filter(name__icontains=search)

    rows = []
    total_qaime = 0
    total_return = 0
    for aptek in aptek_qs:
        clean_name = _clean_aptek_name(aptek.name)
        if aptek.name != clean_name:
            aptek.name = clean_name
            aptek.save(update_fields=['name'])

        qaime_count = aptek.qaime_count or 0
        return_count = aptek.return_count or 0
        total_qaime += qaime_count
        total_return += return_count
        rows.append({
            'id': aptek.id,
            'name': aptek.name,
            'qaime_count': qaime_count,
            'return_count': return_count,
            'last_activity': aptek.last_activity,
            'out_qty': aptek.out_qty or Decimal('0'),
            'in_return_qty': aptek.in_return_qty or Decimal('0'),
        })

    context = {
        'rows': rows,
        'search': search,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'period_label': _date_range_label(date_from, date_to),
        'aptek_count': len(rows),
        'total_qaime': total_qaime,
        'total_return': total_return,
        **_user_context(request),
    }
    return render(request, 'aptekler.html', context)


@login_required
def aptek_detail(request, pk):
    depo = _get_active_depo(request)
    aptek = Aptek.objects.filter(pk=pk, depo=depo).first()
    if not aptek:
        messages.error(request, 'Aptek tapılmadı.')
        return redirect('aptek:aptekler')

    today = timezone.localdate()
    default_from, default_to = _default_date_range(today)
    date_from = _parse_date(request.GET.get('from'), default_from)
    date_to = _parse_date(request.GET.get('to'), default_to)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    search = (request.GET.get('q') or '').strip()
    type_filter = request.GET.get('type') or 'all'
    if type_filter not in {'all', 'out', 'in'}:
        type_filter = 'all'

    movements = (
        AnbarHereket.objects.filter(
            depo=depo,
            aptek=aptek,
            date__gte=date_from,
            date__lte=date_to,
        )
        .select_related('drug', 'qaime')
        .order_by('-date', '-id')
    )

    if type_filter == 'out':
        movements = movements.filter(movement_type=AnbarHereket.MOVEMENT_OUT)
    elif type_filter == 'in':
        movements = movements.filter(movement_type=AnbarHereket.MOVEMENT_IN)

    if search:
        movements = movements.filter(
            Q(drug__med_name__icontains=search)
            | Q(drug__med_full_name__icontains=search)
            | Q(note__icontains=search)
        )

    rows = []
    total_out = Decimal('0')
    total_in = Decimal('0')
    qaime_ids = set()
    return_ids = set()

    for movement in movements:
        is_out = movement.movement_type == AnbarHereket.MOVEMENT_OUT
        if is_out:
            total_out += movement.quantity
        else:
            total_in += movement.quantity

        if movement.qaime_id:
            if movement.qaime.document_type == Qaime.DOC_RETURN:
                return_ids.add(movement.qaime_id)
            else:
                qaime_ids.add(movement.qaime_id)

        if movement.qaime:
            if movement.qaime.document_type == Qaime.DOC_RETURN:
                doc_label = f'Geri qaytarma №{movement.qaime.number}'
            else:
                doc_label = f'Qaimə №{movement.qaime.number}'
        else:
            doc_label = movement.note or '—'

        rows.append({
            'date': movement.date,
            'type': movement.movement_type,
            'type_label': 'Çıxış' if is_out else 'Giriş',
            'doc_label': doc_label,
            'drug': movement.drug.med_name,
            'drug_full': movement.drug.med_full_name or '',
            'quantity': movement.quantity,
            'note': movement.note or '',
            'pdf_url': movement.qaime.pdf.url if movement.qaime and movement.qaime.pdf else '',
            'qaime_id': movement.qaime_id,
        })

    context = {
        'aptek': aptek,
        'rows': rows,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'search': search,
        'type_filter': type_filter,
        'period_label': _date_range_label(date_from, date_to),
        'record_count': len(rows),
        'qaime_count': len(qaime_ids),
        'return_count': len(return_ids),
        'total_out': total_out,
        'total_in': total_in,
        **_user_context(request),
    }
    return render(request, 'aptek_detail.html', context)


@login_required
def export_ledger_excel(request):
    depo = _get_active_depo(request)
    date_from, date_to, aptek_id, status_filter, exclude_ids = _ledger_filters(request)
    rows, totals = _build_ledger(date_from, date_to, aptek_id, status_filter, exclude_ids, depo=depo)

    aptek_label = 'Bütün apteklər'
    if aptek_id:
        aptek = Aptek.objects.filter(pk=aptek_id).first()
        if aptek:
            aptek_label = aptek.name
    if exclude_ids:
        names = list(
            Aptek.objects.filter(id__in=exclude_ids).order_by('name').values_list('name', flat=True)
        )
        if names:
            aptek_label = f'{aptek_label} (istisna: {", ".join(names)})'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Anbar qalığı'

    title_font = Font(bold=True, size=14, color='1F2937')
    subtitle_font = Font(size=10, italic=True, color='6B7280')
    header_font = Font(bold=True, size=11, color='1F2937')

    # Şəkildəki kimi hər sütun üçün fərqli rəng
    header_fills = {
        1: PatternFill('solid', start_color='E2E8F0'),  # #
        2: PatternFill('solid', start_color='E2E8F0'),  # Malın adı
        3: PatternFill('solid', start_color='FBE54D'),  # Əvvələ qalıq - sarı
        4: PatternFill('solid', start_color='86EFC0'),  # Gələn - yaşıl
        5: PatternFill('solid', start_color='F8B4B4'),  # Çıxan - qırmızı
        6: PatternFill('solid', start_color='A9C7F5'),  # Anbarda qalıq - mavi
        7: PatternFill('solid', start_color='D3C4F0'),  # Vəziyyət - bənövşəyi
    }

    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A1'] = f'Anbar qalığı — {date_from.strftime("%d.%m.%Y")} – {date_to.strftime("%d.%m.%Y")}'
    ws['A1'].font = title_font
    ws['A2'] = f'Aptek: {aptek_label}'
    ws['A2'].font = subtitle_font
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')

    headers = ['#', 'Malın adı', 'Əvvələ qalıq', 'Gələn', 'Çıxan', 'Anbarda qalıq', 'Vəziyyət']
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fills[col]
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')
    ws.row_dimensions[4].height = 22

    normal_fill = PatternFill('solid', start_color='DCFCE7')
    normal_font = Font(color='15803D', bold=True, size=10)
    low_fill = PatternFill('solid', start_color='FEE2E2')
    low_font = Font(color='B91C1C', bold=True, size=10)

    row_idx = 5
    for i, row in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=row['ad'])
        for col, key in enumerate(['evvel', 'gelen', 'cixan', 'qalan'], start=3):
            qty_val, qty_fmt = _excel_qty(row[key])
            qty_cell = ws.cell(row=row_idx, column=col, value=qty_val)
            qty_cell.number_format = qty_fmt

        status_cell = ws.cell(row=row_idx, column=7, value='Azalır' if row['low'] else 'Normal')
        if row['low']:
            status_cell.fill = low_fill
            status_cell.font = low_font
        else:
            status_cell.fill = normal_fill
            status_cell.font = normal_font
        status_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
            elif col == 2:
                cell.alignment = Alignment(horizontal='left')
            elif 3 <= col <= 6:
                cell.alignment = Alignment(horizontal='right')

        # Sətir zolaqlı (zebra) fon — oxunaqlılıq üçün
        if i % 2 == 0:
            for col in (1, 2, 3, 4, 5, 6):
                ws.cell(row=row_idx, column=col).fill = PatternFill('solid', start_color='F9FAFB')

        row_idx += 1

    ws.cell(row=row_idx, column=2, value='Cəmi').font = Font(bold=True)
    for col, key in enumerate(['evvel', 'gelen', 'cixan', 'qalan'], start=3):
        qty_val, qty_fmt = _excel_qty(totals[key])
        cell = ws.cell(row=row_idx, column=col, value=qty_val)
        cell.font = Font(bold=True)
        cell.number_format = qty_fmt
        cell.alignment = Alignment(horizontal='right')
        cell.border = Border(top=Side(style='thin', color='9CA3AF'))

    widths = {'A': 6, 'B': 28, 'C': 14, 'D': 10, 'E': 10, 'F': 14, 'G': 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:G{row_idx - 1}'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'anbar_qaligi_{date_from.isoformat()}_{date_to.isoformat()}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def aptek_list(request):
    depo = _get_active_depo(request)
    if request.method == 'POST':
        if request.FILES.get('pdf_file'):
            pdf_file = request.FILES['pdf_file']
            try:
                result = import_qaime_pdf(pdf_file, depo)
                messages.success(request, result['message'])
                if result['missing_drugs']:
                    messages.warning(
                        request,
                        'Tapılmayan dərmanlar: ' + ', '.join(result['missing_drugs']),
                    )
                doc_date = result['doc_date']
                month_start = date(doc_date.year, doc_date.month, 1)
                _, month_last = monthrange(doc_date.year, doc_date.month)
                month_end = date(doc_date.year, doc_date.month, month_last)
                params = urlencode({
                    'from': month_start.isoformat(),
                    'to': month_end.isoformat(),
                    'aptek': result['aptek_id'],
                })
                return redirect(f"{reverse('aptek:anbar_dashboard')}?{params}")
            except QaimeParseError as exc:
                messages.error(request, str(exc))
            except Exception:
                messages.error(request, 'PDF emal edilərkən xəta baş verdi.')
            return redirect('aptek:anbar_dashboard')

        if request.POST.get('manual_qaime') == '1':
            aptek_id = request.POST.get('manual_aptek') or ''
            doc_date_raw = request.POST.get('manual_doc_date') or ''
            doc_date = _parse_date(doc_date_raw, None)
            aptek = (
                Aptek.objects.filter(pk=aptek_id, depo=depo).first() if aptek_id else None
            )

            if not aptek:
                messages.error(request, 'Aptek seçilməyib.')
                return redirect('aptek:anbar_dashboard')
            if not doc_date:
                messages.error(request, 'Tarix düzgün deyil.')
                return redirect('aptek:anbar_dashboard')

            drug_ids = request.POST.getlist('manual_drug_id')
            qty_values = request.POST.getlist('manual_qty')
            items = []
            for drug_id, qty_raw in zip(drug_ids, qty_values):
                if not drug_id:
                    continue
                qty_txt = (qty_raw or '').strip().replace(',', '.')
                if not qty_txt:
                    continue
                try:
                    qty = Decimal(qty_txt)
                except Exception:
                    continue
                if qty <= 0:
                    continue
                drug = Medical.objects.filter(pk=drug_id, status=True).first()
                if not drug:
                    continue
                items.append((drug, qty))

            if not items:
                messages.error(request, 'Ən azı bir dərman və miqdar daxil edin.')
                return redirect('aptek:anbar_dashboard')

            with transaction.atomic():
                qaime_number = (
                    Qaime.objects.filter(
                        depo=depo, aptek=aptek, document_type=Qaime.DOC_QAIME
                    )
                    .order_by('-number')
                    .values_list('number', flat=True)
                    .first() or 0
                ) + 1

                qaime = Qaime.objects.create(
                    depo=depo,
                    aptek=aptek,
                    number=qaime_number,
                    document_type=Qaime.DOC_QAIME,
                    doc_date=doc_date,
                    total=Decimal('0'),
                )

                total_qty = Decimal('0')
                for drug, qty in items:
                    AnbarHereket.objects.create(
                        depo=depo,
                        drug=drug,
                        movement_type=AnbarHereket.MOVEMENT_OUT,
                        quantity=qty,
                        date=doc_date,
                        aptek=aptek,
                        qaime=qaime,
                        note=f'Qaimə №{qaime_number}',
                    )
                    total_qty += qty

                qaime.total = total_qty
                qaime.save(update_fields=['total'])

            messages.success(
                request,
                f'Manuel qaimə №{qaime_number} əlavə olundu ({len(items)} dərman).',
            )
            month_start = date(doc_date.year, doc_date.month, 1)
            _, month_last = monthrange(doc_date.year, doc_date.month)
            month_end = date(doc_date.year, doc_date.month, month_last)
            params = urlencode({
                'from': month_start.isoformat(),
                'to': month_end.isoformat(),
                'aptek': aptek.id,
            })
            return redirect(f"{reverse('aptek:anbar_dashboard')}?{params}")

    date_from, date_to, aptek_id, status_filter, exclude_ids = _ledger_filters(request)
    today = timezone.localdate()
    rows, totals = _build_ledger(
        date_from, date_to, aptek_id, status_filter, exclude_ids, depo=depo
    )
    aptekler = Aptek.objects.filter(depo=depo)
    for aptek in aptekler:
        clean_name = _clean_aptek_name(aptek.name)
        if aptek.name != clean_name:
            aptek.name = clean_name
            aptek.save(update_fields=['name'])
    last_qaime_qs = Qaime.objects.filter(depo=depo, document_type=Qaime.DOC_QAIME)
    if aptek_id:
        last_qaime_qs = last_qaime_qs.filter(aptek_id=aptek_id)
    if exclude_ids:
        last_qaime_qs = last_qaime_qs.exclude(aptek_id__in=exclude_ids)
    last_qaime = last_qaime_qs.select_related('aptek').order_by('-doc_date', '-id').first()

    month_label = AZ_MONTHS.get(date_from.month, '')
    if date_from.month != date_to.month or date_from.year != date_to.year:
        month_label = (
            f'{date_from.strftime("%d.%m.%Y")} — {date_to.strftime("%d.%m.%Y")}'
        )

    excluded_apteks = list(
        Aptek.objects.filter(depo=depo, id__in=exclude_ids).order_by('name')
    )
    user_ctx = _user_context(request)

    context = {
        'rows': rows,
        'totals': totals,
        'aptekler': aptekler,
        'selected_aptek': str(aptek_id) if aptek_id else '',
        'exclude_ids': [str(x) for x in exclude_ids],
        'excluded_apteks': excluded_apteks,
        'exclude_query': _exclude_query(exclude_ids),
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'status_filter': status_filter or 'all',
        'month_label': month_label,
        'record_count': len(rows),
        'last_qaime': last_qaime,
        'manual_drugs': Medical.objects.filter(status=True).order_by('position', 'med_name'),
        'manual_default_date': today.isoformat(),
        **user_ctx,
    }
    return render(request, 'anbar_dashboard.html', context)
