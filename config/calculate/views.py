from django.shortcuts import render
from .models import  Calculate, Medical, Report
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout 
from django.contrib import messages
from vizit.models import Istifadeci
from django.contrib.auth import login as django_login, logout as django_logout   # Bura əlavə edildi
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from decimal import Decimal, InvalidOperation
import json
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



def login_user(request):
    if request.session.get('current_user_data'):
        return redirect('/groups/calculate/')

    if request.method == "POST":
        name_input = request.POST.get("name", "").strip()
        password_input = request.POST.get("password", "").strip()
        
        user = Istifadeci.objects.filter(
            login=name_input, 
            sifre=Istifadeci.hash_sifre(password_input), 
            aktiv=True
        ).first()
        
        if user is not None:
            # Əvvəlcə Django auth user ilə login ol
            django_auth_user, _ = User.objects.get_or_create(
                username=user.login,
                defaults={'first_name': user.ad, 'is_active': True}
            )
            django_auth_user.backend = 'django.contrib.auth.backends.ModelBackend'
            django_login(request, django_auth_user)
            
            # Django_login session flush etdikdən SONRA yaz
            request.session['current_user_data'] = user.session_dict()
            request.session.modified = True
            
            return redirect('/groups/calculate/')
        else:
            messages.error(request, "İstifadəçi adı və ya şifrə səhvdir!")
            return redirect('/groups/login/')
            
    # GET request — sistem user ilə middleware-i keç
    if not request.user.is_authenticated:
        system_user, _ = User.objects.get_or_create(username="system_groups_user")
        system_user.backend = 'django.contrib.auth.backends.ModelBackend'
        django_login(request, system_user)
            
    return render(request, "calculate/login.html")


def logout_user(request):
    request.session.flush()
    return redirect('/groups/login/')

# Qiymət siyahısı (Konstantlar funksiyalardan kənarda saxlanılır)
PRICES_NUMAYENDE_QRUP_1 = {
    "LEVOSTRONG": 0.6, "LIPOMAG+": 3.6, "SOLSEDA": 1.5, "SOLTEP": 0.0,
    "ZEMOVAR": 2.3, "KALVEY": 0.0, "PAINSTOP": 1.5, "BETASOL": 2.1,
    "LITASOL": 1.7, "FENSAVIN": 1.6,
}

PRICES_NUMAYENDE_QRUP_2 = {
    "PROSTAZOLIN": 0.6, "HEPTRAZOL": 1.6, "OPEBLOCK": 2.4, "OPSIDOL": 0.0,
    "SERRASOL": 3.0, "GENOSFER": 1.6, "VITOMER": 1.2, "KARTOVEY": 0.0,
    "SOLTROP": 2.7, "ROPSOL": 1.4, "MOXIVISTA": 0.6,
}

# Menecer qiymətləri şəkildəki qruplara əsasən ikiyə bölündü
PRICES_MENECER_QRUP_1 = {
    "LEVOSTRONG": 0.4,
    "LIPOMAG+": 3.0,     # "LIPOMAQ" -> "LIPOMAG" olaraq düzəldildi
    "SOLSEDA": 1.2,
    "SOLTEP": 0.0,
    "ZEMOVAR": 1.7,
    "KALVEY": 0.0,
    "PAINSTOP": 1.0,
    "BETASOL": 1.4,
    "LITASOL": 1.3,
    "FENSAVIN": 1.2,
}

PRICES_MENECER_QRUP_2 = {
    "PROSTAZOLIN": 0.4,
    "HEPTRAZOL": 1.4,
    "OPEBLOCK": 1.6,
    "OPSIDOL": 0.0,
    "SERRASOL": 2.0,
    "GENOSFER": 1.2,
    "VITOMER": 0.8,     # "Vitomer" adının qorunması təmin edilir
    "KARTOVEY": 0.0,
    "SOLTROP": 1.8,
    "ROPSOL": 0.9,
    "MOXIVISTA": 0.4,
}

PRICES_REHBER = {
    **PRICES_MENECER_QRUP_1,   # QRUP 1
    **PRICES_MENECER_QRUP_2    # QRUP 2
}

# =========================================================================

def hesablamalar(request):
    user_data = request.session.get('current_user_data')
    if not user_data:
        return redirect('/groups/login/')

    # ==================== POST - YADDA SAXLAMA ====================
    if request.method == "POST" and request.POST.get('save_report'):
        items_raw = request.POST.get('items', '[]')
        total_amount_raw = request.POST.get('total_amount', '0')

        # JSON string-i Python obyektinə çevir
        try:
            items_parsed = json.loads(items_raw) if items_raw else []
        except json.JSONDecodeError:
            items_parsed = []

        try:
            total_amount = Decimal(total_amount_raw)
        except (InvalidOperation, TypeError):
            total_amount = Decimal('0')

        if not items_parsed:
            messages.error(request, 'Heç bir dərman seçilməyib, hesabat saxlanılmadı.')
            return redirect('/groups/calculate/')

        try:
            Report.objects.create(
                user_id=user_data.get('id'),   # FK-yə pk veriləndə user_id istifadə olunur
                user_name=user_data.get('ad'),
                user_group=user_data.get('qrup'),
                user_role=user_data.get('rol'),
                total_amount=total_amount,
                items=items_parsed,   # artıq real list/dict, JSONField düzgün saxlayacaq
            )
            messages.success(request, 'Hesabat uğurla yadda saxlanıldı!')
            return redirect('/groups/calculate/')
        except Exception as e:
            messages.error(request, f'Xəta baş verdi: {e}')

    # ==================== GET - Səhifəni göstər ====================
    user_role = user_data.get('rol')   
    user_qrup = user_data.get('qrup')  

    if user_role in ['rehber', 'diviziya_rehb', 'admin']:
        active_prices = PRICES_REHBER
    elif user_role in ['menecer']:
        active_prices = PRICES_MENECER_QRUP_2 if user_qrup == 'QRUP 2' else PRICES_MENECER_QRUP_1
    else:
        active_prices = PRICES_NUMAYENDE_QRUP_2 if user_qrup == 'QRUP 2' else PRICES_NUMAYENDE_QRUP_1

    # Display names...
    display_names = {
        "LEVOSTRONG": "Levostrong", 
        "LIPOMAG": "Lipomag", 
        "SOLSEDA": "Solseda", 
        "SOLTEP": "Soltep", 
        "ZEMOVAR": "Zemovar", 
        "KALVEY": "Kalvey", 
        "PAINSTOP": "Painstop", 
        "BETASOL": "Betasol", 
        "LITASOL": "Litasol", 
        "FENSAVIN": "Fensavin", 
        "PROSTAZOLIN": "Prostazolin", 
        "HEPTRAZOL": "Heptrazol", 
        "OPEBLOCK": "Opeblock", 
        "OPSIDOL": "Opsidol", 
        "SERRASOL": "Serrasol", 
        "GENOSFER": "Genosfer", 
        "VITOMER": "Vitomer D3", 
        "KARTOVEY": "Kartovey", 
        "SOLTROP": "Soltrop", 
        "ROPSOL": "Ropsol", 
        "MOXIVISTA": "Moxivista"
    }  # əvvəlki kimi saxla

    filtered_items = []
    for med_key, price in active_prices.items():
        name_display = display_names.get(med_key, med_key.title())
        filtered_items.append({"med_name": name_display, "azn": price})

    context = {
        "medicals": filtered_items,
        "current_user": {
            "name": user_data.get("ad"),
            "role": user_role, 
            "group": user_qrup
        }
    }
    return render(request, "calculate/calculate.html", context)


def admin_page(request):
    # 1. TƏHLÜKƏSİZLİK: Giriş edilməyibsə dərhal /groups-drugs/login/ səhifəsinə göndər
    user_data = request.session.get('current_user_data')
    if not user_data:
        return redirect('/groups/login/')
    
    # 2. SƏLAHİYYƏT: Yalnız 'admin' və ya 'rehber' bu səhifəyə girə bilsin
    if user_data.get('rol') not in ['admin', 'rehber']:
        messages.error(request, "Bu səhifəyə giriş icazəniz yoxdur!")
        return redirect('/groups/calculate/')

    if request.method == "POST":
        login_input = request.POST.get("name")      
        password_input = request.POST.get("password")
        ad_input = request.POST.get("ad") or login_input 
        role = request.POST.get("role")
        group = request.POST.get("group") or None

        if login_input and password_input and role:
            if Istifadeci.objects.filter(login=login_input).exists():
                messages.error(request, f"'{login_input}' istifadəçi adı ilə artıq kimsə qeydiyyatdan keçib!")
            else:
                yeni_user = Istifadeci(
                    login=login_input,
                    ad=ad_input,
                    rol=role,
                    qrup=group,
                    aktiv=True
                )
                yeni_user.set_password(password_input)
                yeni_user.save()
                
                messages.success(request, "Yeni istifadəçi uğurla əlavə edildi.")
                return redirect('/groups/admin/') 
        else:
            messages.error(request, "Zəhmət olmasa tələb olunan bütün xanaları doldurun.")

    all_users = Istifadeci.objects.all().order_by('-id')
    
    context = {
        "users": all_users,
        "roles": Istifadeci.ROL_CHOICES,    
        "groups_list": Istifadeci.QRUP_CHOICES, 
        "current_user": user_data
    }
    return render(request, "calculate/admin_page.html", context)




# views.py
def reports_list(request):
    user_data = request.session.get('current_user_data')
    if not user_data:
        return redirect('/groups/login/')

    reports = Report.objects.filter(user=user_data.get('id')).order_by('-created_at')
    
    context = {
        'reports': reports,
        'current_user': user_data
    }
    return render(request, 'calculate/report.html', context)


def report_detail_json(request, report_id):
    report = get_object_or_404(Report, id=report_id)
    return JsonResponse({
        'id': report.id,
        'created_at': report.created_at.strftime('%d.%m.%Y %H:%M'),
        'user_name': report.user_name,
        'user_group': report.user_group or '—',
        'user_role': report.user_role or '—',
        'total_amount': float(report.total_amount),
        'note': report.note or '',
        'items': report.items,  # artıq JSON formatındadır
    })  



def report_export_excel(request, report_id):
    report = get_object_or_404(Report, id=report_id)

    items = report.items
    if isinstance(items, str):
        try:
            items = json.loads(items) if items else []
        except json.JSONDecodeError:
            items = []
    if not isinstance(items, list):
        items = []

    wb = Workbook()
    ws = wb.active
    ws.title = f"Hesabat {report.id}"

    # ---------- Stil tərifləri ----------
    header_font = Font(name='Arial', bold=True, size=16, color='1E2937')
    label_font = Font(name='Arial', bold=True, size=9, color='94A3B8')
    value_font = Font(name='Arial', bold=True, size=11, color='334155')
    table_header_font = Font(name='Arial', bold=True, size=10, color='64748B')
    table_header_fill = PatternFill('solid', start_color='F1F5F9')
    total_label_font = Font(name='Arial', bold=True, size=12, color='64748B')
    total_value_font = Font(name='Arial', bold=True, size=14, color='0F766E')
    thin_border = Border(bottom=Side(style='thin', color='E2E8F0'))

    # ---------- Başlıq ----------
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Hesabat #{report.id}"
    ws['A1'].font = header_font
    ws.row_dimensions[1].height = 28

    # ---------- Info sahələri ----------
    ws['A3'] = 'TARİX'
    ws['B3'] = 'İSTİFADƏÇİ'
    ws['C3'] = 'QRUP'
    ws['D3'] = 'ROL'
    for col in ['A3', 'B3', 'C3', 'D3']:
        ws[col].font = label_font

    ws['A4'] = report.created_at.strftime('%d.%m.%Y %H:%M')
    ws['B4'] = report.user_name or '—'
    ws['C4'] = report.user_group or '—'
    ws['D4'] = report.user_role or '—'
    for col in ['A4', 'B4', 'C4', 'D4']:
        ws[col].font = value_font

    # ---------- Dərman cədvəli ----------
    table_start_row = 6
    headers = ['Dərman', 'Miqdar', 'Qiymət (₼)', 'Cəmi (₼)']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=table_start_row, column=i, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

    row_idx = table_start_row + 1
    for item in items:
        name = item.get('name', '—')
        qty = item.get('qty', 0)
        price = item.get('price', 0)
        total = item.get('total', 0)

        ws.cell(row=row_idx, column=1, value=name).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=2, value=qty).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=3, value=float(price)).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=3).number_format = '0.00'
        ws.cell(row=row_idx, column=4, value=float(total)).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=4).number_format = '0.00'

        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = thin_border

        row_idx += 1

    if not items:
        ws.merge_cells(f'A{row_idx}:D{row_idx}')
        ws.cell(row=row_idx, column=1, value='Bu hesabatda dərman məlumatı yoxdur.').font = Font(
            name='Arial', italic=True, size=10, color='94A3B8'
        )
        row_idx += 1

    # ---------- Ümumi məbləğ ----------
    total_row = row_idx + 1
    ws.cell(row=total_row, column=3, value='Ümumi məbləğ:').font = total_label_font
    total_cell = ws.cell(row=total_row, column=4, value=float(report.total_amount))
    total_cell.font = total_value_font
    total_cell.number_format = '0.00'

    # ---------- Sütun enləri ----------
    widths = {'A': 28, 'B': 12, 'C': 14, 'D': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---------- Faylı HTTP cavabı kimi qaytar ----------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="hesabat_{report.id}.xlsx"'
    return response