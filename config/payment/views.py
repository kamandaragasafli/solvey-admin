from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.dateparse import parse_date
from datetime import date
from decimal import Decimal as D
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal
from payment.models import Payment_doctor, Sale, MonthlyDoctorReport, Financial_document, DepoSale
from regions.models import Region
from doctors.models import Doctors, Recipe, RecipeDrug
from medicine.models import Medical
from django.http import JsonResponse
from django.db import transaction
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Sum
from django.core.serializers.json import DjangoJSONEncoder
import json
from collections import defaultdict
from django.db.models.functions import ExtractMonth
from datetime import datetime
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal as d
from collections import defaultdict
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
import openpyxl



def get_doctors_by_region(request):
    region_id = request.GET.get("region_id")
    if region_id:
        doctors = Doctors.objects.filter(bolge_id=region_id).order_by("id")
        doctor_list = [{"id": d.id, "ad": d.ad} for d in doctors]
        return JsonResponse({"doctors": doctor_list})
    return JsonResponse({"doctors": []})


def get_doctors(request):
    region_id = request.GET.get('region_id')
    search = request.GET.get('search', '').strip()
    
    # 1. Bölgə yoxdursa, boş qaytar
    if not region_id:
        return JsonResponse({"doctors": []})
        
    # 2. Həkimləri filtrələ
# Sərt filtrasiya üçün:
    doctors = Doctors.objects.filter(bolge_id=region_id, ad__icontains=search)
    
    # 3. Axtarış varsa, yalnız onu tətbiq et
    if search:
        # 'ad' sahəsinin modelinizdə dəqiq adı nədirsə onu yazın (ad, name, fullname)
        doctors = doctors.filter(ad__icontains=search) 
    
    # 4. JSON formatına çevir
    data = list(doctors.values('id', 'ad'))
    return JsonResponse({"doctors": data})

def add_pay_dr(request, region_id=None):
    region = Region.objects.all().order_by('id')
    doctors = Doctors.objects.none()

    if region_id:
        doctors = Doctors.objects.filter(bolge_id=region_id).prefetch_related('odenisler')
        for d in doctors:
            last = d.odenisler.order_by('-date').first()
            if last:
                d.son_odenis_mebleg = last.pay
                d.son_odenis_tarixi = last.date
                d.son_odenis_novu = last.payment_type

    if request.method == "POST":
        region_id = request.POST.get("region_id")
        doctor_id = request.POST.get("doctor_id")
        payment_type = request.POST.get("payment_type")
        amount = request.POST.get("amount")
        pay_date = request.POST.get("pay_date")

        if not all([region_id, doctor_id, payment_type, amount, pay_date]):
            messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
            return redirect("add-pay-dr-region", region_id=region_id) if region_id else redirect("add-pay-dr")

        try:
            borc_miqdari = Decimal(amount)
        except:
            messages.error(request, "Ödəniş məbləği düzgün formatda deyil.")
            return redirect("add-pay-dr-region", region_id=region_id)

        try:
            Payment_doctor.objects.create(
                area_id=region_id,
                doctor_id=doctor_id,
                payment_type=payment_type,
                pay=borc_miqdari,
                date=pay_date
            )
            messages.success(request, "Ödəniş uğurla əlavə edildi.")
            return redirect("doctor_detail", doctor_id=doctor_id)
        except Exception as e:
            messages.error(request, f"Xəta baş verdi: {e}")
            return redirect("add-pay-dr-region", region_id=region_id)

    context = {
        "region": region,
        "doctors": doctors,
        "selected_region_id": int(region_id) if region_id and str(region_id).isdigit() else None,
    }
    return render(request, "crud/addpay_dr.html", context)



def document_add(request):
    region = Region.objects.all().order_by('id')  # Bölgələri template'a göndərin
    
    if request.method == 'POST':
        try:
            region_id = request.POST.get('region_id')
            doctor_id = request.POST.get('doctor_id')
            check_photo = request.FILES.get('check_photo')
            check_date = request.POST.get('check_date')
            
            if not all([region_id, doctor_id, check_photo, check_date]):
                messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
                return redirect('document_add')
            
            # Fayl ölçüsünü yoxla
            if check_photo.size > 5 * 1024 * 1024:  # 5MB
                messages.error(request, "Fayl ölçüsü 5MB-dan çox ola bilməz")
                return redirect('document_add')
            
            # Sənədi yadda saxla
            document = Financial_document(
                check_photo=check_photo,
                check_dr_id=doctor_id,
                check_region_id=region_id,
                check_date=check_date
            )
            document.save()
            
            messages.success(request, 'Sənəd uğurla əlavə edildi!')
            return redirect('document_add')
            
        except Exception as e:
            messages.error(request, f'Xəta baş verdi: {str(e)}')
            return redirect('document_add')
    
    context = {
        'region': region,
    }
    return render(request, 'crud/add_document.html', context)


def financial_documents(request):
    documents = Financial_document.objects.select_related(
        'check_dr', 'check_region'
    ).order_by('-check_date')
    
    # Filter by region if provided
    region_id = request.GET.get('region')
    if region_id:
        documents = documents.filter(check_region_id=region_id)
    
    # Filter by doctor name if provided
    doctor_name = request.GET.get('doctor')
    if doctor_name:
        documents = documents.filter(check_dr__ad__icontains=doctor_name)
    
    # Statistics
    total_documents = documents.count()
    this_month = timezone.now().month
    this_month_count = documents.filter(
        check_date__month=this_month
    ).count()
    
    regions = Region.objects.all()
    
    context = {
        'documents': documents,
        'total_documents': total_documents,
        'this_month_count': this_month_count,
        'regions': regions,
    }
    
    return render(request, 'finance-document.html', context)

def create_sale(request):
    drug_all = Medical.objects.active().order_by('id')
    region_all = Region.objects.all().order_by('id')

    if request.method == "POST":
        region_id = request.POST.get("region")
        date_str = request.POST.get("date")

        if not region_id or not date_str:
            messages.error(request, "Bölgə və tarixi seçməlisiniz!")
            return redirect('index')

        # Tarixi date tipinə çeviririk
        try:
            sale_date = parse_date(date_str)
            if sale_date is None:
                raise ValueError
        except ValueError:
            messages.error(request, "Tarix düzgün deyil!")
            return redirect('index')

        region = Region.objects.get(id=region_id)
        sales_created = False
        errors = []

        # Bu ayın ilk günü (aylığa görə yoxlamaq üçün)
        month_start = sale_date.replace(day=1)

        for key, value in request.POST.items():
            if key.startswith('quantity_') and value.isdigit() and int(value) > 0:
                drug_id = key.split("_")[1]
                quantity = int(value)
                drug = Medical.objects.get(id=drug_id)

                # Eyni ay, bölgə, dərman üçün artıq satış var?
                exists = Sale.objects.filter(
                    region=region,
                    drug=drug,
                    sale_date__year=month_start.year,
                    sale_date__month=month_start.month
                ).exists()

                if exists:
                    errors.append(f"{drug.med_name} dərmanı üçün bu ay artıq satış əlavə olunub.")
                else:
                    Sale.objects.create(
                        region=region,
                        drug=drug,
                        quantity=quantity,
                        sale_date=sale_date
                    )
                    sales_created = True

        if sales_created:
            from doctors.views import recalc_region_report_for_date
            from django.utils.html import format_html
            recalc_region_report_for_date(region.id, sale_date)
            report_url = (
                reverse("region_report", args=[region.id])
                + f"?month={sale_date.month}&year={sale_date.year}"
            )
            messages.success(
                request,
                format_html(
                    'Satışlar uğurla əlavə olundu. Komissiya hesablandı. '
                    '<a href="{}" style="color:#93c5fd;text-decoration:underline;">'
                    'Cədvəllərə bax ({} {})</a>',
                    report_url,
                    sale_date.month,
                    sale_date.year,
                ),
            )

        if errors:
            for error in errors:
                messages.warning(request, error)

        if not sales_created and not errors:
            messages.warning(request, 'Heç bir dərman seçilməyib')

        return redirect('index')

    context = {
        "drug_all": drug_all,
        "region_all": region_all,
    }
    return render(request, 'crud/add-sales.html', context)


def update_sale(request):
    """
    Region üzrə aylıq satışları redaktə et (mövcud miqdarları yüklə, upsert et).
    """
    drug_all = Medical.objects.all().order_by("id")
    region_all = Region.objects.all().order_by("id")

    selected_region_id = request.GET.get("region") or request.POST.get("region") or ""
    date_str = request.GET.get("date") or request.POST.get("date") or date.today().strftime("%Y-%m-%d")
    default_date = date.today().strftime("%Y-%m-%d")
    existing_qty = {}

    # Handle POST (save/upsert)
    if request.method == "POST":
        if not selected_region_id or not date_str:
            messages.error(request, "Bölgə və tarixi seçməlisiniz!")
            return redirect("update-sell")

        try:
            sale_date = parse_date(date_str)
            if sale_date is None:
                raise ValueError
        except ValueError:
            messages.error(request, "Tarix düzgün deyil!")
            return redirect("update-sell")

        try:
            region = Region.objects.get(id=selected_region_id)
        except Region.DoesNotExist:
            messages.error(request, "Seçilmiş bölgə tapılmadı.")
            return redirect("update-sell")

        month_start = sale_date.replace(day=1)

        for drug in drug_all:
            key = f"quantity_{drug.id}"
            val = request.POST.get(key)
            if val is None:
                continue
            try:
                qty = int(val) if val != "" else 0
            except ValueError:
                qty = 0

            qs = Sale.objects.filter(
                region=region,
                drug=drug,
                sale_date__year=month_start.year,
                sale_date__month=month_start.month,
            )
            if qty > 0:
                if qs.exists():
                    sale_obj = qs.first()
                    sale_obj.quantity = qty
                    sale_obj.sale_date = sale_date
                    sale_obj.save()
                else:
                    Sale.objects.create(
                        region=region,
                        drug=drug,
                        quantity=qty,
                        sale_date=sale_date,
                    )
            else:
                qs.delete()  # qty == 0 → sil

        messages.success(request, "Satış məlumatları yeniləndi.")
        from doctors.views import recalc_region_report_for_date
        from django.utils.html import format_html
        recalc_region_report_for_date(region.id, sale_date)
        report_url = (
            reverse("region_report", args=[region.id])
            + f"?month={month_start.month}&year={month_start.year}"
        )
        messages.info(
            request,
            format_html(
                'Komissiya hesablandı. '
                '<a href="{}" style="color:#93c5fd;text-decoration:underline;">'
                'Cədvəllərə bax</a>',
                report_url,
            ),
        )
        return redirect(
            f"{reverse('sales')}?month={month_start.month}&year={month_start.year}"
            + (f"&region={region.id}" if selected_region_id else "")
        )

    # Handle GET (load existing)
    try:
        sel_date = parse_date(date_str)
    except Exception:
        sel_date = date.today()

    if selected_region_id and sel_date:
        existing = Sale.objects.filter(
            region_id=selected_region_id,
            sale_date__year=sel_date.year,
            sale_date__month=sel_date.month,
        )
        existing_qty = {s.drug_id: s.quantity for s in existing}

    context = {
        "drug_all": drug_all,
        "region_all": region_all,
        "selected_region_id": int(selected_region_id) if selected_region_id else "",
        "selected_date": sel_date.strftime("%Y-%m-%d") if sel_date else default_date,
        "default_date": default_date,
        "existing_qty": existing_qty,
    }
    return render(request, "crud/update-sales.html", context)

def sales(request):
    """
    Region × Drug sales matrix with filters.
    Filters: region_search (name contains), region (id), month, year.
    İlk açılışda cari ay/il default; boş seçim (Hamısı) saxlanılır.
    """
    today = date.today()
    all_region = Region.objects.all().order_by("region_name")
    all_drug = Medical.objects.active().order_by("position", "id")

    years = list(Sale.objects.dates("sale_date", "year").distinct())
    years = sorted([y.year for y in years], reverse=True)
    current_year = today.year
    if current_year not in years:
        years.insert(0, current_year)
    if 2026 not in years:
        years.insert(0, 2026)
    years = sorted(set(years), reverse=True)

    region_search = request.GET.get("region_search", "").strip()
    region_id = request.GET.get("region", "").strip()
    # İlk açılışda cari ay/il URL-ə yazılsın (Excel və bookmark üçün)
    if "month" not in request.GET and "year" not in request.GET:
        q = request.GET.copy()
        q["month"] = str(today.month)
        q["year"] = str(today.year)
        return redirect(f"{request.path}?{q.urlencode()}")

    month = (request.GET.get("month") or "").strip()
    year = (request.GET.get("year") or "").strip()

    sales_queryset = Sale.objects.all()

    if region_search:
        all_region = all_region.filter(region_name__icontains=region_search)

    if region_id:
        try:
            region_id_int = int(region_id)
            all_region = all_region.filter(id=region_id_int)
            sales_queryset = sales_queryset.filter(region_id=region_id_int)
        except ValueError:
            region_id_int = None
    else:
        region_id_int = None

    month_int = None
    if month:
        try:
            month_int = int(month)
            sales_queryset = sales_queryset.filter(sale_date__month=month_int)
        except ValueError:
            month_int = None

    year_int = None
    if year:
        try:
            year_int = int(year)
            sales_queryset = sales_queryset.filter(sale_date__year=year_int)
        except ValueError:
            year_int = None

    # Bir sorğu ilə region×drug cəmləri
    sales_dict = {r.id: {d.id: 0 for d in all_drug} for r in all_region}
    totals_per_region = {r.id: 0 for r in all_region}
    totals_per_drug = {d.id: 0 for d in all_drug}
    grand_total = 0

    region_ids = list(all_region.values_list("id", flat=True))
    agg = (
        sales_queryset.filter(region_id__in=region_ids)
        .values("region_id", "drug_id")
        .annotate(total=Sum("quantity"))
    )
    for row in agg:
        rid, did = row["region_id"], row["drug_id"]
        qty = row["total"] or 0
        if rid in sales_dict and did in sales_dict[rid]:
            sales_dict[rid][did] = qty
            totals_per_region[rid] += qty
            totals_per_drug[did] += qty
            grand_total += qty

    month_names = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
        5: "May", 6: "İyun", 7: "İyul", 8: "Avqust",
        9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr",
    }
    period_label = "Bütün dövrlər"
    if month_int and year_int:
        period_label = f"{month_names.get(month_int, month_int)} {year_int}"
        edit_date = date(year_int, month_int, 1).strftime("%Y-%m-%d")
    elif year_int:
        period_label = str(year_int)
        edit_date = date(year_int, today.month, 1).strftime("%Y-%m-%d")
    elif month_int:
        period_label = month_names.get(month_int, str(month_int))
        edit_date = date(today.year, month_int, 1).strftime("%Y-%m-%d")
    else:
        edit_date = today.strftime("%Y-%m-%d")

    context = {
        "all_region": all_region,
        "all_drug": all_drug,
        "sales_dict": sales_dict,
        "totals_per_region": totals_per_region,
        "totals_per_drug": totals_per_drug,
        "grand_total": grand_total,
        "years": years,
        "selected_region": region_id_int,
        "selected_month": month_int,
        "selected_year": year_int,
        "region_search": region_search,
        "period_label": period_label,
        "edit_date": edit_date,
        "today": today.strftime("%Y-%m-%d"),
    }
    return render(request, "reports/sales.html", context)


def export_sales_excel(request):
    """Aylıq Satışlar səhifəsinin Excel çıxarışı (reports/sales)."""
    today = date.today()
    all_region = Region.objects.all().order_by("region_name")
    all_drug = Medical.objects.active().order_by("position", "id")

    region_search = request.GET.get("region_search", "").strip()
    region_id = request.GET.get("region", "").strip()
    if "month" not in request.GET:
        month = str(today.month)
    else:
        month = (request.GET.get("month") or "").strip()
    if "year" not in request.GET:
        year = str(today.year)
    else:
        year = (request.GET.get("year") or "").strip()

    sales_queryset = Sale.objects.all()

    if region_search:
        all_region = all_region.filter(region_name__icontains=region_search)
    if region_id:
        try:
            rid = int(region_id)
            all_region = all_region.filter(id=rid)
            sales_queryset = sales_queryset.filter(region_id=rid)
        except ValueError:
            pass
    if month:
        try:
            sales_queryset = sales_queryset.filter(sale_date__month=int(month))
        except ValueError:
            pass
    if year:
        try:
            sales_queryset = sales_queryset.filter(sale_date__year=int(year))
        except ValueError:
            pass

    sales_dict = {r.id: {d.id: 0 for d in all_drug} for r in all_region}
    totals_per_region = {r.id: 0 for r in all_region}
    totals_per_drug = {d.id: 0 for d in all_drug}
    grand_total = 0
    region_ids = list(all_region.values_list("id", flat=True))
    agg = (
        sales_queryset.filter(region_id__in=region_ids)
        .values("region_id", "drug_id")
        .annotate(total=Sum("quantity"))
    )
    for row in agg:
        rid, did = row["region_id"], row["drug_id"]
        qty = row["total"] or 0
        if rid in sales_dict and did in sales_dict[rid]:
            sales_dict[rid][did] = qty
            totals_per_region[rid] += qty
            totals_per_drug[did] += qty
            grand_total += qty

    wb = Workbook()
    ws = wb.active
    ws.title = "Aylıq Satışlar"

    headers = ["#", "Bölgə"] + [d.med_name for d in all_drug] + ["Total"]
    ws.append(headers)

    bold_font = Font(bold=True, color="060411")
    header_fill = PatternFill(fill_type="solid", fgColor="E0E0E0")
    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_fill = PatternFill(fill_type="solid", fgColor="C8E6C9")

    for col, cell in enumerate(ws[1], 1):
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, region in enumerate(all_region, start=1):
        row = [i, region.region_name]
        for drug in all_drug:
            row.append(sales_dict.get(region.id, {}).get(drug.id, 0))
        row.append(totals_per_region.get(region.id, 0))
        ws.append(row)

    total_row = ["", "CƏMİ"] + [totals_per_drug[d.id] for d in all_drug] + [grand_total]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = thin_border

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    for i in range(len(all_drug)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12
    ws.column_dimensions[get_column_letter(3 + len(all_drug))].width = 12

    month_names = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "İyun",
        7: "İyul", 8: "Avqust", 9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr",
    }
    period = ""
    if month and year:
        try:
            period = f"{month_names.get(int(month), month)}_{year}"
        except (ValueError, KeyError):
            period = f"{month}_{year}"
    elif year:
        period = str(year)
    else:
        period = datetime.now().strftime("%Y-%m-%d")
    filename = f"Ayliq_Satislar_{period}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
  
def depo_sales(request):
    all_region = Region.objects.all().order_by("id")
    all_drug = Medical.objects.all().order_by("id")

    # ================= 1. 1C EXCEL IMPORT (POST) =================
    if request.method == 'POST' and request.FILES.get('excel_file'):
        selected_depo_raw = request.POST.get('modal_depo')
        selected_month    = request.POST.get('modal_month')
        excel_file        = request.FILES['excel_file']

        depo_mapping = {'avromed': 'Avromed', 'zeytun': 'Zeytun', 'bine': 'Binə'}
        depo_name    = depo_mapping.get(selected_depo_raw, 'Binə')

        drug_translation = {
            'бетасол':      'betasol',
            'витомер д3':   'vitomer d3',
            'витомер d3':   'vitomer d3',
            'витомер kids': 'vitomer kids',
            'витомер':      'vitomer d3',
            'гептразол':    'heptrazol',
            'сольтроп':     'soltrop',
            'soltrop':      'soltrop',
            'солтр':        'soltrop',
            'фенсавин':     'fensavin',
            'сольтеп':      'soltep',
            'литасол':      'litasol',
            'простазолин':  'prostazolin',
            'провитал':     'provital',
            'геносфер':     'genosfer',
            'серрасол':     'serrasol',
            'картовей':     'kartovey',
            'левостронг':   'levostrong',
            'опеблок':      'opeblock',
            'ропсол':       'ropsol',
            'фесola':       'fesola',
            'зимвоар':      'zemovar',
            'земовар':      'zemovar',
            'опсайдол':     'opsidol',
            'солсайдол':    'solseda',
            'солседа':      'solseda',
            'бетакон':      'betacon',
            'моксивиста':   'moksivista',
            'пейнстоп':     'peynstop',
        }

        try:
            wb    = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            current_drug = None
            saved_count  = 0

            with transaction.atomic():
                for row in range(14, sheet.max_row + 1):
                    b_obj  = sheet.cell(row=row, column=2)
                    c_val  = sheet.cell(row=row, column=3).value
                    b_val  = b_obj.value

                    if b_val is None:
                        continue

                    b_str = str(b_val).strip().lower()

                    # ── ADDIM 1: Dərman sətirini aşkar et ──────────────────────────
                    # Kiril dərman adı varsa → bu dərman başlıq sətiridir
                    is_drug_row      = False
                    matched_latin    = None

                    for kiril_key, latin_val in drug_translation.items():
                        if kiril_key in b_str:
                            is_drug_row   = True
                            matched_latin = latin_val
                            break

                    # "флак", "шт", "ампулы" və s. vahid sözləri də dərman sətiridir
                    if not is_drug_row:
                        drug_unit_kw = ["флак", "шт", "ампулы", "капсул",
                                        "капли", "таблет", "мг №", "раствор", "саше", "сироп"]
                        if any(kw in b_str for kw in drug_unit_kw):
                            is_drug_row = True

                    if is_drug_row:
                        drug_obj = None

                        if matched_latin:
                            drug_obj = next(
                                (d for d in all_drug
                                 if d.med_name.lower().strip() == matched_latin),
                                None
                            )
                        # Translation ilə tapılmadısa birbaşa mətndən axtar
                        if not drug_obj:
                            drug_obj = next(
                                (d for d in all_drug
                                 if d.med_name.lower().strip() in b_str),
                                None
                            )

                        if drug_obj:
                            current_drug = drug_obj
                        continue  # dərman sətirini data kimi yazma

                    # ── ADDIM 2: Miqdar olmayan sətirləri atla ──────────────────────
                    if current_drug is None or c_val is None:
                        continue

                    # ── ADDIM 3: Yalnız indent=1 sətirləri → depo/region cəmi ──────
                    #
                    # Excel hierarxiyası:
                    #   indent=0  → "Итог" (ümumi cəm) — skip
                    #   indent=1  → Depo/Region cəmi:
                    #               "Bakı", "Abşeron", "Bölgə (Şəki Depo)",
                    #               "Şamaxı", "Sumqayıt", "Xaçmaz" ...
                    #   indent=2  → Alt-region (Şəki, Gəncə, Qəbələ ...) — skip
                    #   indent=3  → Alt müştəri (VIP class ...) — skip
                    #
                    # indent=1 olan sətir = bizim bazaya yazacağımız region cəmidir.

                    try:
                        indent = b_obj.alignment.indent if b_obj.alignment else 0
                        indent = indent if indent is not None else 0
                    except Exception:
                        indent = 0

                    if indent != 1:
                        continue  # alt-region, alt-müştəri, yekun → hamısı skip

                    # ── ADDIM 4: Region adını tap ───────────────────────────────────
                    matched_region = None
                    for r in all_region:
                        r_name = r.region_name.lower().strip()
                        if r_name in b_str:
                            matched_region = r
                            break

                    if not matched_region:
                        continue  # naməlum region → skip

                    # ── ADDIM 5: Miqdarı təmizlə və bazaya yaz ─────────────────────
                    try:
                        clean_qty = (
                            str(c_val)
                            .replace(" ", "")
                            .replace("\xa0", "")
                            .replace(",", ".")
                            .strip()
                        )
                        qty = float(clean_qty)

                        DepoSale.objects.create(
                            depo_type = depo_name,
                            region    = matched_region,
                            drug      = current_drug,
                            quantity  = qty,
                            sale_date = f"2026-{str(selected_month).zfill(2)}-01"
                        )
                        saved_count += 1

                    except (ValueError, TypeError):
                        pass  # rəqəmə çevrilməyən dəyəri atla

            messages.success(
                request,
                f"{depo_name} üçün {saved_count} region satışı uğurla idxal edildi!"
            )

        except Exception as e:
            messages.error(request, f"Excel oxunarkən xəta baş verdi: {str(e)}")

        return redirect('depo-sales')

    # ================= 2. AJAX FILTER (GET) =================
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        depo_mapping   = {'avromed': 'Avromed', 'zeytun': 'Zeytun', 'bine': 'Binə'}
        selected_depo  = depo_mapping.get(request.GET.get('depo', 'bine'), 'Binə')
        region_type    = request.GET.get('region_type', 'all')
        selected_month = request.GET.get('month', 'all')

        regions_queryset = Region.objects.all().order_by("id")
        if region_type != 'all':
            regions_queryset = regions_queryset.filter(region_type=region_type)

        sales_data = []
        for region in regions_queryset:
            region_sales = []
            for drug in all_drug:
                filters = {
                    'depo_type': selected_depo,
                    'region':    region,
                    'drug':      drug,
                }
                if selected_month != 'all':
                    filters['sale_date__month'] = int(selected_month)

                total_qty = (
                    DepoSale.objects
                    .filter(**filters)
                    .aggregate(total=Sum('quantity'))['total'] or 0
                )
                region_sales.append(float(total_qty))

            sales_data.append({
                "region": region.region_name,
                "sales":  region_sales,
            })

        return JsonResponse({'sales': sales_data})

    # ================= 3. İLKİN SƏHİFƏ YÜKLƏNMƏSİ =================
    context = {
        "all_drug":     all_drug,
        "region_types": ['Bakı', 'Digər', 'Şəhər'],
        "months": [
            {'num': 1,  'name': 'Yanvar'},
            {'num': 2,  'name': 'Fevral'},
            {'num': 3,  'name': 'Mart'},
            {'num': 4,  'name': 'Aprel'},
            {'num': 5,  'name': 'May'},
            {'num': 6,  'name': 'İyun'},
            {'num': 7,  'name': 'İyul'},
            {'num': 8,  'name': 'Avqust'},
            {'num': 9,  'name': 'Sentyabr'},
            {'num': 10, 'name': 'Oktyabr'},
            {'num': 11, 'name': 'Noyabr'},
            {'num': 12, 'name': 'Dekabr'},
        ],
    }
    return render(request, "depo-sales.html", context)

def report_list(request):
    today = date.today()

    if "month" not in request.GET and "year" not in request.GET:
        q = request.GET.copy()
        q["month"] = str(today.month)
        q["year"] = str(today.year)
        return redirect(f"{request.path}?{q.urlencode()}")

    try:
        selected_month = int(request.GET.get("month") or today.month)
    except ValueError:
        selected_month = today.month

    try:
        selected_year = int(request.GET.get("year") or today.year)
    except ValueError:
        selected_year = today.year

    region = Region.objects.all().order_by('id')
    drug = Medical.objects.active().order_by('id')
    years = list(range(today.year - 1, today.year + 2))

    context = {
        "region": region,
        "drug": drug,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "years": years,
    }
    return render(request, "reports/report.html", context )


def d(v):
    try:
        return Decimal(str(v or 0))
    except (TypeError, ValueError):
        return Decimal('0')
    

def ajax_region_report(request):
    region_id = request.GET.get("region_id")
    month = request.GET.get("month")
    year = request.GET.get("year")
    borc_filter = request.GET.get("borc")
    page = request.GET.get("page", 1)  # Page parametri
    per_page = 20 

    if not region_id:
        return JsonResponse({"results": [], "total_pages": 0, "current_page": 1})

    doctors = Doctors.objects.filter(bolge_id=region_id, is_active=True).order_by("id")
    result = []

    # Satışlar (bütün region üçün filtr)
    sales = Sale.objects.filter(region_id=region_id)
    if month:
        try:
            ay = int(month)
            sales = sales.filter(sale_date__month=ay)
        except ValueError:
            sales = Sale.objects.none()
    
    if year:
        try:
            il = int(year)
            sales = sales.filter(sale_date__year=il)
        except ValueError:
            pass

    sales_exist = sales.exists()

    for doctor in doctors:
        # BÜTÜN dəyişənləri əvvəlcədən təyin et
        previous_debt = d(0)
        borc = d(0)
        avans = d(0)
        investisiya = d(0)
        geriqaytarma = d(0)  
        datasiya = d(0)
        hekimden_silinen = d(0)
        hesablanan_miqdar = d(0)

        # Seçilən ay üçün mövcud hesabatı tap
        monthly_report = None
        if month:
            try:
                il = int(year) if year else date.today().year
                report_month = date(il, int(month), 1)
                monthly_report = MonthlyDoctorReport.objects.filter(
                    doctor=doctor,
                    report_month=report_month
                ).first()
            except ValueError:
                pass

        if monthly_report:
            previous_debt = d(monthly_report.yekun_borc or 0)
            borc = d(monthly_report.borc or 0)
            avans = d(monthly_report.avans or 0)
            investisiya = d(monthly_report.investisiya or 0)
            datasiya = d(0)
            hekimden_silinen = d(monthly_report.hekimden_silinen or 0)
            hesablanan_miqdar = d(monthly_report.hesablanan_miqdar or 0)
            geriqaytarma = d(monthly_report.geriqaytarma or 0)
        else:
            # Hesabat bağlanmayıbsa, seçilmiş ay və il üzrə ödənişləri hesabla
            previous_debt = d(doctor.previous_debt or 0)
            borc = d(doctor.borc or 0)
            datasiya = d(doctor.datasiya or 0)
            hekimden_silinen = d(doctor.hekimden_silinen or 0)
            hesablanan_miqdar = d(doctor.hesablanan_miqdar or 0)
            
            # Ödənişləri seçilmiş ay və il üzrə filtrələ
            payments = Payment_doctor.objects.filter(doctor=doctor)
            if month and year:
                try:
                    ay = int(month)
                    il = int(year)
                    payments = payments.filter(date__month=ay, date__year=il)
                except ValueError:
                    payments = Payment_doctor.objects.none()
            elif month:
                try:
                    ay = int(month)
                    payments = payments.filter(date__month=ay)
                except ValueError:
                    payments = Payment_doctor.objects.none()
            elif year:
                try:
                    il = int(year)
                    payments = payments.filter(date__year=il)
                except ValueError:
                    payments = Payment_doctor.objects.none()
            
            # Ödəniş növlərinə görə topla
            avans = d(0)
            investisiya = d(0)
            geriqaytarma = d(0)
            
            for payment in payments:
                if payment.payment_type == 'Avans':
                    avans += d(payment.pay)
                elif payment.payment_type == 'İnvest':
                    investisiya += d(payment.pay)
                elif payment.payment_type == 'Geri_qaytarma':
                    geriqaytarma += d(payment.pay)

        # Əgər satış yoxdursa bu iki sahə sıfır olsun
        if not sales_exist:
            hekimden_silinen = d(0)
            hesablanan_miqdar = d(0)

        # Dərman məlumatları (reseptlər)
        recipe_drugs = RecipeDrug.objects.filter(
            recipe__dr=doctor,
            recipe__region_id=region_id,
            drug__status=True,
        )
        if month:
            try:
                ay = int(month)
                recipe_drugs = recipe_drugs.filter(recipe__date__month=ay)
            except ValueError:
                recipe_drugs = RecipeDrug.objects.none()
        
        if year:
            try:
                il = int(year)
                recipe_drugs = recipe_drugs.filter(recipe__date__year=il)
            except ValueError:
                pass

        drugs_agg = recipe_drugs.values('drug__med_name').annotate(total_count=Sum('number'))

        drugs = []
        total = 0
        for d_item in drugs_agg:
            drugs.append({
                "name": d_item['drug__med_name'],
                "count": d_item['total_count']
            })
            total += d_item['total_count']

        yekun_borc = previous_debt + avans + investisiya - geriqaytarma +  datasiya - hekimden_silinen

        if borc_filter == "borclu" and yekun_borc <= 0:
            continue
        if borc_filter == "borcsuz" and yekun_borc > 0:
            continue

        result.append({
            "bolge": doctor.bolge.region_name if doctor.bolge else "",
            "doctor": doctor.ad,
            "doctor_id": doctor.id,
            "barcode": doctor.barkod,
            "city": doctor.city.city_name if doctor.city else "",
            "derece": doctor.get_derece_display(),
            "ixtisas": doctor.get_ixtisas_display(),
            "previous_debt": float(previous_debt),
            "borc": float(borc),
            "avans": float(avans),
            "investisiya": float(investisiya),
            "geriqaytarma": float(geriqaytarma),  # Artıq həmişə təyin olunub
            "datasiya": float(datasiya),
            "hekimden_silinen": float(hekimden_silinen),
            "hesablanan_miqdar": float(hesablanan_miqdar),
            "drugs": drugs,
            "total": float(total),
            "yekun_borc": float(yekun_borc),
        })

    paginator = Paginator(result, per_page)
    
    try:
        current_page = paginator.page(page)
        paginated_results = list(current_page.object_list)
    except:
        current_page = paginator.page(1)
        paginated_results = list(current_page.object_list)

    return JsonResponse({
        "results": paginated_results,
        "total_pages": paginator.num_pages,
        "current_page": current_page.number,
        "has_previous": current_page.has_previous(),
        "has_next": current_page.has_next(),
        "total_results": len(result)
    }, json_dumps_params={'ensure_ascii': False})



@csrf_exempt
def hesabat_bagla(request):
    if request.method == "POST":
        try:
            ay = int(request.POST.get("ay"))
            il = int(request.POST.get("il"))
            region_raw = (request.POST.get("region_id") or "").strip()
            ay_tarixi = date(il, ay, 1)

            # Boş, "all" və s. — bütün bölgələrin həkimləri üçün bir dəfədə bağlama
            close_all_regions = (
                not region_raw
                or region_raw.lower() == "all"
                or region_raw == "__all__"
            )
            if close_all_regions:
                doctors_qs = Doctors.objects.all().select_related("bolge")
            else:
                try:
                    region_id = int(region_raw)
                except (TypeError, ValueError):
                    return JsonResponse(
                        {"success": False, "message": "Bölgə seçimi etibarsızdır."},
                        status=400,
                    )
                doctors_qs = Doctors.objects.filter(bolge_id=region_id).select_related(
                    "bolge"
                )

            with transaction.atomic():
                for doctor in doctors_qs.select_for_update():
                    # @property metodlarından oxu (avans, investisiya)
                    avans_total = D(doctor.avans or 0)
                    invest_total = D(doctor.investisiya or 0)
                    geriqaytarma_total = D(doctor.geriqaytarma or 0)
                    
                    # Database field-lərindən oxu
                    previous_debt = D(doctor.previous_debt or 0)
                    borc = D(doctor.borc or 0)
                    hekimden_silinen = D(doctor.hekimden_silinen or 0)
                    datasiya = D(doctor.datasiya or 0)
                    hesablanan_miqdar = D(doctor.hesablanan_miqdar or 0)

                    # Yekun borcu hesabla
                    yekun_borc = (previous_debt + borc + avans_total + invest_total + 
                                 datasiya - hekimden_silinen - geriqaytarma_total)

                    # Reseptləri götür və dərman sayını hesabla
                    recipes = Recipe.objects.filter(
                        dr=doctor,
                        region=doctor.bolge,
                        date__year=ay_tarixi.year,
                        date__month=ay_tarixi.month
                    )
                    total_drugs = sum(item.number for recipe in recipes for item in recipe.drugs.all())

                    # Cari ay üçün hesabat yarat və ya yenilə - YALNIZ MÖVCUD FIELD-LƏR
                    MonthlyDoctorReport.objects.update_or_create(
                        doctor=doctor,
                        report_month=ay_tarixi,
                        defaults={
                            "region": doctor.bolge,
                            "borc": float(borc),
                            "avans": float(avans_total),
                            "investisiya": float(invest_total),
                            "hekimden_silinen": float(hekimden_silinen),
                            "hesablanan_miqdar": float(hesablanan_miqdar),
                            "yekun_borc": float(yekun_borc),
                            "recipe_total_drugs": total_drugs,
                        }
                    )   

                    Payment_doctor.objects.filter(
                        doctor=doctor,
                        date__year=ay_tarixi.year,
                        date__month=ay_tarixi.month
                    ).update(is_closed=True)


                    # Növbəti aya devr et:
                    # 1. Köhnə borcu yekun_borc et
                    doctor.previous_debt = yekun_borc
                    
                    # 2. Cari ayın məlumatlarını sıfırla (sadece database field-ləri)
                    doctor.borc = 0
                    doctor.hekimden_silinen = 0
                    doctor.datasiya = 0
                    doctor.hesablanan_miqdar = 0
                    
                    doctor.save()

            aylar_ad = {
                1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "İyun",
                7: "İyul", 8: "Avqust", 9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
            }
            ay_ad = aylar_ad.get(ay, ay_tarixi.strftime('%m'))
            if close_all_regions:
                msg = f"{ay_ad} {il} — bütün bölgələr üzrə hesabat uğurla bağlandı."
            else:
                msg = f"{ay_ad} {il} ayının hesabatı uğurla bağlandı."
            return JsonResponse({"success": True, "message": msg})

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Xəta baş verdi: {str(e)}"})

    return JsonResponse({"success": False, "message": "Yalnız POST icazəlidir."})


    
def export_region_report_excel(request):
    region_id = request.GET.get("region_id")
    month = request.GET.get("month")
    year = request.GET.get("year")
    borc_filter = request.GET.get("borc")
    drugs = list(Medical.objects.active().order_by('id'))

    if not region_id:
        return HttpResponse("Bölgə seçilməyib.", status=400)

    doctors = Doctors.objects.filter(bolge_id=region_id, is_active=True).select_related('bolge')
    
    # Satışların olub-olmadığını yoxla (AJAX funksiyası ilə eyni məntiq)
    sales = Sale.objects.filter(region_id=region_id)
    if month:
        try:
            ay = int(month)
            sales = sales.filter(sale_date__month=ay)
        except ValueError:
            sales = Sale.objects.none()
    
    if year:
        try:
            il = int(year)
            sales = sales.filter(sale_date__year=il)
        except ValueError:
            pass
    
    sales_exist = sales.exists()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bölgə Hesabatı"

    drug_totals = [0] * len(drugs)
    total_total = d(0)
    hesablanan_miqdar_total = d(0)
    hekimden_silinen_total = d(0)
    avans_total = d(0)
    investisiya_total = d(0)
    geriqaytarma_total = d(0)
    datasiya_total = d(0)
    yekun_borc_total = d(0)
    previous_debt_total = d(0)
    show_geriqaytarma = False
    show_datasiya = False
    show_seher = False
    pending_rows = []
    row_idx = 0

    for doctor in doctors:
        # Seçilən ay üçün mövcud hesabatı tap (AJAX ilə uyğunlaşdır)
        monthly_report = None
        if month:
            try:
                ay = int(month)
                il = int(year) if year else date.today().year
                report_month = date(il, ay, 1)
                monthly_report = MonthlyDoctorReport.objects.filter(
                    doctor=doctor,
                    report_month=report_month
                ).first()
            except ValueError:
                pass

        if monthly_report:
            # Hesabat bağlanıbsa, həmin ayın dəyərlərini göstər
            previous_debt = d(monthly_report.yekun_borc or 0)
            borc = d(monthly_report.borc or 0)
            avans = d(monthly_report.avans or 0)
            investisiya = d(monthly_report.investisiya or 0)
            geriqaytarma = d(monthly_report.geriqaytarma or 0)
            datasiya = d(0)
            hekimden_silinen = d(monthly_report.hekimden_silinen or 0)
            hesablanan_miqdar = d(monthly_report.hesablanan_miqdar or 0)
        else:
            # Hesabat hələ bağlanmayıbsa, seçilmiş ay və il üzrə ödənişləri hesabla
            previous_debt = d(doctor.previous_debt or 0)
            borc = d(doctor.borc or 0)
            datasiya = d(doctor.datasiya or 0)
            hekimden_silinen = d(doctor.hekimden_silinen or 0)
            hesablanan_miqdar = d(doctor.hesablanan_miqdar or 0)
            
            # Ödənişləri seçilmiş ay və il üzrə filtrələ
            payments = Payment_doctor.objects.filter(doctor=doctor)
            if month and year:
                try:
                    ay = int(month)
                    il = int(year)
                    payments = payments.filter(date__month=ay, date__year=il)
                except ValueError:
                    payments = Payment_doctor.objects.none()
            elif month:
                try:
                    ay = int(month)
                    payments = payments.filter(date__month=ay)
                except ValueError:
                    payments = Payment_doctor.objects.none()
            elif year:
                try:
                    il = int(year)
                    payments = payments.filter(date__year=il)
                except ValueError:
                    payments = Payment_doctor.objects.none()
            
            # Ödəniş növlərinə görə topla
            avans = d(0)
            investisiya = d(0)
            geriqaytarma = d(0)
            
            for payment in payments:
                if payment.payment_type == 'Avans':
                    avans += d(payment.pay)
                elif payment.payment_type == 'İnvest':
                    investisiya += d(payment.pay)
                elif payment.payment_type == 'Geri_qaytarma':
                    geriqaytarma += d(payment.pay)

        # ƏGər satış yoxdursa, bu iki sahəni sıfırla (AJAX ilə eyni məntiq)
        if not sales_exist:
            hekimden_silinen = d(0)
            hesablanan_miqdar = d(0)

        yekun_borc = previous_debt + avans + investisiya - geriqaytarma + datasiya - hekimden_silinen
        
        # Borc filter tətbiq et
        if borc_filter == "borclu" and yekun_borc <= 0:
            continue
        if borc_filter == "borcsuz" and yekun_borc > 0:
            continue

        if geriqaytarma > 0:
            show_geriqaytarma = True
        if datasiya > 0:
            show_datasiya = True

        city_name = doctor.city.city_name if doctor.city else ""
        if city_name.strip():
            show_seher = True

        # Recipes və drug məlumatları (month filter ilə)
        recipes = Recipe.objects.filter(dr=doctor, region_id=region_id)
        if month:
            try:
                ay = int(month)
                recipes = recipes.filter(date__month=ay)
            except ValueError:
                recipes = Recipe.objects.none()
        
        if year:
            try:
                il = int(year)
                recipes = recipes.filter(date__year=il)
            except ValueError:
                pass

        drug_map = defaultdict(int)
        total = 0
        for recipe in recipes:
            # RecipeDrug vasitəsilə dərmanları al (AJAX ilə uyğunlaşdır)
            recipe_drugs = RecipeDrug.objects.filter(recipe=recipe, drug__status=True)
            for item in recipe_drugs:
                drug_map[item.drug.med_name] += item.number
                total += item.number

        previous_debt_total += previous_debt
        hesablanan_miqdar_total += hesablanan_miqdar
        hekimden_silinen_total += hekimden_silinen
        avans_total += avans
        investisiya_total += investisiya
        geriqaytarma_total += geriqaytarma
        datasiya_total += datasiya
        yekun_borc_total += yekun_borc
        total_total += total

        row_idx += 1
        drug_values = []
        for i, drug in enumerate(drugs):
            val = drug_map.get(drug.med_name, 0)
            drug_totals[i] += val
            drug_values.append(val)

        pending_rows.append([
            row_idx,
            doctor.bolge.region_name,
            doctor.ad,
            doctor.city.city_name if doctor.city else "",
            doctor.get_derece_display(),
            doctor.get_ixtisas_display(),
            float(previous_debt),
            *drug_values,
            float(total),
            float(hesablanan_miqdar),
            float(hekimden_silinen),
            float(avans),
            float(investisiya),
            float(geriqaytarma),
            float(datasiya),
            float(yekun_borc),
        ])

    tail_headers = [
        "Total", "Hesablanan Miqdar", "Həkimdən Silinən", "Avans", "İnvestisiya",
    ]
    if show_geriqaytarma:
        tail_headers.append("Geri qaytarma")
    if show_datasiya:
        tail_headers.append("Datasiya")
    tail_headers.append("Yekun Borc")

    prefix_headers = ["№", "Bölgə", "Həkim"]
    if show_seher:
        prefix_headers.append("Şəhər")
    prefix_headers += ["Dərəcə", "İxtisas", "Əvvəlki Borc"]

    headers = prefix_headers + [d.med_name for d in drugs] + tail_headers

    ws.append([])
    ws.append(headers)

    bold_font = Font(bold=True, color="060411")
    header_fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
    thin = Side(style='thin', color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[2]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)

    ws.freeze_panes = "A3"

    prefix_cols = 6 + (1 if show_seher else 0)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    for i in range(len(drugs)):
        col_letter = get_column_letter(prefix_cols + 1 + i)
        ws.column_dimensions[col_letter].width = 4

    base_col = prefix_cols + 1 + len(drugs)
    for i in range(len(tail_headers)):
        col_letter = get_column_letter(base_col + i)
        ws.column_dimensions[col_letter].width = 8

    drug_start = 7
    for row in pending_rows:
        prefix_values = [row[0], row[1], row[2]]
        if show_seher:
            prefix_values.append(row[3])
        prefix_values.extend([row[4], row[5], row[6]])
        drug_values = row[drug_start:drug_start + len(drugs)]
        tail_values = row[drug_start + len(drugs):drug_start + len(drugs) + 5]
        optional_values = row[drug_start + len(drugs) + 5:drug_start + len(drugs) + 7]
        yekun_value = row[drug_start + len(drugs) + 7]
        export_row = prefix_values + drug_values + tail_values
        if show_geriqaytarma:
            export_row.append(optional_values[0])
        if show_datasiya:
            export_row.append(optional_values[1])
        export_row.append(yekun_value)
        ws.append(export_row)

        current_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Alt sətir (cəmi) - bütün cəmləri göstər
    total_row_idx = ws.max_row + 1
    
    # Əvvəlcə bütün xanaları boş doldur
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row_idx, column=col, value="")
    
    # "Cəmi" labelini "Bölgə" sütununa qoy (column 2)
    cemi_cell = ws.cell(row=total_row_idx, column=2, value="Cəmi")
    cemi_cell.font = Font(bold=True)
    cemi_cell.border = thin_border
    cemi_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Əvvəlki borcun cəmi
    prev_debt_cell = ws.cell(row=total_row_idx, column=prefix_cols, value=float(previous_debt_total))
    prev_debt_cell.font = Font(bold=True)
    prev_debt_cell.border = thin_border
    prev_debt_cell.alignment = Alignment(horizontal="center", vertical="center")

    drug_start_col = prefix_cols + 1
    for i, total_val in enumerate(drug_totals):
        cell = ws.cell(row=total_row_idx, column=drug_start_col + i, value=total_val)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    final_totals = [
        ("Total", float(total_total)),
        ("Hesablanan Miqdar", float(hesablanan_miqdar_total)),
        ("Həkimdən Silinən", float(hekimden_silinen_total)),
        ("Avans", float(avans_total)),
        ("İnvestisiya", float(investisiya_total)),
    ]
    if show_geriqaytarma:
        final_totals.append(("Geri qaytarma", float(geriqaytarma_total)))
    if show_datasiya:
        final_totals.append(("Datasiya", float(datasiya_total)))
    final_totals.append(("Yekun Borc", float(yekun_borc_total)))
    
    for i, (label, value) in enumerate(final_totals):
        cell = ws.cell(row=total_row_idx, column=base_col + i, value=value)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Qalan boş xanalara border əlavə et
    empty_border_cols = [1, 3]
    if show_seher:
        empty_border_cols.extend([4, 5, 6])
    else:
        empty_border_cols.extend([4, 5])
    for col in empty_border_cols:
        cell = ws.cell(row=total_row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fayl adı
    region_name = doctors[0].bolge.region_name if doctors.exists() else "Region"
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{region_name}_Borc_{today_str}.xlsx".replace(" ", "_")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
     
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    wb.save(response)
    return response

def kohne_hesabat(request):
    region_list = Region.objects.all()
    drug_list = Medical.objects.active().order_by('id')

    aylar = [
        (1, "Yanvar"),
        (2, "Fevral"),
        (3, "Mart"),
        (4, "Aprel"),
        (5, "May"),
        (6, "İyun"),
        (7, "İyul"),
        (8, "Avqust"),
        (9, "Sentyabr"),
        (10, "Oktyabr"),
        (11, "Noyabr"),
        (12, "Dekabr"),
    ]

    # Mövcud illəri gətir
    current_year = timezone.now().year
    years = range(current_year - 5, current_year + 1)

    context = {
        'region': region_list,
        'drug': drug_list,
        'aylar': aylar,
        'years': reversed(list(years)),  # Ən son illər üstə
    }
    return render(request, 'reports/old-report.html', context)


def kohne_region_ajax(request):
    region_id = request.GET.get('region_id')
    month = request.GET.get('month')
    year = request.GET.get('year', timezone.now().year)  # İl parametrini də əlavə et

    if not region_id or not month:
        return JsonResponse({'error': 'Region və ay seçilməyib'}, status=400)

    try:
        report_month = date(year=int(year), month=int(month), day=1)
    except ValueError:
        return JsonResponse({'error': 'Yanlış tarix dəyəri'}, status=400)

    # Hesabatları gətir
    reports = MonthlyDoctorReport.objects.filter(
        region_id=region_id,
        report_month=report_month
    ).select_related('doctor', 'region')

    results = []
    for report in reports:
        # Drugs siyahısını JSON olaraq oxu
        try:
            drugs_list = json.loads(report.recipe_drugs_list or '[]')
        except json.JSONDecodeError:
            drugs_list = []

        # Əgər drugs_list boşdursa, həkimə aid RecipeDrug məlumatlarından yarat
        if not drugs_list:
            counts_qs = RecipeDrug.objects.filter(
                recipe__dr_id=report.doctor.id,
                recipe__date__year=report_month.year,
                recipe__date__month=report_month.month,
                drug__status=True,
            )
            temp_drugs = []
            total_drugs = 0
            for item in counts_qs:
                temp_drugs.append({
                    'name': item.drug.med_name,
                    'count': float(item.number or 0)
                })
                total_drugs += item.number or 0
            drugs_list = temp_drugs
            report.recipe_total_drugs = total_drugs  # toplam dərman sayını saxla
            report.recipe_drugs_list = json.dumps(drugs_list)
            report.save()

        # Həkimə aid əsas məlumatlar
        doctor = report.doctor
        results.append({
            'bolge': report.region.region_name if report.region else '',
            'doctor': doctor.ad,
            'barcode': getattr(doctor, 'barcode', ''),
            'kategoriya': getattr(doctor, 'kategoriya', ''),
            'derece': getattr(doctor, 'derece', ''),
            'ixtisas': getattr(doctor, 'ixtisas', ''),
            'previous_debt': float(report.borc or 0),
            'drugs': drugs_list,
            'total': report.recipe_total_drugs or 0,
            'hekimden_silinen': float(report.hekimden_silinen or 0),
            'hesablanan_miqdar': float(report.hesablanan_miqdar or 0),
            'avans': float(report.avans or 0),
            'investisiya': float(report.investisiya or 0),
            'yekun_borc': float(report.yekun_borc or 0),
        })

    return JsonResponse({'results': results})


