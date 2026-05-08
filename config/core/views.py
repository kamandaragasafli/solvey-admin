
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse
from django.shortcuts import render,redirect,get_object_or_404
from medicine.models import Medical
from decimal import Decimal
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.contrib import messages
from datetime import datetime
from django.db.models import Sum, DecimalField
from django.http import JsonResponse

from django.db.models import Count,  Sum
from django.utils import timezone
from django.utils.timezone import make_aware, is_aware
from datetime import datetime, time, timedelta

from django.db.models import Count
from django.db.models.functions import Coalesce
from doctors.models import Doctors, Recipe, RecipeDrug, RealSales, RealSalesDrug
from regions.models import Region, Hospital, City
from payment.models import Sale
from core.models import DeletedRecipeDrugLog
from itertools import chain
from operator import attrgetter
from decimal import Decimal, ROUND_HALF_UP
import urllib.parse
from django.db.models import Sum, DecimalField, Q
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import requests
import os
from django.contrib.auth.decorators import login_required



def ensure_aware_datetime(value):
    """Giriş dəyəri datetime.date və ya datetime.datetime ola bilər"""
    if isinstance(value, datetime):
        dt = value
    else:  # əgər bu datetime.date-dirsə
        dt = datetime.combine(value, time.min)
    
    if not is_aware(dt):
        return make_aware(dt)
    return dt


# Əsas Səhifə
def index(request):
    # Həkimlər, xəstəxanalar, şəhərlər
    doctors = Doctors.objects.select_related('bolge').all()
    total_doctors = doctors.count()
    total_hospitals = Hospital.objects.count()
    total_cities = City.objects.count()
    all_drug = Medical.objects.all()



    # Şəhər statistikası
    city_stats = Hospital.objects.values('city__city_name').annotate(say=Count('id'))
    city_labels = [x['city__city_name'] for x in city_stats]
    city_counts = [x['say'] for x in city_stats]

    baki_hospitals_b = Hospital.objects.filter(region_net__region_type='Bakı').count()

    # Son reseptlər
    last_recipes = Recipe.objects.select_related('dr').order_by('-date')[:10]

    # Son silinmiş loglar
    deleted_logs = DeletedRecipeDrugLog.objects.order_by('-deleted_at')[:10]
    for log in deleted_logs:
        try:
            log.recipe = Recipe.objects.select_related('dr').get(id=log.recipe_id)
        except Recipe.DoesNotExist:
            log.recipe = None

    # Tip ayırd edilməsi üçün obyektlərə atribut əlavə edək
    for r in last_recipes:
        r.event_type = "added"
        r.event_date = ensure_aware_datetime(r.date)

    for l in deleted_logs:
        l.event_type = "deleted"
        l.event_date = ensure_aware_datetime(l.deleted_at)

    # Birləşdir və sırala
    combined_events = sorted(
        chain(last_recipes, deleted_logs),
        key=attrgetter("event_date"),
        reverse=True
    )[:10]

    # Aktiv həkim sayı
    active_count = (
        Doctors.objects
        .annotate(derman_sayi=Count('recipe__drugs'))
        .filter(derman_sayi__gt=1)
        .count()
    )



    today = timezone.localdate()
    first_day_of_month = today.replace(day=1)

    # Digər bölgələr
    diger_region = Region.objects.filter(region_type="Digər")
    baki_region = Region.objects.filter(region_type="Bakı")

    # Bakı üçün data
    baki_region_drug_counts_daily = {}
    baki_region_drug_counts_monthly = {}
    baki_region_daily_totals = {}
    baki_region_monthly_totals = {}

    for region in baki_region:
        baki_region_drug_counts_daily[region.region_name] = {}
        baki_region_drug_counts_monthly[region.region_name] = {}

        for drug in all_drug:
            baki_region_drug_counts_daily[region.region_name][drug.med_name] = 0
            baki_region_drug_counts_monthly[region.region_name][drug.med_name] = 0

        baki_region_daily_totals[region.region_name] = 0
        baki_region_monthly_totals[region.region_name] = 0

    # RecipeDrug məlumatları
    drugs_data = (
        RecipeDrug.objects
        .filter(recipe__region__in=diger_region)
        .values("recipe__region__region_name", "drug__med_name", "recipe__date")
        .annotate(total=Sum("number"))
    )

    region_drug_counts = {}
    region_drug_counts_monthly = {}
    region_daily_totals = {}
    region_monthly_totals = {}

    # Dövr: hər bölgə üçün gündəlik və aylıq hesabla
    for region in diger_region:
        region_drug_counts[region.region_name] = {}
        region_drug_counts_monthly[region.region_name] = {}
        daily_total = Decimal(0)
        monthly_total = Decimal(0)

        for drug in all_drug:
            # Gündəlik
            daily = sum(
                item["total"]
                for item in drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and item["recipe__date"] == today
            )
            # Aylıq
            monthly = sum(
                item["total"]
                for item in drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and first_day_of_month <= item["recipe__date"] <= today
            )

            region_drug_counts[region.region_name][drug.med_name] = daily
            region_drug_counts_monthly[region.region_name][drug.med_name] = monthly
            daily_total += daily
            monthly_total += monthly

        region_daily_totals[region.region_name] = daily_total
        region_monthly_totals[region.region_name] = monthly_total

    # ✅ Dövr bitdi, indi sıralama et
    diger_region = sorted(
        diger_region,
        key=lambda r, totals=region_monthly_totals: totals.get(r.region_name, 0),
        reverse=True
    )
    # Bakı regionları üçün
    baki_drugs_data = (
        RecipeDrug.objects
        .filter(recipe__region__in=baki_region)
        .values("recipe__region__region_name", "drug__med_name", "recipe__date")
        .annotate(total=Sum("number"))
    )

    baki_region_drug_counts_daily = {}
    baki_region_drug_counts_monthly = {}
    baki_region_daily_totals = {}
    baki_region_monthly_totals = {}

    for region in baki_region:
        baki_region_drug_counts_daily[region.region_name] = {}
        baki_region_drug_counts_monthly[region.region_name] = {}
        daily_total = Decimal(0)
        monthly_total = Decimal(0)

        for drug in all_drug:
            # Günlük
            daily = sum(
                item["total"]
                for item in baki_drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and item["recipe__date"] == today
            )
            # Aylıq
            monthly = sum(
                item["total"]
                for item in baki_drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and first_day_of_month <= item["recipe__date"] <= today
            )

            baki_region_drug_counts_daily[region.region_name][drug.med_name] = daily
            baki_region_drug_counts_monthly[region.region_name][drug.med_name] = monthly
            daily_total += daily
            monthly_total += monthly

        baki_region_daily_totals[region.region_name] = daily_total
        baki_region_monthly_totals[region.region_name] = monthly_total

    baki_region = sorted(
    baki_region,
    key=lambda r, totals=baki_region_monthly_totals: totals.get(r.region_name, 0),
    reverse=True
)

    # Şəhər tipli bölgələr (Region.region_type = "Şəhər") — cityModal üçün
    seher_region = Region.objects.filter(region_type="Şəhər")
    seher_drugs_data = (
        RecipeDrug.objects
        .filter(recipe__region__in=seher_region)
        .values("recipe__region__region_name", "drug__med_name", "recipe__date")
        .annotate(total=Sum("number"))
    )
    seher_region_drug_counts = {}
    seher_region_drug_counts_monthly = {}
    seher_region_daily_totals = {}
    seher_region_monthly_totals = {}
    for region in seher_region:
        seher_region_drug_counts[region.region_name] = {}
        seher_region_drug_counts_monthly[region.region_name] = {}
        daily_total = Decimal(0)
        monthly_total = Decimal(0)
        for drug in all_drug:
            daily = sum(
                item["total"]
                for item in seher_drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and item["recipe__date"] == today
            )
            monthly = sum(
                item["total"]
                for item in seher_drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and first_day_of_month <= item["recipe__date"] <= today
            )
            seher_region_drug_counts[region.region_name][drug.med_name] = daily
            seher_region_drug_counts_monthly[region.region_name][drug.med_name] = monthly
            daily_total += daily
            monthly_total += monthly
        seher_region_daily_totals[region.region_name] = daily_total
        seher_region_monthly_totals[region.region_name] = monthly_total
    seher_region = sorted(
        seher_region,
        key=lambda r, totals=seher_region_monthly_totals: totals.get(r.region_name, 0),
        reverse=True,
    )

    today = timezone.localdate()  # Cari tarix
    current_month = today.month   # Cari ay
    current_year = today.year  
    total_other = (
        RecipeDrug.objects
        .filter(
            recipe__region__in=diger_region,
            recipe__date__month=current_month,
            recipe__date__year=current_year
        )
        .aggregate(total=Coalesce(Sum('number', output_field=DecimalField()), Decimal('0.0')))
    )['total']
    total_other = total_other.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

    total_baku = (
        RecipeDrug.objects
        .filter(
            recipe__region__in=baki_region,
            recipe__date__month=current_month,
            recipe__date__year=current_year
        )
        .aggregate(total=Coalesce(Sum('number', output_field=DecimalField()), Decimal('0.0')))
    )['total']
    total_baku = total_baku.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

    # Context dövrün içində deyil, dövr bitdikdən sonra
    context = {
        'doctors': doctors,
        'aktiv_sayi': active_count,
        'combined_events': combined_events,
        'total_doctors': total_doctors,
        'total_hospitals': total_hospitals,
        'total_cities': total_cities,
        'city_labels': city_labels,
        'city_counts': city_counts,
        'baki_hospitals_bage': baki_hospitals_b,
        'total_other': total_other,
        'total_baku': total_baku,
        'doctor_count': total_doctors,
        'city_count': total_cities,
        'hospital_count': total_hospitals,
        "all_drug": all_drug,
        "diger_region": diger_region,
        "baki_region":baki_region,
        "region_drug_counts": region_drug_counts,
        "region_drug_counts_monthly": region_drug_counts_monthly,
        "region_daily_totals": region_daily_totals,
        "region_monthly_totals": region_monthly_totals,

        # geriyə uyğunluq: köhnə açar aylıq idi, saxlayırıq
        "baki_region_drug_counts": baki_region_drug_counts_monthly,
        "baki_region_drug_counts_daily": baki_region_drug_counts_daily,
        "baki_region_drug_counts_monthly": baki_region_drug_counts_monthly,
        "baki_region_daily_totals": baki_region_daily_totals,
        "baki_region_monthly_totals": baki_region_monthly_totals,

        "seher_region": seher_region,
        "seher_region_drug_counts": seher_region_drug_counts,
        "seher_region_drug_counts_monthly": seher_region_drug_counts_monthly,
        "seher_region_daily_totals": seher_region_daily_totals,
        "seher_region_monthly_totals": seher_region_monthly_totals,

        # Məlumat mərkəzi
        "top_doctors": list(
            Doctors.objects
            .filter(recipe__date__month=first_day_of_month.month, recipe__date__year=first_day_of_month.year)
            .annotate(rd_count=Count("recipe__drugs"))
            .filter(rd_count__gt=0)
            .order_by("-rd_count")[:10]
            .values("ad", "rd_count")
        ),
        "top_drugs": list(
            RecipeDrug.objects
            .filter(recipe__date__month=first_day_of_month.month, recipe__date__year=first_day_of_month.year)
            .values("drug__med_name")
            .annotate(total=Sum("number"))
            .order_by("-total")[:10]
        ),
        "region_comparison": list(
            Region.objects
            .annotate(
                total_qeyd=Coalesce(
                    Sum("doctors__recipe__drugs__number", filter=Q(doctors__recipe__date__month=first_day_of_month.month, doctors__recipe__date__year=first_day_of_month.year)),
                    0,
                    output_field=DecimalField()
                )
            )
            .filter(total_qeyd__gt=0)
            .order_by("-total_qeyd")[:15]
            .values("region_name", "total_qeyd")
        ),
    }

    return render(request, 'index.html', context)


@require_http_methods(["GET"])
def region_modal_monthly_data(request):
    """
    Region Qeydiyyat modalı üçün interval üzrə aylıq cəmlər.
    Giriş: date_range = 'YYYY-MM-DD - YYYY-MM-DD'
    Çıxış: hər bölgə üçün hər dərman üzrə total + ümumi total.
    """
    date_range = (request.GET.get("date_range") or "").strip()
    if " - " not in date_range:
        return JsonResponse({"ok": False, "error": "date_range tələb olunur"}, status=400)

    try:
        start_str, end_str = date_range.split(" - ", 1)
        start_date = datetime.strptime(start_str.strip(), "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str.strip(), "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"ok": False, "error": "date_range formatı yanlışdır"}, status=400)

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    regions = list(Region.objects.filter(region_type="Digər").values_list("region_name", flat=True))
    drugs = list(Medical.objects.all().order_by("id").values_list("med_name", flat=True))

    # Default 0 strukturu
    region_map = {
        rn: {"region": rn, "drugs": {dn: "0" for dn in drugs}, "total": "0"}
        for rn in regions
    }

    qs = (
        RecipeDrug.objects
        .filter(
            recipe__region__region_type="Digər",
            recipe__date__gte=start_date,
            recipe__date__lte=end_date,
        )
        .values("recipe__region__region_name", "drug__med_name")
        .annotate(total=Coalesce(Sum("number"), 0, output_field=DecimalField()))
    )

    # doldur
    for row in qs:
        rn = row["recipe__region__region_name"]
        dn = row["drug__med_name"]
        total = row["total"] or 0
        if rn in region_map and dn in region_map[rn]["drugs"]:
            region_map[rn]["drugs"][dn] = str(total)

    # totals
    for rn, payload in region_map.items():
        t = Decimal("0")
        for dn in drugs:
            try:
                t += Decimal(payload["drugs"][dn])
            except Exception:
                pass
        payload["total"] = str(t)

    return JsonResponse({"ok": True, "start": start_date.isoformat(), "end": end_date.isoformat(), "regions": list(region_map.values())})


@require_http_methods(["GET"])
def baku_modal_monthly_data(request):
    """
    Bakı Qeydiyyat modalı üçün interval üzrə aylıq cəmlər.
    Giriş: date_range = 'YYYY-MM-DD - YYYY-MM-DD'
    """
    date_range = (request.GET.get("date_range") or "").strip()
    if " - " not in date_range:
        return JsonResponse({"ok": False, "error": "date_range tələb olunur"}, status=400)

    try:
        start_str, end_str = date_range.split(" - ", 1)
        start_date = datetime.strptime(start_str.strip(), "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str.strip(), "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"ok": False, "error": "date_range formatı yanlışdır"}, status=400)

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    regions = list(Region.objects.filter(region_type="Bakı").values_list("region_name", flat=True))
    drugs = list(Medical.objects.all().order_by("id").values_list("med_name", flat=True))

    region_map = {
        rn: {"region": rn, "drugs": {dn: "0" for dn in drugs}, "total": "0"}
        for rn in regions
    }

    qs = (
        RecipeDrug.objects
        .filter(
            recipe__region__region_type="Bakı",
            recipe__date__gte=start_date,
            recipe__date__lte=end_date,
        )
        .values("recipe__region__region_name", "drug__med_name")
        .annotate(total=Coalesce(Sum("number"), 0, output_field=DecimalField()))
    )

    for row in qs:
        rn = row["recipe__region__region_name"]
        dn = row["drug__med_name"]
        total = row["total"] or 0
        if rn in region_map and dn in region_map[rn]["drugs"]:
            region_map[rn]["drugs"][dn] = str(total)

    for rn, payload in region_map.items():
        t = Decimal("0")
        for dn in drugs:
            try:
                t += Decimal(payload["drugs"][dn])
            except Exception:
                pass
        payload["total"] = str(t)

    return JsonResponse({"ok": True, "start": start_date.isoformat(), "end": end_date.isoformat(), "regions": list(region_map.values())})


# Region data Start
def _get_chart_month_year(request):
    """GET-dən ay/il götür, yoxdursa cari ay/il."""
    now = timezone.localdate()
    try:
        month = int(request.GET.get('month', now.month))
        year = int(request.GET.get('year', now.year))
        if 1 <= month <= 12 and 2020 <= year <= 2030:
            return month, year
    except (ValueError, TypeError):
        pass
    return now.month, now.year


def region_drug_data_other(request):
    current_month, current_year = _get_chart_month_year(request)

    regions = (
        Region.objects
        .filter(region_type='Digər')
        .annotate(
            drug_count=Coalesce(
                Sum(
                    'doctors__recipe__drugs__number',
                    filter=(
                        Q(doctors__recipe__date__month=current_month) &
                        Q(doctors__recipe__date__year=current_year)
                    )
                ),
                0,
                output_field=DecimalField()
            )
        )
        .order_by('region_name')
    )

    labels = [r.region_name for r in regions]
    counts = [float(r.drug_count) for r in regions]

    return JsonResponse({'labels': labels, 'data': counts})


def region_drug_data_baku(request):
    current_month, current_year = _get_chart_month_year(request)

    regions = (
        Region.objects
        .filter(region_type='Bakı')
        .annotate(
            drug_count=Coalesce(
                Sum(
                    'doctors__recipe__drugs__number',
                    filter=(
                        Q(doctors__recipe__date__month=current_month) &
                        Q(doctors__recipe__date__year=current_year)
                    )
                ),
                0,
                output_field=DecimalField()
            )
        )
        .order_by('region_name')
    )

    labels = [r.region_name for r in regions]
    counts = [float(r.drug_count) for r in regions]

    return JsonResponse({'labels': labels, 'data': counts})


def region_sales_data_other(request):
    """Digər bölgələr üzrə satış (Sale modeli)."""
    current_month, current_year = _get_chart_month_year(request)

    regions = (
        Region.objects
        .filter(region_type='Digər')
        .annotate(
            sales_total=Coalesce(
                Sum(
                    'sales__quantity',
                    filter=(
                        Q(sales__sale_date__month=current_month) &
                        Q(sales__sale_date__year=current_year)
                    )
                ),
                0,
                output_field=DecimalField()
            )
        )
        .order_by('region_name')
    )

    labels = [r.region_name for r in regions]
    counts = [float(r.sales_total) for r in regions]

    return JsonResponse({'labels': labels, 'data': counts})


def region_sales_data_baku(request):
    """Bakı bölgələri üzrə satış (RealSales + RealSalesDrug)."""
    current_month, current_year = _get_chart_month_year(request)

    regions = (
        Region.objects
        .filter(region_type='Bakı')
        .annotate(
            sales_total=Coalesce(
                Sum(
                    'realsales__drug_name__numbers',
                    filter=(
                        Q(realsales__date_sale__month=current_month) &
                        Q(realsales__date_sale__year=current_year)
                    )
                ),
                0,
                output_field=DecimalField()
            )
        )
        .order_by('region_name')
    )

    labels = [r.region_name for r in regions]
    counts = [float(r.sales_total) for r in regions]

    return JsonResponse({'labels': labels, 'data': counts})


# Region data end
# Login
# def user_login(request):
#     if request.method == "POST":
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         user = authenticate(request, username=username, password=password)
        
#         if user is not None:
#             login(request, user)

#             if user.is_superuser:
#                 return redirect('/admin')  # Superuser admin panelə
#             elif user.groups.filter(name="Moderator").exists():
#                 return redirect('/admin')  # Moderator dashboard
#             elif user.groups.filter(name="İstifadəçi").exists():
#                 return redirect('movqe_gonder_view')  # Normal istifadəçi
#             else:
#                 # Qrup təyin olunmayıbsa normal istifadəçi kimi
#                 return redirect('movqe_gonder_view')  
#         else:
#             error = "İstifadəçi adı və ya şifrə yanlışdır"
#             return render(request, 'login.html', {'error': error})
    
#     return render(request, 'login.html')

def user_login(request):
    if request.method == "POST":
        from django.contrib.auth import get_user_model

        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''

        user = authenticate(request, username=username, password=password)

        # Username böyük/kiçik hərfə görə fərqlənə bilər (mobil/klaviatura/autofill).
        # Django default backend username-i case-sensitive yoxlayır.
        if user is None and username:
            User = get_user_model()
            matched = User.objects.filter(username__iexact=username).only("username", "is_active").first()
            if matched:
                user = authenticate(request, username=matched.username, password=password)
        
        if user is not None:
            if hasattr(user, "is_active") and not user.is_active:
                messages.error(request, "Bu hesab deaktiv edilib.")
                return render(request, 'login.html')
            login(request, user)
            return redirect('/admin')  # Superuser admin panelə
        else:
            messages.error(request, "İstifadəçi adı və ya şifrə yanlışdır.")
            return render(request, 'login.html')
    
    return render(request, 'login.html')



def user_logout(request):
    logout(request)
    return redirect('login')

# Aylıq və günlük Excel Faylı Çıxarışı start
def export_excel_ayliq_region(request):
    today = timezone.localdate()
    first_day_of_month = today.replace(day=1)
    mode = (request.GET.get("mode") or "daily").strip().lower()
    date_range = (request.GET.get("date_range") or "").strip()

    # aylıq interval (default: ayın əvvəlindən bu günə)
    monthly_start = first_day_of_month
    monthly_end = today
    if " - " in date_range:
        try:
            start_str, end_str = date_range.split(" - ", 1)
            monthly_start = datetime.strptime(start_str.strip(), "%Y-%m-%d").date()
            monthly_end = datetime.strptime(end_str.strip(), "%Y-%m-%d").date()
        except ValueError:
            pass
    if monthly_start > monthly_end:
        monthly_start, monthly_end = monthly_end, monthly_start

    # Digər bölgələr
    diger_region = Region.objects.filter(region_type="Digər")
    all_drug = Medical.objects.all()

    # RecipeDrug məlumatları (günlük + aylıq/interval)
    daily_map = {}
    monthly_map = {}

    daily_qs = (
        RecipeDrug.objects
        .filter(recipe__region__in=diger_region, recipe__date=today)
        .values("recipe__region__region_name", "drug__med_name")
        .annotate(total=Coalesce(Sum("number"), 0, output_field=DecimalField()))
    )
    for row in daily_qs:
        daily_map[(row["recipe__region__region_name"], row["drug__med_name"])] = row["total"]

    monthly_qs = (
        RecipeDrug.objects
        .filter(recipe__region__in=diger_region, recipe__date__gte=monthly_start, recipe__date__lte=monthly_end)
        .values("recipe__region__region_name", "drug__med_name")
        .annotate(total=Coalesce(Sum("number"), 0, output_field=DecimalField()))
    )
    for row in monthly_qs:
        monthly_map[(row["recipe__region__region_name"], row["drug__med_name"])] = row["total"]

    region_drug_counts_daily = {}
    region_drug_counts_monthly = {}
    region_daily_totals = {}
    region_monthly_totals = {}

    for region in diger_region:
        region_drug_counts_daily[region.region_name] = {}
        region_drug_counts_monthly[region.region_name] = {}
        daily_total = Decimal(0)
        monthly_total = Decimal(0)

        for drug in all_drug:
            daily = daily_map.get((region.region_name, drug.med_name), 0)
            monthly = monthly_map.get((region.region_name, drug.med_name), 0)

            region_drug_counts_daily[region.region_name][drug.med_name] = daily
            region_drug_counts_monthly[region.region_name][drug.med_name] = monthly
            daily_total += daily
            monthly_total += monthly

        region_daily_totals[region.region_name] = daily_total
        region_monthly_totals[region.region_name] = monthly_total

    # Aylıq qeydiyyata görə azalan sırayla sort
    diger_region = sorted(
        diger_region,
        key=lambda r: region_monthly_totals.get(r.region_name, 0),
        reverse=True
    )

    # Excel faylı yarat
    wb = Workbook()
    ws = wb.active
    ws.title = "Region Qeydiyyat"

    # Stil tərifləri
    bold_font = Font(bold=True, name="Calibri", size=12)
    calibri_font = Font(name="Calibri", size=11)
    bottom_alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="8ab1e3", end_color="8ab1e3", fill_type="solid")

    # Yuxarıda tarix
    if mode == "monthly":
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_drug) + 2)
        cell = ws.cell(row=1, column=1)
        cell.value = f"Tarix aralığı: {monthly_start} - {monthly_end}"
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_drug) + 3)
        cell = ws.cell(row=1, column=1)
        cell.value = f"Günlük Tarix: {today} | Aylıq interval: {monthly_start} - {monthly_end}"
    cell.font = bold_font
    cell.alignment = center_alignment

    # Header
    if mode == "monthly":
        headers = ["№", "Bölgə"] + [drug.med_name for drug in all_drug] + ["Aylıq Qeydiyyat"]
    else:
        headers = ["№", "Bölgə"] + [drug.med_name for drug in all_drug] + ["Gündəlik Qeydiyyat", "Aylıq Qeydiyyat"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = bottom_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Məzmun
    for index, region in enumerate(diger_region, start=1):
        row_num = index + 3  # 3-cü sətirdən sonra yazmağa başlayırıq

        # № sütunu
        cell = ws.cell(row=row_num, column=1)
        cell.value = index
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

        # Bölgə sütunu
        cell = ws.cell(row=row_num, column=2)
        cell.value = region.region_name
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

        # Dərmanlar
        for col_num, drug in enumerate(all_drug, start=3):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = (
                region_drug_counts_monthly[region.region_name][drug.med_name]
                if mode == "monthly"
                else region_drug_counts_daily[region.region_name][drug.med_name]
            )
            cell.font = calibri_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Cəmlər
        if mode == "monthly":
            i = len(all_drug) + 3
            cell = ws.cell(row=row_num, column=i)
            cell.value = region_monthly_totals[region.region_name]
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill
        else:
            for i, value in enumerate(
                [region_daily_totals[region.region_name], region_monthly_totals[region.region_name]],
                start=len(all_drug) + 3,
            ):
                cell = ws.cell(row=row_num, column=i)
                cell.value = value
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = header_fill

    # Aşağıda cəm
    total_row = len(diger_region) + 4

    cell = ws.cell(row=total_row, column=2)
    cell.value = "Cəm"
    cell.font = Font(bold=True, name="Calibri", color="FFFFFF", size=12)
    cell.alignment = center_alignment
    cell.border = thin_border
    cell.fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")


    for col_num, drug in enumerate(all_drug, start=3):
        cell = ws.cell(row=total_row, column=col_num)
        if mode == "monthly":
            cell.value = sum(region_drug_counts_monthly[reg.region_name][drug.med_name] for reg in diger_region)
        else:
            cell.value = sum(region_drug_counts_daily[reg.region_name][drug.med_name] for reg in diger_region)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # ümumi cəm
    if mode == "monthly":
        i = len(all_drug) + 3
        cell = ws.cell(row=total_row, column=i)
        cell.value = sum(region_monthly_totals.values())
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill
    else:
        for i, value in enumerate([sum(region_daily_totals.values()), sum(region_monthly_totals.values())], start=len(all_drug) + 3):
            cell = ws.cell(row=total_row, column=i)
            cell.value = value
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill

    # Excel faylını göndər
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if mode == "monthly":
        filename = f"Region Qeydiyyat (Aylıq) - {monthly_start}_to_{monthly_end}.xlsx"
    else:
        filename = f"Region Qeydiyyat (Günlük) - {today}.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    wb.save(response)
    return response


def export_excel_ayliq_seher(request):
    """Digər region export ilə eyni düstur; yalnız region_type = Şəhər."""
    today = timezone.localdate()
    first_day_of_month = today.replace(day=1)

    seher_region = Region.objects.filter(region_type="Şəhər")
    all_drug = Medical.objects.all()

    drugs_data = (
        RecipeDrug.objects
        .filter(recipe__region__in=seher_region)
        .values("recipe__region__region_name", "drug__med_name", "recipe__date")
        .annotate(total=Sum("number"))
    )

    region_drug_counts_daily = {}
    region_drug_counts_monthly = {}
    region_daily_totals = {}
    region_monthly_totals = {}

    for region in seher_region:
        region_drug_counts_daily[region.region_name] = {}
        region_drug_counts_monthly[region.region_name] = {}
        daily_total = Decimal(0)
        monthly_total = Decimal(0)

        for drug in all_drug:
            daily = sum(
                item["total"]
                for item in drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and item["recipe__date"] == today
            )
            monthly = sum(
                item["total"]
                for item in drugs_data
                if item["recipe__region__region_name"] == region.region_name
                and item["drug__med_name"] == drug.med_name
                and first_day_of_month <= item["recipe__date"] <= today
            )

            region_drug_counts_daily[region.region_name][drug.med_name] = daily
            region_drug_counts_monthly[region.region_name][drug.med_name] = monthly
            daily_total += daily
            monthly_total += monthly

        region_daily_totals[region.region_name] = daily_total
        region_monthly_totals[region.region_name] = monthly_total

    seher_region = sorted(
        seher_region,
        key=lambda r: region_monthly_totals.get(r.region_name, 0),
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Aylıq Şəhər Qeydiyyat"

    bold_font = Font(bold=True, name="Calibri", size=12)
    calibri_font = Font(name="Calibri", size=11)
    bottom_alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="8ab1e3", end_color="8ab1e3", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_drug) + 3)
    cell = ws.cell(row=1, column=1)
    cell.value = f"Tarix: {today} (Şəhər bölgələri)"
    cell.font = bold_font
    cell.alignment = center_alignment

    headers = ["№", "Bölgə"] + [drug.med_name for drug in all_drug] + ["Gündəlik Qeydiyyat", "Aylıq Qeydiyyat"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = bottom_alignment
        cell.border = thin_border
        cell.fill = header_fill

    for index, region in enumerate(seher_region, start=1):
        row_num = index + 3
        cell = ws.cell(row=row_num, column=1)
        cell.value = index
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

        cell = ws.cell(row=row_num, column=2)
        cell.value = region.region_name
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

        for col_num, drug in enumerate(all_drug, start=3):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = region_drug_counts_daily[region.region_name][drug.med_name]
            cell.font = calibri_font
            cell.alignment = center_alignment
            cell.border = thin_border

        for i, value in enumerate(
            [
                region_daily_totals[region.region_name],
                region_monthly_totals[region.region_name],
            ],
            start=len(all_drug) + 3,
        ):
            cell = ws.cell(row=row_num, column=i)
            cell.value = value
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill

    total_row = len(seher_region) + 4

    cell = ws.cell(row=total_row, column=2)
    cell.value = "Cəm"
    cell.font = Font(bold=True, name="Calibri", color="FFFFFF", size=12)
    cell.alignment = center_alignment
    cell.border = thin_border
    cell.fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")

    for col_num, drug in enumerate(all_drug, start=3):
        cell = ws.cell(row=total_row, column=col_num)
        cell.value = sum(
            region_drug_counts_daily[reg.region_name][drug.med_name] for reg in seher_region
        )
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    for i, value in enumerate(
        [sum(region_daily_totals.values()), sum(region_monthly_totals.values())],
        start=len(all_drug) + 3,
    ):
        cell = ws.cell(row=total_row, column=i)
        cell.value = value
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"Sumqayıt Naxçıvan Qeydiyyatı - {today}.xlsx"
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    wb.save(response)
    return response


def export_excel_ayliq_baki(request):
    today = timezone.localdate()
    first_day_of_month = today.replace(day=1)
    mode = (request.GET.get("mode") or "daily").strip().lower()
    date_range = (request.GET.get("date_range") or "").strip()

    monthly_start = first_day_of_month
    monthly_end = today
    if " - " in date_range:
        try:
            start_str, end_str = date_range.split(" - ", 1)
            monthly_start = datetime.strptime(start_str.strip(), "%Y-%m-%d").date()
            monthly_end = datetime.strptime(end_str.strip(), "%Y-%m-%d").date()
        except ValueError:
            pass
    if monthly_start > monthly_end:
        monthly_start, monthly_end = monthly_end, monthly_start

    # Digər bölgələr
    baku_region = Region.objects.filter(region_type="Bakı")
    all_drug = Medical.objects.all()

    daily_map = {}
    monthly_map = {}

    daily_qs = (
        RecipeDrug.objects
        .filter(recipe__region__in=baku_region, recipe__date=today)
        .values("recipe__region__region_name", "drug__med_name")
        .annotate(total=Coalesce(Sum("number"), 0, output_field=DecimalField()))
    )
    for row in daily_qs:
        daily_map[(row["recipe__region__region_name"], row["drug__med_name"])] = row["total"]

    monthly_qs = (
        RecipeDrug.objects
        .filter(recipe__region__in=baku_region, recipe__date__gte=monthly_start, recipe__date__lte=monthly_end)
        .values("recipe__region__region_name", "drug__med_name")
        .annotate(total=Coalesce(Sum("number"), 0, output_field=DecimalField()))
    )
    for row in monthly_qs:
        monthly_map[(row["recipe__region__region_name"], row["drug__med_name"])] = row["total"]

    region_drug_counts_daily = {}
    region_drug_counts_monthly = {}
    region_daily_totals = {}
    region_monthly_totals = {}

    for region in baku_region:
        region_drug_counts_daily[region.region_name] = {}
        region_drug_counts_monthly[region.region_name] = {}
        daily_total = Decimal(0)
        monthly_total = Decimal(0)

        for drug in all_drug:
            daily = daily_map.get((region.region_name, drug.med_name), 0)
            monthly = monthly_map.get((region.region_name, drug.med_name), 0)

            region_drug_counts_daily[region.region_name][drug.med_name] = daily
            region_drug_counts_monthly[region.region_name][drug.med_name] = monthly
            daily_total += daily
            monthly_total += monthly

        region_daily_totals[region.region_name] = daily_total
        region_monthly_totals[region.region_name] = monthly_total

    # Aylıq qeydiyyata görə azalan sırayla sort
    baku_region = sorted(
        baku_region,
        key=lambda r: region_monthly_totals.get(r.region_name, 0),
        reverse=True
    )

    # Excel faylı yarat
    wb = Workbook()
    ws = wb.active
    ws.title = "Region Qeydiyyat"

     # Stil tərifləri
    bold_font = Font(bold=True, name="Calibri", size=12)
    calibri_font = Font(name="Calibri", size=11)
    bottom_alignment = Alignment(horizontal="center", vertical="bottom", text_rotation= 90 )
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), 
        right=Side(style="thin"), 
        top=Side(style="thin"), 
        bottom=Side(style="thin")
    )
    medium_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium")
    )
    header_fill = PatternFill(start_color="8ab1e3", end_color="8ab1e3", fill_type="solid")


    # Yuxarıda tarix
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_drug)+3)
    cell = ws.cell(row=1, column=1)
    if mode == "monthly":
        cell.value = f"Tarix aralığı: {monthly_start} - {monthly_end}"
    else:
        cell.value = f"Günlük Tarix: {today}"
    cell.font = bold_font
    cell.alignment = center_alignment


    # Header
    if mode == "monthly":
        headers = ["Bölgə"] + [drug.med_name for drug in all_drug] + ["Aylıq Qeydiyyat"]
    else:
        headers = ["Bölgə"] + [drug.med_name for drug in all_drug] + ["Gündəlik Qeydiyyat", "Aylıq Qeydiyyat"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)  # Header 3-cü sətirdə
        cell.value = header
        cell.font = bold_font
        cell.alignment = bottom_alignment 
        cell.border = thin_border
        cell.fill = header_fill
        if col_num == 1 or col_num > len(all_drug):
            cell.fill = header_fill

    # Məzmun
    for row_num, region in enumerate(baku_region, start=4):
        # Region adı bold və çərçivəli
        cell = ws.cell(row=row_num, column=1)
        cell.value = region.region_name
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = header_fill

        cell.border = thin_border

        # Dərmanlar
        for col_num, drug in enumerate(all_drug, start=2):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = (
                region_drug_counts_monthly[region.region_name][drug.med_name]
                if mode == "monthly"
                else region_drug_counts_daily[region.region_name][drug.med_name]
            )
            cell.font = calibri_font
            cell.alignment = center_alignment

            cell.border = thin_border

        if mode == "monthly":
            total_col_index = len(all_drug) + 2
            cell = ws.cell(row=row_num, column=total_col_index)
            cell.value = region_monthly_totals[region.region_name]
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill
        else:
            # günlük toplam
            daily_col_index = len(all_drug) + 2
            cell = ws.cell(row=row_num, column=daily_col_index)
            cell.value = region_daily_totals[region.region_name]
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill
            # aylıq toplam
            monthly_col_index = len(all_drug) + 3
            cell = ws.cell(row=row_num, column=monthly_col_index)
            cell.value = region_monthly_totals[region.region_name]
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill

    # Aşağıda cəm
    total_row = len(baku_region) + 3
    ws.cell(row=total_row, column=1).value = "Cəm"
    ws.cell(row=total_row, column=1).font = bold_font
    ws.cell(row=total_row, column=1).alignment = center_alignment
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=1).fill = header_fill

    for col_num, drug in enumerate(all_drug, start=2):
        cell = ws.cell(row=total_row, column=col_num)
        if mode == "monthly":
            cell.value = sum(region_drug_counts_monthly[reg.region_name][drug.med_name] for reg in baku_region)
        else:
            cell.value = sum(region_drug_counts_daily[reg.region_name][drug.med_name] for reg in baku_region)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    if mode == "monthly":
        total_col_index = len(all_drug) + 2
        cell = ws.cell(row=total_row, column=total_col_index)
        cell.value = sum(region_monthly_totals.values())
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill
    else:
        daily_col_index = len(all_drug) + 2
        cell = ws.cell(row=total_row, column=daily_col_index)
        cell.value = sum(region_daily_totals.values())
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

        monthly_col_index = len(all_drug) + 3
        cell = ws.cell(row=total_row, column=monthly_col_index)
        cell.value = sum(region_monthly_totals.values())
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if mode == "monthly":
        filename = f"Bakı Qeydiyyatı.xlsx"
    else:
        filename = f"Bakı Qeydiyyatı.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    wb.save(response)
    return response

 
  # Aylıq və günlük Excel Faylı Çıxarışı son

import json
import requests
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods


@csrf_exempt
@require_http_methods(["POST"])
def openai_chat(request):
    """OpenAI API integration for chat assistant with database query capabilities"""
    try:
        from core.ai_queries import FUNCTIONS, FUNCTION_MAP

        data = json.loads(request.body)
        message = data.get('message', '').strip()
        model = data.get('model', 'gpt-4o')  # gpt-4o tövsiyə olunur - function calling daha yaxşıdır
        conversation_history = data.get('history', [])

        if not message:
            return JsonResponse({'error': 'Message is required'}, status=400)

        from django.conf import settings
        api_key = getattr(settings, 'OPENAI_API_KEY', '')

        if not api_key:
            return JsonResponse({
                'reply': 'Üzr istəyirəm, AI xidməti hazırda mövcud deyil. Zəhmət olmasa sistem administratoru ilə əlaqə saxlayın.'
            }, status=503)

        # ------------------------------------------------------------------ #
        # FUNCTIONS → TOOLS formatına çevir  (yeni OpenAI API tələbi)
        # FUNCTIONS siyahısındakı hər element belə görünür:
        #   { "name": "...", "description": "...", "parameters": {...} }
        # Tools formatı isə:
        #   { "type": "function", "function": { "name": "...", ... } }
        # ------------------------------------------------------------------ #
        tools = [{"type": "function", "function": fn} for fn in FUNCTIONS]

        # Sistem promptu
        system_prompt = {
            'role': 'system',
            'content': (
                'Sən Solvey tibbi şirkətinin admin paneli üçün TƏK bir AI assistant botsan.\n'
                'Azərbaycan dilində cavab ver. Qısa, dəqiq və faydalı cavablar ver.\n'
                'İlk cavabda Salam Kamandar deyin.\n'
                'Emoji, smaylik və dekorativ simvollardan HƏR DƏFƏ istifadə etmə.\n'
                'İstifadəçini lazımsız suallarla yormadan birbaşa cavab hazırla.\n'
                '"Başqa bir həkim istəyirsiniz?", "başqa nə lazımdır?" kimi follow-up suallarını yazma.\n'
                'Əgər mütləq kritik məlumat çatışmırsa, maksimum 1 qısa dəqiqləşdirici sual ver.\n\n'
                'Mütləq qayda:\n'
                '- Verilənlər bazasından məlumat almaq üçün HƏMİŞƏ təqdim olunmuş funksiyalardan istifadə et.\n'
                '- Heç vaxt özündən rəqəm, statistika və ya tarix UYDURMA, yalnız funksiyaların qaytardığı nəticələri istifadə et.\n'
                '- Funksiya sıfır nəticə qaytaranda bunu aydın yaz: "Bu həkim üçün bu dövrdə məlumat tapılmadı."\n\n'
                'Kontekst:\n'
                '- İstifadəçi ardıcıl suallar verəndə əvvəlki mesajdakı həkim adını və tarixi yadında saxla.\n'
                '- Məs: "Məmmədov Əsəd keçən ay nə qədər qeydiyyatı var?" + "noyabrda bəs?" → eyni həkim üçün noyabr ayına aid sorğu.\n'
                '- "noyabrda bəs?", "bu ay necə?", "keçən ay nə qədər?" kimi qısa cümlələri HƏKİM ADI kimi yox, əvvəlki sualın davamı kimi şərh et.\n\n'
                'Tarix şərhi:\n'
                '- "keçən ay" → cari tarixdən əvvəlki ay.\n'
                '- "bu ay" → cari ay.\n'
                '- "noyabrda" kimi ay adları veriləndə uyğun aya çevir.\n\n'
                'Cavab formatı:\n'
                '- Mümkün qədər sadə saxla, yalnız istifadəçi əlavə detal istəyəndə daha detallı məlumat ver.\n'
                '- Tapılmadı hallarında qısa yaz və alternativi sual kimi yox, təlimat kimi yaz. Məs: "Tam ad (ad+soyad) və ya barkod yazın."\n'
            )
        }

        messages = [system_prompt]
        messages.extend(conversation_history[-10:])
        messages.append({'role': 'user', 'content': message})

        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

        # ------------------------------------------------------------------ #
        # İlk sorğu
        # ------------------------------------------------------------------ #
        payload = {
            'model': model,
            'messages': messages,
            'tools': tools,
            'tool_choice': 'auto',
            'max_tokens': 1500,
            'temperature': 0.3,  # Daha deterministik cavablar üçün aşağı saxladıq
        }

        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers=headers,
            json=payload,
            timeout=30
        )

        if response.status_code != 200:
            error_data = response.json() if response.text else {}
            error_msg = error_data.get('error', {}).get('message', 'Xəta baş verdi')
            return JsonResponse(
                {'reply': f'Üzr istəyirəm, xəta baş verdi: {error_msg}'},
                status=response.status_code
            )

        result = response.json()
        message_obj = result['choices'][0]['message']
        finish_reason = result['choices'][0].get('finish_reason', '')

        # ------------------------------------------------------------------ #
        # Model tool çağırmaq istəyirsə  (finish_reason == "tool_calls")
        # ------------------------------------------------------------------ #
        if finish_reason == 'tool_calls' and message_obj.get('tool_calls'):
            # Model birdən çox tool çağıra bilər; bütün nəticələri toplayırıq
            messages.append(message_obj)  # assistant mesajını tarixçəyə əlavə et

            all_results = {}

            for tool_call in message_obj['tool_calls']:
                tool_call_id = tool_call['id']
                function_name = tool_call['function']['name']

                try:
                    function_args = json.loads(tool_call['function']['arguments'])
                except json.JSONDecodeError:
                    function_args = {}

                if function_name in FUNCTION_MAP:
                    try:
                        function_result = FUNCTION_MAP[function_name](**function_args)
                    except Exception as e:
                        function_result = {'error': str(e)}
                else:
                    function_result = {'error': f'{function_name} funksiyası tapılmadı.'}

                all_results[function_name] = function_result

                # Hər tool_call üçün ayrı tool mesajı əlavə et
                messages.append({
                    'role': 'tool',
                    'tool_call_id': tool_call_id,
                    'content': json.dumps(function_result, ensure_ascii=False)
                })

            # ---------------------------------------------------------------- #
            # İkinci sorğu – model nəticələri görüb son cavabı hazırlayır
            # Burada tools parametrini göndərməyə ehtiyac yoxdur (opsional)
            # ---------------------------------------------------------------- #
            payload2 = {
                'model': model,
                'messages': messages,
                'max_tokens': 1500,
                'temperature': 0.3,
            }

            response2 = requests.post(
                'https://api.openai.com/v1/chat/completions',
                headers=headers,
                json=payload2,
                timeout=30
            )

            if response2.status_code == 200:
                result2 = response2.json()
                reply = result2['choices'][0]['message']['content']
            else:
                # Fallback: nəticəni özümüz formatlayırıq
                # Birinci çağırılan funksiyaya baxırıq
                first_fn = list(all_results.keys())[0] if all_results else None
                first_result = all_results.get(first_fn) if first_fn else None
                reply = format_function_result(first_fn, first_result) if first_fn else 'Cavab alınamadı.'

            return JsonResponse({'reply': reply})

        # ------------------------------------------------------------------ #
        # Birbaşa mətn cavabı (tool çağırılmadı)
        # ------------------------------------------------------------------ #
        reply = message_obj.get('content', 'Cavab alınamadı.')
        return JsonResponse({'reply': reply})

    except requests.exceptions.Timeout:
        return JsonResponse({
            'reply': 'Üzr istəyirəm, sorğu zaman aşımına uğradı. Zəhmət olmasa yenidən cəhd edin.'
        }, status=504)
    except requests.exceptions.RequestException:
        return JsonResponse({
            'reply': 'Üzr istəyirəm, bağlantı xətası baş verdi. Zəhmət olmasa yenidən cəhd edin.'
        }, status=500)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'reply': f'Gözlənilməz xəta: {str(e)}'}, status=500)


# --------------------------------------------------------------------------- #
# format_function_result – fallback formatlaşdırma (dəyişiklik yoxdur)
# --------------------------------------------------------------------------- #
def format_function_result(function_name, result):
    """Format function results for display when second API call fails"""
    if not result:
        return 'Nəticə tapılmadı.'

    if function_name == 'get_recent_doctors':
        if not result:
            return 'Ən son əlavə olunan həkim tapılmadı.'
        text = 'Ən son əlavə olunan həkimlər:\n\n'
        for i, doctor in enumerate(result, 1):
            text += f"{i}. {doctor['ad']}\n"
            text += f"   Bölgə: {doctor['bolge']}, Şəhər: {doctor['city']}\n"
            text += f"   Klinika: {doctor['klinika']}, İxtisas: {doctor['ixtisas']}\n"
            text += f"   Dərəcə: {doctor['derece']}, Tarix: {doctor['created_at']}\n\n"
        return text

    elif function_name == 'get_doctor_statistics':
        stats = result
        text = '📊 Həkim Statistikaları:\n\n'
        text += f"Ümumi həkim sayı: {stats['total_doctors']}\n"
        text += f"Bu ay əlavə olunan: {stats['new_this_month']}\n\n"
        if stats.get('by_degree'):
            text += 'Dərəcə üzrə:\n'
            for degree, count in stats['by_degree'].items():
                text += f'  - {degree}: {count}\n'
        if stats.get('by_region'):
            text += '\nBölgə üzrə (top 5):\n'
            sorted_regions = sorted(stats['by_region'].items(), key=lambda x: x[1], reverse=True)[:5]
            for region, count in sorted_regions:
                text += f'  - {region}: {count}\n'
        return text

    elif function_name == 'get_region_statistics':
        if not result:
            return 'Bölgə tapılmadı.'
        text = '📍 Bölgə Statistikaları:\n\n'
        for region in result:
            text += f"{region['region_name']} ({region['region_type']}):\n"
            text += f"  Həkim: {region['doctor_count']}, Şəhər: {region['city_count']}, Xəstəxana: {region['hospital_count']}\n\n"
        return text

    elif function_name == 'search_doctors':
        if not result:
            return 'Axtarışa uyğun həkim tapılmadı.'
        text = f'Axtarış nəticələri ({len(result)} həkim):\n\n'
        for i, doctor in enumerate(result, 1):
            text += f"{i}. {doctor['ad']}\n"
            text += f"   Bölgə: {doctor['bolge']}\n\n"
        return text

    elif function_name == 'get_financial_summary':
        stats = result
        text = '💰 Maliyyə Ümumi Məlumatları:\n\n'
        text += f"Ümumi borc: {stats['total_debt']:.2f} ₼\n"
        text += f"Əvvəlki borc: {stats['total_previous_debt']:.2f} ₼\n"
        text += f"Borclu həkim sayı: {stats['doctors_with_debt']} / {stats['total_doctors']}\n"
        return text

    elif function_name == 'get_doctors_by_region':
        if not result:
            return 'Bu bölgədə həkim tapılmadı.'
        text = f'Bölgə həkimləri ({len(result)} həkim):\n\n'
        for i, doctor in enumerate(result, 1):
            text += f"{i}. {doctor['ad']}\n"
            text += f"   Şəhər: {doctor['city']}, Klinika: {doctor['klinika']}\n\n"
        return text

    elif function_name == 'get_doctor_financial_details':
        if not result:
            return 'Bu ada uyğun həkim tapılmadı.'
        if len(result) > 1:
            text = f'Axtarış nəticələri ({len(result)} həkim):\n\n'
            for i, doctor in enumerate(result, 1):
                text += f"{i}. {doctor['ad']}\n"
                text += f"   Bölgə: {doctor['bolge']}\n\n"
            text += 'Zəhmət olmasa daha dəqiq ad və ya barkod qeyd edin.'
            return text
        doctor = result[0]
        text = f"💳 Həkim: {doctor['ad']}\n\n"
        text += f"  - Əvvəlki borc: {doctor['previous_debt']:.2f} ₼\n"
        text += f"  - Cari borc: {doctor['borc']:.2f} ₼\n"
        text += f"  - Yekun borc: {doctor['yekun_borc']:.2f} ₼\n\n"
        for p in doctor.get('payments', []):
            text += f"  Ödəniş: {p['date']}: {p['pay']:.2f} ₼\n"
        return text

    elif function_name == 'get_doctor_prescription_stats':
        if not result.get('doctor_found'):
            return result.get('message', 'Bu ada uyğun həkim tapılmadı.')
        doctor = result['doctor']
        text = f"🧾 {doctor['ad']} – Resept statistikası:\n\n"
        text += f"  - Ümumi resept: {result['total_recipes']}\n"
        text += f"  - Ümumi dərman: {result['total_drug_count']:.1f}\n\n"
        for d in result.get('drugs', [])[:10]:
            text += f"  - {d['name']}: {d['count']:.1f} ədəd\n"
        return text

    return str(result)


    """
Yeni AI alətləri:
  1. Fayl analizi  – PDF, Excel, Word, şəkil (base64)
  2. Veb axtarış   – Google Custom Search API + xülasə
  3. İcazə sistemi – DB-yə yazmazdan əvvəl istifadəçidən təsdiq alır
"""

import json
import base64
import tempfile
import os
import requests as http_requests
from django.conf import settings
import pdfplumber


# ════════════════════════════════════════════════════════════════════════════ #
#  1. FAYL ANALİZİ
# ════════════════════════════════════════════════════════════════════════════ #

def analyze_file(file_content_b64: str, file_name: str, question: str = "") -> dict:
    """
    Base64 kodlanmış faylı analiz edir.
    file_content_b64 : base64 string
    file_name        : 'report.pdf', 'data.xlsx', 'scan.png' ...
    question         : istifadəçinin fayldan nə istədiyi
    """
    try:
        file_bytes = base64.b64decode(file_content_b64)
        ext = file_name.rsplit(".", 1)[-1].lower()
        extracted_text = ""

        # ── PDF ────────────────────────────────────────────────────────────
        if ext == "pdf":
            try:
                import pdfplumber
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                with pdfplumber.open(tmp_path) as pdf:
                    pages = []
                    for page in pdf.pages[:10]:          # maks 10 səhifə
                        text = page.extract_text() or ""
                        pages.append(text)
                extracted_text = "\n\n".join(pages)
                os.unlink(tmp_path)
            except ImportError:
                return {"success": False, "error": "pdfplumber quraşdırılmayıb: pip install pdfplumber"}

        # ── EXCEL ──────────────────────────────────────────────────────────
        elif ext in ("xlsx", "xls", "csv"):
            try:
                import pandas as pd
                with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                if ext == "csv":
                    df = pd.read_csv(tmp_path)
                else:
                    df = pd.read_excel(tmp_path)
                os.unlink(tmp_path)
                extracted_text = (
                    f"Cəmi sətir: {len(df)}, Sütunlar: {list(df.columns)}\n\n"
                    f"İlk 20 sətir:\n{df.head(20).to_string()}\n\n"
                    f"Statistika:\n{df.describe().to_string()}"
                )
            except ImportError:
                return {"success": False, "error": "pandas quraşdırılmayıb: pip install pandas openpyxl"}

        # ── WORD ───────────────────────────────────────────────────────────
        elif ext in ("docx", "doc"):
            try:
                import docx
                with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                doc = docx.Document(tmp_path)
                extracted_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                os.unlink(tmp_path)
            except ImportError:
                return {"success": False, "error": "python-docx quraşdırılmayıb: pip install python-docx"}

        # ── ŞƏKİL (OCR) ────────────────────────────────────────────────────
        elif ext in ("png", "jpg", "jpeg", "webp", "gif"):
            # GPT-4o vision ilə OCR
            api_key = getattr(settings, "OPENAI_API_KEY", "")
            vision_payload = {
                "model": "gpt-4o",
                "max_tokens": 1000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": question or "Bu şəkildə nə var? Bütün mətni və məzmunu izah et."},
                        {"type": "image_url", "image_url": {
                            "url": f"data:image/{ext};base64,{file_content_b64}"
                        }}
                    ]
                }]
            }
            resp = http_requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json=vision_payload, timeout=30
            )
            if resp.status_code == 200:
                return {
                    "success": True,
                    "file_type": "image",
                    "analysis": resp.json()["choices"][0]["message"]["content"]
                }
            else:
                return {"success": False, "error": f"Vision API xətası: {resp.text[:200]}"}

        else:
            # Sadə mətn faylları (.txt, .log, .json ...)
            try:
                extracted_text = file_bytes.decode("utf-8", errors="ignore")
            except Exception:
                return {"success": False, "error": f"Dəstəklənməyən fayl formatı: {ext}"}

        # ── Çıxarılmış mətni AI ilə analiz et ─────────────────────────────
        if not extracted_text.strip():
            return {"success": False, "error": "Fayldan mətn çıxarıla bilmədi."}

        api_key = getattr(settings, "OPENAI_API_KEY", "")
        analysis_prompt = (
            f"Aşağıdakı sənədi analiz et.\n"
            f"Sual/tapşırıq: {question or 'Əsas məlumatları xülasələ.'}\n\n"
            f"Sənəd məzmunu:\n{extracted_text[:8000]}"   # maks 8000 simvol
        )
        resp2 = http_requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "gpt-4o",
                "messages": [
                    {"role": "system", "content": "Sən Azərbaycan dilində sənəd analiz edən köməkçisən."},
                    {"role": "user", "content": analysis_prompt}
                ],
                "max_tokens": 1000,
                "temperature": 0.3
            },
            timeout=40
        )
        if resp2.status_code == 200:
            return {
                "success": True,
                "file_type": ext,
                "raw_text_preview": extracted_text[:500],
                "analysis": resp2.json()["choices"][0]["message"]["content"]
            }
        else:
            return {"success": False, "error": f"Analiz API xətası: {resp2.text[:200]}"}

    except Exception as e:
        return {"success": False, "error": str(e)}


# ════════════════════════════════════════════════════════════════════════════ #
#  2. VEB AXTARIŞ
# ════════════════════════════════════════════════════════════════════════════ #

def web_search(query: str, num_results: int = 5) -> dict:
    """
    Google Custom Search API ilə axtarış edir, nəticələri AI ilə analiz edir.
    Tələb olunur:
      settings.GOOGLE_API_KEY        – Google Cloud API açarı
      settings.GOOGLE_SEARCH_ENGINE_ID – Programmable Search Engine ID
    """
    try:
        google_api_key = getattr(settings, "GOOGLE_API_KEY", "")
        search_engine_id = getattr(settings, "GOOGLE_SEARCH_ENGINE_ID", "")

        if not google_api_key or not search_engine_id:
            # Fallback: DuckDuckGo instant answers (API açarı tələb etmir)
            return _duckduckgo_search(query, num_results)

        params = {
            "key": google_api_key,
            "cx": search_engine_id,
            "q": query,
            "num": min(num_results, 10),
            "hl": "az",        # Azərbaycan dili üstünlüyü
        }
        resp = http_requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params=params, timeout=15
        )
        if resp.status_code != 200:
            return {"success": False, "error": f"Google API xətası: {resp.text[:200]}"}

        items = resp.json().get("items", [])
        if not items:
            return {"success": True, "results": [], "summary": "Nəticə tapılmadı."}

        results = []
        for item in items:
            results.append({
                "title": item.get("title", ""),
                "link": item.get("link", ""),
                "snippet": item.get("snippet", ""),
            })

        # AI ilə xülasə
        summary = _summarize_search_results(query, results)
        return {"success": True, "query": query, "results": results, "summary": summary}

    except Exception as e:
        return {"success": False, "error": str(e)}


def _duckduckgo_search(query: str, num_results: int = 5) -> dict:
    """Google açarı olmadıqda DuckDuckGo instant answer API istifadə edir."""
    try:
        resp = http_requests.get(
            "https://api.duckduckgo.com/",
            params={"q": query, "format": "json", "no_redirect": 1, "no_html": 1},
            timeout=10
        )
        data = resp.json()
        results = []

        if data.get("AbstractText"):
            results.append({
                "title": data.get("Heading", query),
                "link": data.get("AbstractURL", ""),
                "snippet": data["AbstractText"]
            })
        for topic in data.get("RelatedTopics", [])[:num_results]:
            if isinstance(topic, dict) and topic.get("Text"):
                results.append({
                    "title": topic.get("Text", "")[:60],
                    "link": topic.get("FirstURL", ""),
                    "snippet": topic.get("Text", "")
                })

        if not results:
            return {"success": True, "results": [], "summary": f"'{query}' üçün nəticə tapılmadı."}

        summary = _summarize_search_results(query, results)
        return {"success": True, "query": query, "results": results, "summary": summary,
                "note": "DuckDuckGo instant answers istifadə edildi (Google açarı yoxdur)."}

    except Exception as e:
        return {"success": False, "error": str(e)}


def _summarize_search_results(query: str, results: list) -> str:
    """Axtarış nəticələrini AI ilə xülasələyir."""
    try:
        api_key = getattr(settings, "OPENAI_API_KEY", "")
        snippets = "\n\n".join(
            f"[{i+1}] {r['title']}\n{r['snippet']}\nMənbə: {r['link']}"
            for i, r in enumerate(results)
        )
        resp = http_requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "gpt-4o",
                "messages": [
                    {"role": "system", "content": "Azərbaycan dilində axtarış nəticələrini xülasələ."},
                    {"role": "user", "content": f"Sorğu: {query}\n\nNəticələr:\n{snippets}\n\nQısa xülasə yaz."}
                ],
                "max_tokens": 500,
                "temperature": 0.3
            },
            timeout=20
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"]
        return snippets[:500]
    except Exception:
        return "\n".join(r["snippet"] for r in results[:3])


# ════════════════════════════════════════════════════════════════════════════ #
#  3. İCAZƏ SİSTEMİ  (pending_actions cache)
# ════════════════════════════════════════════════════════════════════════════ #
#
#  Axın:
#    AI "add_doctor" çağırmaq istəyir
#    → request_permission() çağrılır  (action_id qaytarır)
#    → İstifadəçiyə "Bəli / Xeyr" göstərilir
#    → İstifadəçi "bəli" deyir → confirm_action(action_id) çağrılır → icra edilir
#    → İstifadəçi "xeyr" deyir → cancel_action(action_id)
#
#  Sadə in-memory cache (Django cache framework istifadə olunur).
# ═══════════════════════════════════════════════════════════════════════════ #

import uuid as _uuid
from django.core.cache import cache as _cache

PERMISSION_TTL = 300   # 5 dəqiqə (saniyə ilə)


def request_permission(action_type: str, action_data: dict, description: str) -> dict:
    """
    DB-yə yazma əməliyyatından əvvəl icazə tələb edir.
    action_type  : 'add_doctor' | 'update_doctor' | 'delete_doctor' | 'add_payment' | ...
    action_data  : əməliyyat üçün lazım olan parametrlər
    description  : istifadəçiyə göstəriləcək insan dostu açıqlama
    Qaytarır: { action_id, description, preview }
    """
    action_id = str(_uuid.uuid4())[:8]   # qısa ID: 'a3f7b2c1'
    _cache.set(f"pending_action:{action_id}", {
        "type": action_type,
        "data": action_data,
        "description": description,
        "status": "pending"
    }, timeout=PERMISSION_TTL)

    return {
        "action_id": action_id,
        "status": "awaiting_approval",
        "description": description,
        "preview": action_data,
        "message": (
            f"⚠️ Təsdiq tələb olunur!\n\n"
            f"Əməliyyat: {description}\n\n"
            f"Davam etmək üçün 'bəli #{action_id}' yazın.\n"
            f"Ləğv etmək üçün 'xeyr #{action_id}' yazın."
        )
    }


def confirm_action(action_id: str) -> dict:
    """
    İstifadəçi 'bəli' dedikdən sonra çağrılır.
    Əməliyyatı icra edir.
    """
    pending = _cache.get(f"pending_action:{action_id}")
    if not pending:
        return {"success": False, "error": f"#{action_id} tapılmadı və ya vaxtı keçib."}

    action_type = pending["type"]
    action_data = pending["data"]

    try:
        result = _execute_approved_action(action_type, action_data)
        _cache.delete(f"pending_action:{action_id}")
        return {"success": True, "action_id": action_id,
                "action_type": action_type, "result": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


def cancel_action(action_id: str) -> dict:
    """İstifadəçi 'xeyr' dedikdən sonra çağrılır."""
    pending = _cache.get(f"pending_action:{action_id}")
    if not pending:
        return {"success": False, "error": f"#{action_id} tapılmadı."}
    _cache.delete(f"pending_action:{action_id}")
    return {"success": True, "message": f"#{action_id} ləğv edildi."}


def _execute_approved_action(action_type: str, action_data: dict) -> dict:
    """
    Təsdiqlənmiş əməliyyatı icra edir.
    Buraya öz modellərinizdən import əlavə edin.
    """
    # ── Öz modellərinizə görə genişləndirin ─────────────────────────────
    # from doctors.models import Doctor
    # from payments.models import Payment
    # ...

    if action_type == "add_doctor":
        # Doctor.objects.create(**action_data)
        return {"message": f"Həkim əlavə edildi: {action_data.get('name', '?')}",
                "data": action_data}

    elif action_type == "update_doctor":
        doctor_id = action_data.pop("id")
        # Doctor.objects.filter(pk=doctor_id).update(**action_data)
        return {"message": f"Həkim #{doctor_id} yeniləndi", "data": action_data}

    elif action_type == "delete_doctor":
        doctor_id = action_data.get("id")
        # Doctor.objects.filter(pk=doctor_id).delete()
        return {"message": f"Həkim #{doctor_id} silindi"}

    elif action_type == "add_payment":
        # Payment.objects.create(**action_data)
        return {"message": "Ödəniş əlavə edildi", "data": action_data}

    else:
        raise ValueError(f"Bilinməyən əməliyyat tipi: {action_type}")


# ════════════════════════════════════════════════════════════════════════════ #
#  AI Assistant Page + API endpoints
# ════════════════════════════════════════════════════════════════════════════ #


@login_required(login_url="login")
def ai_assistant_page(request):
    return render(request, "ai-assistant.html")


@csrf_exempt
@require_http_methods(["POST"])
def ai_analyze_file_api(request):
    try:
        data = json.loads(request.body or "{}")
        file_content_b64 = (data.get("file_content_b64") or "").strip()
        file_name = (data.get("file_name") or "").strip()
        question = (data.get("question") or "").strip()

        if not file_content_b64 or not file_name:
            return JsonResponse({"success": False, "error": "file_content_b64 və file_name tələb olunur."}, status=400)

        return JsonResponse(analyze_file(file_content_b64=file_content_b64, file_name=file_name, question=question))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def ai_web_search_api(request):
    try:
        data = json.loads(request.body or "{}")
        query = (data.get("query") or "").strip()
        num_results = data.get("num_results", 5)

        if not query:
            return JsonResponse({"success": False, "error": "query tələb olunur."}, status=400)

        try:
            num_results = int(num_results)
        except Exception:
            num_results = 5

        return JsonResponse(web_search(query=query, num_results=max(1, min(num_results, 10))))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def ai_confirm_action_api(request):
    try:
        data = json.loads(request.body or "{}")
        action_id = (data.get("action_id") or "").strip().lstrip("#")
        if not action_id:
            return JsonResponse({"success": False, "error": "action_id tələb olunur."}, status=400)
        return JsonResponse(confirm_action(action_id))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def ai_cancel_action_api(request):
    try:
        data = json.loads(request.body or "{}")
        action_id = (data.get("action_id") or "").strip().lstrip("#")
        if not action_id:
            return JsonResponse({"success": False, "error": "action_id tələb olunur."}, status=400)
        return JsonResponse(cancel_action(action_id))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)