from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_POST
from .models import Doctors, RecipeDrug, Recipe, RealSales, RealSalesDrug
from medicine.models import Medical
from django.contrib import messages
from regions.models import Region , Hospital, City
from payment.models import Payment_doctor, Sale, MonthlyDoctorReport
from django.http import JsonResponse
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import openpyxl
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from django.core.paginator import Paginator
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Sum, Q, OuterRef, Subquery
from django.utils import timezone
from collections import defaultdict
from datetime import datetime
from core.models import DeletedRecipeDrugLog 
from django.contrib.auth.models import User 
from regions.models import Region, Hospital, City
from django.db import transaction, IntegrityError
from django.db.models import Max
import urllib.parse
from datetime import date, timedelta
from django.core.paginator import Paginator
import json
from io import BytesIO


def doctors_list(request):
    regions = Region.objects.all()
    queryset = Doctors.objects.all().prefetch_related('odenisler')

    # Arxiv / Aktiv filter
    if request.GET.get('archived') == '1':
        queryset = queryset.filter(is_active=False)
    else:
        queryset = queryset.filter(is_active=True)

    # Region filter
    region_id = request.GET.get('region_filter')
    if region_id:
        queryset = queryset.filter(bolge_id=region_id)

    # Debt filter
    debt_filter = request.GET.get('debt_filter')
    if debt_filter == 'greater':
        queryset = queryset.filter(previous_debt__gt=0)
    elif debt_filter == 'zero':
        queryset = queryset.filter(previous_debt=0)
    elif debt_filter == 'less':
        queryset = queryset.filter(previous_debt__lt=100)

    # Search filter
    search_query = request.GET.get('search')
    if search_query:
        queryset = queryset.filter(ad__icontains=search_query)

    # Prepare data with last payment info
    doctor_data = []
    for doctor in queryset:
        last_payment = doctor.odenisler.order_by('-date').first()
        doctor_data.append({
            "doctor": doctor,
            "last_payment_date": last_payment.date if last_payment else None,
            "last_payment_amount": last_payment.pay if last_payment else 0
        })

    # Pagination – 30 həkim / səhifə
    paginator = Paginator(doctor_data, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "doctors": page_obj,
        "regions": regions,
        "is_archived_view": request.GET.get('archived') == '1',
    }
    return render(request, "doctors.html", context)


def doctors_export_excel(request):
    """Export doctors list to Excel with filters applied"""
    queryset = Doctors.objects.all()

    # Arxiv / Aktiv filter
    if request.GET.get('archived') == '1':
        queryset = queryset.filter(is_active=False)
    else:
        queryset = queryset.filter(is_active=True)

    # Apply same filters as doctors_list
    region_id = request.GET.get('region_filter')
    if region_id:
        queryset = queryset.filter(bolge_id=region_id)

    debt_filter = request.GET.get('debt_filter')
    if debt_filter == 'greater':
        queryset = queryset.filter(previous_debt__gt=0)
    elif debt_filter == 'zero':
        queryset = queryset.filter(previous_debt=0)
    elif debt_filter == 'less':
        queryset = queryset.filter(previous_debt__lt=100)

    search_query = request.GET.get('search')
    if search_query:
        queryset = queryset.filter(ad__icontains=search_query)

    # Son ödəniş: hər həkim üçün prefetch yox — 2 correlated subquery (yaddın / zaman)
    latest_payment = Payment_doctor.objects.filter(doctor_id=OuterRef('pk')).order_by(
        '-date', '-id'
    )
    queryset = queryset.select_related('bolge', 'city', 'klinika').annotate(
        _export_last_pay_date=Subquery(latest_payment.values('date')[:1]),
        _export_last_pay_amt=Subquery(latest_payment.values('pay')[:1]),
    )

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Həkimlər Siyahısı"

    # Headers
    headers = [
        "№", "Barkod", "Bölgə", "Şəhər", "Ad", "Sonuncu Ödəniş Tarixi",
        "Sonuncu Ödəniş Məbləği", "İxtisas", "Dərəcə", "Kateqoriya",
        "Klinika", "Əlaqə", "Yekun Borc"
    ]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data rows
    for idx, doctor in enumerate(queryset, start=1):
        last_payment = doctor.odenisler.order_by('-date').first()
        last_payment_date = last_payment.date.strftime("%d.%m.%Y") if last_payment else "-"
        last_payment_amount = float(last_payment.pay) if last_payment else 0.0

        row = [
            idx,
            doctor.barkod or "",
            doctor.bolge.region_name if doctor.bolge else "",
            doctor.city.city_name if doctor.city else "",
            doctor.ad or "",
            last_payment_date,
            last_payment_amount,
            doctor.get_ixtisas_display() if hasattr(doctor, 'get_ixtisas_display') else (doctor.ixtisas or ""),
            doctor.get_derece_display() if hasattr(doctor, 'get_derece_display') else (doctor.derece or ""),
            f"{doctor.kategoriya} kategoriya" if doctor.kategoriya else "",
            doctor.klinika.hospital_name if doctor.klinika else "",
            doctor.number or "",
            float(doctor.previous_debt) if doctor.previous_debt else 0.0
        ]
        ws.append(row)

        # Apply border to data cells
        for cell in ws[ws.max_row]:
            cell.border = border

    # Auto-adjust column widths
    column_widths = {
        'A': 6,   # №
        'B': 12,  # Barkod
        'C': 15,  # Bölgə
        'D': 15,  # Şəhər
        'E': 25,  # Ad
        'F': 18,  # Sonuncu Ödəniş Tarixi
        'G': 20,  # Sonuncu Ödəniş Məbləği
        'H': 12,  # İxtisas
        'I': 15,  # Dərəcə
        'J': 12,  # Kateqoriya
        'K': 20,  # Klinika
        'L': 15,  # Əlaqə
        'M': 12,  # Yekun Borc
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Create filename
    filename = f"Həkimlər_siyahısı_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    return response


def create_doctor(request):
    regions = Region.objects.all()
    hospitals = Hospital.objects.all()
    cities = City.objects.all()  
    
    if request.method == "POST":
        ad = request.POST.get("ad")
        ixtisas = request.POST.get("ixtisas")
        kategoriya = request.POST.get("kategoriya")
        derece = request.POST.get("derece")
        cinsiyyet = request.POST.get("cinsiyyet")
        bolge_id = request.POST.get("bolge_id")    
        city_id = request.POST.get("city_id")  # yeni əlavə etdik
        klinika_id = request.POST.get("klinika_id")
        number = request.POST.get("number")

        if not all([ad, ixtisas, kategoriya, bolge_id, klinika_id]):
            messages.error(request, "Zəhmət olmasa bütün vacib sahələri doldurun.")
            return redirect("add-doctor")

        bolge = get_object_or_404(Region, id=bolge_id)
        klinika = get_object_or_404(Hospital, id=klinika_id)
        city = City.objects.filter(id=city_id).first() if city_id else None


        
        doctor = Doctors.objects.create(
            ad=ad,
            ixtisas=ixtisas,
            kategoriya=kategoriya,
            derece=derece or 'II',
            cinsiyyet=cinsiyyet or 'Kişi',
            bolge=bolge,
            city=city,
            klinika=klinika,
            number=number,

        )


        messages.success(request, "Həkim uğurla əlavə edildi.")
        return redirect("doctor_detail", doctor_id=doctor.id)

    context = {
        "regions": regions,
        "hospitals": hospitals,
        "cities": cities,
    }
    return render(request, "crud/add-doctor.html", context)


def get_hospitals_by_region(request):
    region_id = request.GET.get("region_id")
    hospitals = Hospital.objects.filter(region_net_id=region_id).values("id", "hospital_name")
    return JsonResponse({"hospitals": list(hospitals)})

def get_cities_by_region(request):
    region_id = request.GET.get("region_id")
    cities = City.objects.filter(region_id=region_id)

    city_list = [
        {
            "id": city.id,
            "city_name": city.get_city_name_display()  # Model metodunu burada çağıra bilərsən
        }
        for city in cities
    ]

    return JsonResponse({"cities": city_list})




def del_all(request):
    Doctors.objects.all().delete()
    messages.success(request, "Bütün həkimlər uğurla silindi.")
    return redirect("doctors")  # Silindikdən sonra uyğun səhifəyə yönləndir


def doctor_detail(request, doctor_id):
    doctor = get_object_or_404(Doctors, id=doctor_id)
    
    
    payments = Payment_doctor.objects.filter(doctor=doctor).order_by("-date")[:10]
    recipe = RecipeDrug.objects.filter(recipe__dr=doctor).select_related(
        "recipe", "drug", "recipe__region"
    )

    recipe_date_from = ""
    recipe_date_to = ""
    fd = request.GET.get("recipe_date_from", "").strip()
    td = request.GET.get("recipe_date_to", "").strip()
    if fd:
        try:
            df = date.fromisoformat(fd)
            recipe = recipe.filter(recipe__date__gte=df)
            recipe_date_from = fd
        except ValueError:
            pass
    if td:
        try:
            dt = date.fromisoformat(td)
            recipe = recipe.filter(recipe__date__lte=dt)
            recipe_date_to = td
        except ValueError:
            pass

    recipe = recipe.order_by("recipe__date") 


    # Əgər resept modeli varsa
    recibe_total = RecipeDrug.objects.filter(recipe__dr=doctor).aggregate(total=Sum('number'))['total'] or 0


    monthly_reports = MonthlyDoctorReport.objects.filter(doctor=doctor).order_by('-report_month')
    silinme_list = []
    for report in monthly_reports:
        silinme_list.append({
            "month": report.report_month.strftime("%B %Y"),  
            "hekimden_silinen": report.hekimden_silinen
        })



    regions = Region.objects.all().order_by("region_name")

    context = {
        "doctor": doctor,
        "payments": payments,
        "recibe_total": recibe_total,
        "recipe": recipe,
        "silinme_list": silinme_list,
        "regions": regions,
        "recipe_date_from": recipe_date_from,
        "recipe_date_to": recipe_date_to,
    }
    return render(request, "doctor-details.html", context)

def ajax_doctors_by_region(request):
    region_id = request.GET.get('region_id')
    if region_id:
        doctors = Doctors.objects.filter(bolge=region_id).values('id', 'ad')
        doctors_list = list(doctors)
        return JsonResponse({'doctors': doctors_list})
    else:
        return JsonResponse({'doctors': []})




def icaze_var(il, ay, region_id):
    from datetime import date
    cari_tarix = date.today()
    cari_il = cari_tarix.year
    cari_ay = cari_tarix.month

    hesabat_baglidir = MonthlyDoctorReport.objects.filter(
        report_month__year=il,
        report_month__month=ay,
        region_id=region_id
    ).exists()

    # Əlavə şərt: cari ay üçün hesabat hələ açılmayıb, 
    # amma keçən ay bağlıdırsa → cari aya icazə ver
    if (il, ay) == (cari_il, cari_ay) and not hesabat_baglidir:
        kecen_ay = cari_ay - 1
        kecen_il = cari_il
        if kecen_ay == 0:
            kecen_ay = 12
            kecen_il -= 1

        kecen_ay_baglidir = MonthlyDoctorReport.objects.filter(
            report_month__year=kecen_il,
            report_month__month=kecen_ay,
            region_id=region_id
        ).exists()

        if kecen_ay_baglidir:
            return True  

    if hesabat_baglidir:
        return (il, ay) >= (cari_il, cari_ay)
    else:
        return (il, ay) < (cari_il, cari_ay)


from .utils import fix_recipe_drug_sequence


def _doctor_display_name(pk):
    if pk is None or pk == "":
        return ""
    try:
        ad = Doctors.objects.filter(pk=int(pk)).values_list("ad", flat=True).first()
        return ad or ""
    except (ValueError, TypeError):
        return ""


def create_recipe(request):
    regions = Region.objects.all().order_by("region_name")
    drugs = Medical.objects.all().order_by('id')
    last_recipes = RecipeDrug.objects.all().order_by("-created_at", "-id")[:5]

    selected_region = ""
    selected_doctor = ""
    selected_date = ""
    doctors = Doctors.objects.none()

    if request.method == "POST":
        region_id = request.POST.get("region", "")
        doctor_id = request.POST.get("doctor", "")
        date_str = request.POST.get("date", "")

        selected_region = region_id
        selected_doctor = doctor_id
        selected_date = date_str

        # Tarixi parse et
        try:
            istifade_olunacaq_tarix = date.fromisoformat(date_str)
        except ValueError:
            messages.error(request, "Zəhmət olmasa düzgün tarix seçin.")
            doctors = Doctors.objects.filter(bolge_id=region_id).order_by("id") if region_id else Doctors.objects.none()
            return render(request, "crud/add-recipe.html", {
                "regions": regions,
                "doctors": doctors,
                "drugs": drugs,
                "selected_region": selected_region,
                "selected_doctor": selected_doctor,
                "selected_doctor_name": _doctor_display_name(selected_doctor),
                "selected_date": selected_date
            })

        ay = istifade_olunacaq_tarix.month
        il = istifade_olunacaq_tarix.year

        # İcazəni yoxla
        if not icaze_var(il, ay, region_id):
            messages.error(request, f"{istifade_olunacaq_tarix.strftime('%Y-%m')} ayı üçün əlavə etməyə icazə yoxdur.")
            doctors = Doctors.objects.filter(bolge_id=region_id) if region_id else Doctors.objects.none()
            return render(request, "crud/add-recipe.html", {
                "regions": regions,
                "doctors": doctors,
                "drugs": drugs,
                "selected_region": selected_region,
                "selected_doctor": selected_doctor,
                "selected_doctor_name": _doctor_display_name(selected_doctor),
                "selected_date": selected_date
            })

        # Həkimləri göstər
        doctors = Doctors.objects.filter(bolge_id=region_id) if region_id else Doctors.objects.none()

        # Region və həkim yoxdursa
        if not (region_id and doctor_id):
            messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
            return render(request, "crud/add-recipe.html", {
                "regions": regions,
                "doctors": doctors,
                "drugs": drugs,
                "selected_region": selected_region,
                "selected_doctor": selected_doctor,
                "selected_doctor_name": _doctor_display_name(selected_doctor),
                "selected_date": selected_date
            })

        # Recipe yarat - SEQUENCE PROBLEMİNƏ QARŞI QORUMA İLƏ
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                with transaction.atomic():
                    # Recipe yarat
                    recipe = Recipe.objects.create(
                        region_id=region_id,
                        dr_id=doctor_id,
                        date=istifade_olunacaq_tarix
                    )
                    doctor = Doctors.objects.get(id=doctor_id)

                    # Əlavə olunan dərmanlar
                    for key in request.POST:
                        if key.startswith("quantity_"):
                            drug_id = key.split("_")[1]
                            count = request.POST.get(key)
                            if count and float(count) > 0:
                                RecipeDrug.objects.create(
                                    recipe=recipe,
                                    drug_id=drug_id,
                                    number=count
                                )

                    messages.success(request, f"{doctor.ad} həkimə resept {selected_date} tarixi ilə əlavə olundu.")
                    break  # Uğurlu oldu, döngüdən çıx

            except IntegrityError as e:
                retry_count += 1
                if 'duplicate key' in str(e) and 'doctors_recipedrug_pkey' in str(e):
                    # Sequence problemi - sıfırla və yenidən cəhd et
                    if retry_count < max_retries:
                        fix_recipe_drug_sequence()  # ← İndi utils-dən gəlir
                        continue
                    else:
                        messages.error(request, "Texniki xəta. Zəhmət olmasa bir neçə dəqiqədən sonra yenidən cəhd edin.")
                        return redirect('create_recipe')
                else:
                    # Digər IntegrityError
                    messages.error(request, f"Verilənlər bazası xətası: {str(e)}")
                    return redirect('create_recipe')
                    
            except Exception as e:
                messages.error(request, f"Gözlənilməz xəta: {str(e)}")
                return redirect('create_recipe')

    else:
        doctors = Doctors.objects.none()

    return render(request, "crud/add-recipe.html", {
        "regions": regions,
        "doctors": doctors,
        "last_recipes": last_recipes,
        "drugs": drugs,
        "selected_region": selected_region,
        "selected_doctor": selected_doctor,
        "selected_doctor_name": _doctor_display_name(selected_doctor),
        "selected_date": selected_date
    })



def create_detail_recipe(request):
    regions = Region.objects.all().order_by("region_name")
    drugs = Medical.objects.all().order_by("med_full_name")

    selected_region = ""
    selected_doctor = None
    selected_date = ""

    # GET sorğusundan doctor_id al
    doctor_id = request.GET.get("doctor_id")
    if doctor_id:
        try:
            selected_doctor = Doctors.objects.get(id=doctor_id)
            selected_region = selected_doctor.bolge.id
        except Doctors.DoesNotExist:
            selected_doctor = None

    if request.method == "POST":
        region_id = request.POST.get("region", "")
        doctor_id = request.POST.get("doctor", "")
        date = request.POST.get("date", "")

        selected_region = region_id
        selected_doctor = Doctors.objects.get(id=doctor_id) if doctor_id else None
        selected_date = date

        # Regiona görə həkimləri filtre et
        if selected_region:
            doctors = Doctors.objects.filter(bolge_id=selected_region).order_by("ad")
        else:
            doctors = Doctors.objects.none()

        if not (region_id and doctor_id and date):
            messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
            return render(request, "crud/add-details-recipe.html", {
                "regions": regions,
                "doctors": doctors,
                "drugs": drugs,
                "selected_region": selected_region,
                "selected_doctor": selected_doctor,
                "selected_date": selected_date
            })

        # Recipe yarat
        recipe = Recipe.objects.create(
            region_id=region_id,
            dr_id=doctor_id,
            date=date
        )
        doctor = Doctors.objects.get(id=doctor_id)

        # Əlavə olunan dərmanları qeyd et
        for key in request.POST:
            if key.startswith("quantity_"):
                drug_id = key.split("_")[1]
                count = request.POST.get(key)
                if count and int(count) > 0:
                    RecipeDrug.objects.create(
                        recipe=recipe,
                        drug_id=drug_id,
                        number=count
                    )

        messages.success(request, f"{doctor.ad} həkimə resept uğurla əlavə olundu.")

    else:
        # GET zamanı regiona görə həkimləri seçilmiş həkim varsa filtre et
        if selected_region:
            doctors = Doctors.objects.filter(bolge_id=selected_region).order_by("ad")
        else:
            doctors = Doctors.objects.none()

    return render(request, "crud/add-details-recipe.html", {
        "regions": regions,
        "doctors": doctors,
        "drugs": drugs,
        "selected_region": selected_region,
        "selected_doctor": selected_doctor,
        "selected_date": selected_date
    })



def del_recipe(request, id):
    rm_recipe = get_object_or_404(RecipeDrug, id=id)
    
    # Silinmə əməliyyatını loglamaq
    DeletedRecipeDrugLog.objects.create(
        drug_name=rm_recipe.drug,
        recipe_id=rm_recipe.recipe.id,
        deleted_by=request.user if request.user.is_authenticated else None,
    )
    
    # Həkimin ID-sini al
    doctor_id = rm_recipe.recipe.dr.id
    
    # Resepti sil
    rm_recipe.delete()
    
    next_url = request.GET.get('next')
    if next_url == 'create_recipe':
        return redirect('create_recipe')
    return redirect('doctor_detail', doctor_id=doctor_id)


@require_POST
def bulk_del_recipe_lines(request, doctor_id):
    doctor = get_object_or_404(Doctors, id=doctor_id)
    raw_ids = request.POST.getlist("recipe_ids")
    id_list = []
    for x in raw_ids:
        try:
            id_list.append(int(x))
        except (TypeError, ValueError):
            continue

    if not id_list:
        messages.warning(request, "Silinəcək resept sətiri seçilməyib.")
        return redirect('doctor_detail', doctor_id=doctor.id)

    lines = list(
        RecipeDrug.objects.filter(
            pk__in=id_list,
            recipe__dr_id=doctor.pk,
        ).select_related("recipe", "drug")
    )

    if not lines:
        messages.warning(request, "Seçilmiş sətirlər tapılmadı və ya bu həkimə aid deyil.")
        return redirect('doctor_detail', doctor_id=doctor.id)

    with transaction.atomic():
        DeletedRecipeDrugLog.objects.bulk_create(
            [
                DeletedRecipeDrugLog(
                    drug_name=str(rd.drug),
                    recipe_id=rd.recipe_id,
                    deleted_by=request.user if request.user.is_authenticated else None,
                )
                for rd in lines
            ]
        )
        RecipeDrug.objects.filter(
            pk__in=[rd.pk for rd in lines],
            recipe__dr_id=doctor.pk,
        ).delete()

    messages.success(request, f"{len(lines)} resept sətiri silindi.")
    return redirect('doctor_detail', doctor_id=doctor.id)


def del_payments(request, id):
    payment = get_object_or_404(Payment_doctor, id=id)
    doctor_id = payment.doctor.id
    payment.delete()
    return redirect('doctor_detail', doctor_id=doctor_id)




def update_recipe(request, id):
    """Resept sətirini (RecipeDrug) və əsas reseptin tarixini/bölgəsini yenilə."""
    rd = get_object_or_404(
        RecipeDrug.objects.select_related("recipe__dr", "recipe__region", "drug"),
        id=id,
    )
    rp = rd.recipe
    regions = Region.objects.all().order_by("region_name")

    if request.method == "POST":
        number_raw = (request.POST.get("number") or "").strip().replace(",", ".")
        date_str = (request.POST.get("date") or "").strip()
        region_id_raw = (request.POST.get("region_id") or "").strip()

        err = []
        try:
            num = Decimal(number_raw)
            if num <= 0:
                err.append("Say 0-dan böyük olmalıdır.")
        except InvalidOperation:
            err.append("Düzgün say daxil edin.")

        parsed_date = None
        try:
            parsed_date = date.fromisoformat(date_str)
        except ValueError:
            err.append("Düzgün tarix seçin.")

        if not region_id_raw.isdigit():
            err.append("Bölgə seçin.")
        elif not Region.objects.filter(pk=int(region_id_raw)).exists():
            err.append("Bölgə tapılmadı.")

        is_ajax = (
            request.POST.get("ajax") == "1"
            or request.headers.get("X-Requested-With") == "XMLHttpRequest"
        )

        if err:
            if is_ajax:
                return JsonResponse({"ok": False, "errors": err}, status=400)
            for m in err:
                messages.error(request, m)
        else:
            num = num.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
            rp.date = parsed_date
            rp.region_id = int(region_id_raw)
            rd.number = num
            with transaction.atomic():
                rp.save(update_fields=("date", "region_id"))
                rd.save(update_fields=("number",))
            ok_msg = f"{rd.drug.med_name} — resept sətiri yeniləndi."
            if is_ajax:
                return JsonResponse(
                    {
                        "ok": True,
                        "redirect_url": reverse("doctor_detail", args=[rp.dr_id]),
                        "message": ok_msg,
                    }
                )
            messages.success(request, ok_msg)
            return redirect("doctor_detail", doctor_id=rp.dr_id)

    return render(
        request,
        "crud/update_recipe.html",
        {
            "rd": rd,
            "regions": regions,
        },
    )



def load_cities_hospitals(request):
    region_id = request.GET.get('region_id')
    if not region_id or not region_id.isdigit():
        return JsonResponse({'cities': [], 'hospitals': []})

    cities = list(City.objects.filter(region_id=region_id).values('id', 'city_name'))
    hospitals = list(Hospital.objects.filter(region_net_id=region_id).values('id', 'hospital_name'))
    return JsonResponse({'cities': cities, 'hospitals': hospitals})




def update_doctor(request, pk):
    doctor = get_object_or_404(Doctors, pk=pk)
    if request.method == "POST":
        doctor.ad = request.POST.get('ad', doctor.ad)
        doctor.ixtisas = request.POST.get('ixtisas', doctor.ixtisas)
        doctor.kategoriya = request.POST.get('kategoriya', doctor.kategoriya)
        doctor.cinsiyyet = request.POST.get('cinsiyyet', doctor.cinsiyyet)
        doctor.derece = request.POST.get('derece', doctor.derece)
        doctor.number = request.POST.get('number', doctor.number)
        doctor.is_active = request.POST.get('is_active') == '1'

        # Region, city və klinika yenilənməsi
        bolge_id = request.POST.get('region_id')
        city_id = request.POST.get('city_id')
        klinika_id = request.POST.get('hospital_id')

        if bolge_id:
            try:
                doctor.bolge = Region.objects.get(id=bolge_id)
            except Region.DoesNotExist:
                pass

        if city_id:
            try:
                doctor.city = City.objects.get(id=city_id)
            except City.DoesNotExist:
                doctor.city = None
        else:
            doctor.city = None

        if klinika_id:
            try:
                doctor.klinika = Hospital.objects.get(id=klinika_id)
            except Hospital.DoesNotExist:
                pass

        doctor.save()
        return redirect('doctor_detail', doctor_id=doctor.pk)
    return render(request, 'doctor-details.html', {'doctor': doctor})


def create_real_sales(request):
    regions = Region.objects.all().order_by("region_name")
    doctors = Doctors.objects.all().order_by("id")
    drugs = Medical.objects.all().order_by('id')

    selected_region = None
    selected_doctor = None
    selected_date = None

    if request.method == "POST":
        region_id = request.POST.get("region")
        doctor_id = request.POST.get("doctor")
        date = request.POST.get("date")

        selected_region = region_id
        selected_doctor = doctor_id
        selected_date = date

        if not (region_id and doctor_id and date):
            messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
            return render(request, "crud/add-real-sales.html", {
                "regions": regions,
                "doctors": doctors,
                "drugs": drugs,
                "selected_region": selected_region,
                "selected_doctor": selected_doctor,
                "selected_date": selected_date
            })

        # Dərmanların olub-olmadığını yoxlayaq
        selected_drugs = []
        for key in request.POST:
            if key.startswith("quantity_"):
                drug_id = key.split("_")[1]
                count = request.POST.get(key)
                if count and int(count) > 0:
                    selected_drugs.append((int(drug_id), int(count)))

        if not selected_drugs:
            messages.error(request, "Zəhmət olmasa ən az bir dərman miqdarı daxil edin.")
            return render(request, "crud/add-real-sales.html", {
                "regions": regions,
                "doctors": doctors,
                "drugs": drugs,
                "selected_region": selected_region,
                "selected_doctor": selected_doctor,
                "selected_date": selected_date
            })

        # Real satış yaradılır
        real_sale = RealSales.objects.create(
            region_n_id=region_id,
            dr_name_id=doctor_id,
            date_sale=date
        )
        
        total_commission = Decimal('0')
        total_quantity = 0  # Real satış miqdarını burada topluyuruq

        for drug_id, count in selected_drugs:
            RealSalesDrug.objects.create(
                real_sale=real_sale,
                drug_name_id=drug_id,
                numbers=count
            )


            # Həkimin reseptindən də azaldırıq (tarixi köhnədən yeniyə)
            recipe_drugs = RecipeDrug.objects.filter(
                recipe__region_id=region_id,
                recipe__dr_id=doctor_id,
                drug_id=drug_id
            ).order_by('recipe__date')
            
            remaining_to_subtract = count
            for rd in recipe_drugs:
                if remaining_to_subtract <= 0:
                    break
                
                if rd.number >= remaining_to_subtract:
                    rd.number -= remaining_to_subtract
                    remaining_to_subtract = 0
                else:
                    remaining_to_subtract -= rd.number
                    rd.number = 0
                
                rd.save()

            # Komissiyanı hesabla
            drug = Medical.objects.get(id=drug_id)
            komissiya = drug.komissiya * count
            total_commission += komissiya
            
            # Real satış miqdarını topluyuruq
            total_quantity += count

        # Həkimin miqdarını və pulunu artırırıq
        doctor = Doctors.objects.get(id=doctor_id)
        

        
        # Komissiyanı (məbləği) həkimdən silinənə əlavə edirik
        doctor.hekimden_silinen += total_commission  # MƏBLƏĞ (manat)
        
        doctor.save()

        messages.success(request, "Satış uğurla əlavə olundu. Həkimin miqdarı və komissiyası yeniləndi.")
        return redirect("create_real_sales")

    return render(request, "crud/add-real-sales.html", {
        "regions": regions,
        "doctors": doctors,
        "drugs": drugs,
        "selected_region": selected_region,
        "selected_doctor": selected_doctor,
        "selected_date": selected_date
    })

# def create_real_sales(request):
#     regions = Region.objects.all().order_by("region_name")
#     doctors = Doctors.objects.all().order_by("id")
#     drugs = Medical.objects.all().order_by('id')

#     selected_region = None
#     selected_doctor = None
#     selected_date = None

#     if request.method == "POST":
#         region_id = request.POST.get("region")
#         doctor_id = request.POST.get("doctor")
#         date = request.POST.get("date")

#         selected_region = region_id
#         selected_doctor = doctor_id
#         selected_date = date

#         if not (region_id and doctor_id and date):
#             messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
#             return render(request, "crud/add-real-sales.html", {
#                 "regions": regions,
#                 "doctors": doctors,
#                 "drugs": drugs,
#                 "selected_region": selected_region,
#                 "selected_doctor": selected_doctor,
#                 "selected_date": selected_date
#             })

#         # Dərmanların olub-olmadığını yoxlayaq
#         selected_drugs = []
#         for key in request.POST:
#             if key.startswith("quantity_"):
#                 drug_id = key.split("_")[1]
#                 count = request.POST.get(key)
#                 if count and int(count) > 0:
#                     selected_drugs.append((int(drug_id), int(count)))

#         if not selected_drugs:
#             messages.error(request, "Zəhmət olmasa ən az bir dərman miqdarı daxil edin.")
#             return render(request, "crud/add-real-sales.html", {
#                 "regions": regions,
#                 "doctors": doctors,
#                 "drugs": drugs,
#                 "selected_region": selected_region,
#                 "selected_doctor": selected_doctor,
#                 "selected_date": selected_date
#             })

#         # Real satış yaradılır
#         real_sale = RealSales.objects.create(
#             region_n_id=region_id,
#             dr_name_id=doctor_id,
#             date_sale=date
#         )
        
#         total_commission = Decimal('0')

#         for drug_id, count in selected_drugs:
#             RealSalesDrug.objects.create(
#                 real_sale=real_sale,
#                 drug_name_id=drug_id,
#                 numbers=count
#             )

#             # Bölgə satışını azaldırıq
#             sale_qs = Sale.objects.filter(region_id=region_id, drug_id=drug_id).first()
#             if sale_qs:
#                 sale_qs.quantity = max(0, sale_qs.quantity - count)
#                 sale_qs.save()

#             # Həkimin reseptindən də azaldırıq (tarixi köhnədən yeniyə)
#             recipe_drugs = RecipeDrug.objects.filter(
#                 recipe__region_id=region_id,
#                 recipe__dr_id=doctor_id,
#                 drug_id=drug_id
#             ).order_by('recipe__date')
            
#             remaining_to_subtract = count
#             for rd in recipe_drugs:
#                 if remaining_to_subtract <= 0:
#                     break
                
#                 if rd.number >= remaining_to_subtract:
#                     rd.number -= remaining_to_subtract
#                     remaining_to_subtract = 0
#                 else:
#                     remaining_to_subtract -= rd.number
#                     rd.number = 0
                
#                 rd.save()

#             # Komissiyanı hesabla
#             drug = Medical.objects.get(id=drug_id)
#             komissiya = drug.komissiya * count
#             total_commission += komissiya

#         # Həkimin miqdarını və pulunu artırırıq
#         doctor = Doctors.objects.get(id=doctor_id)
        
#         # Real satış QUTU sayını hesablanan miqdarına əlavə edirik
#         total_quantity = sum(count for _, count in selected_drugs)
#         doctor.hesablanan_miqdar += total_quantity # QUTU sayı
        
#         # Komissiyanı (məbləği) həkimdən silinənə əlavə edirik
#         doctor.hekimden_silinen += total_commission  # MƏBLƏĞ (manat)
        
#         doctor.save()

#         messages.success(request, "Satış uğurla əlavə olundu. Həkimin miqdarı və komissiyası yeniləndi.")
#         return redirect("create_real_sales")

#     return render(request, "crud/add-real-sales.html", {
#         "regions": regions,
#         "doctors": doctors,
#         "drugs": drugs,
#         "selected_region": selected_region,
#         "selected_doctor": selected_doctor,
#         "selected_date": selected_date
#     })

def create_datasiya(request):
    regions = Region.objects.all().order_by("region_name")
    selected_region = request.GET.get("region") or None

    doctors = Doctors.objects.filter(region_id=selected_region).order_by("ad") if selected_region else []

    if request.method == "POST":
        region_id = request.POST.get("region")
        date = request.POST.get("date")

        if not region_id or not date:
            messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
            return render(request, "crud/add-datasiya.html", {
                "regions": regions,
                "doctors": doctors,
                "selected_region": region_id,
                "selected_date": date,
            })

        # Həkimlərin borcunu yenilə
        for key, value in request.POST.items():
            if key.startswith("given_") or key.startswith("received_"):
                parts = key.split("_")
                prefix = parts[0]
                doctor_id = parts[1]

                if value.strip() == "":
                    continue
                try:
                    amount = Decimal(value)
                except ValueError:
                    amount = 0

                if amount == 0:
                    continue

                try:
                    doctor = Doctors.objects.get(id=doctor_id)
                except Doctors.DoesNotExist:
                    continue

                if prefix == "given":
                    doctor.datasiya += amount
                elif prefix == "received":
                    doctor.datasiya -= amount

                doctor.save()

        messages.success(request, "Datasiya uğurla əlavə etdiniz.")
        return redirect("datasiya")

    return render(request, "crud/add-datasiya.html", {
        "regions": regions,
        "doctors": doctors,
        "selected_region": selected_region,
    })

def finance_view(request):
    regions = Region.objects.all().order_by("region_name")
    selected_region = request.GET.get("region")

    # HTML input type="month" → "2024-11" formatında gəlir
    selected_month_val = request.GET.get("month")

    today = date.today()
    current_month = today.month
    current_year = today.year

    # Ay və ili düzgün parse edirik
    if selected_month_val:
        try:
            year, month = selected_month_val.split("-")
            selected_year = int(year)
            selected_month = int(month)
        except Exception:
            selected_year = current_year
            selected_month = current_month
    else:
        # Heç bir ay seçilməyibsə, default olaraq keçən ayı göstər
        if current_month == 1:
            selected_month = 12
            selected_year = current_year - 1
        else:
            selected_month = current_month - 1
            selected_year = current_year
        selected_month_val = f"{selected_year:04d}-{selected_month:02d}"

    doctors = []

    if selected_region:
        doctors = Doctors.objects.filter(region_id=selected_region).order_by("id")

        for doctor in doctors:
            payments = Payment_doctor.objects.filter(
                doctor=doctor,
                area_id=selected_region,
                date__month=selected_month,
                date__year=selected_year
            ).values("payment_type").annotate(total=Sum("pay"))

            doctor.avans = Decimal("0.00")
            doctor.investisiya = Decimal("0.00")

            for p in payments:
                if p["payment_type"] == "Avans":
                    doctor.avans = p["total"] or Decimal("0.00")
                elif p["payment_type"] == "İnvest":
                    doctor.investisiya = p["total"] or Decimal("0.00")

            # Excel və template üçün əvvəlki borc dəyəri (əgər None-dursa 0.00 olsun)
            doctor.previous_debt = doctor.previous_debt or Decimal("0.00")

    return render(request, "finance.html", {
        "regions": regions,
        "doctors": doctors,
        "selected_region": selected_region,
        "selected_month": selected_month_val,  # seçilmiş (və ya default) ayı template-ə göndəririk
    })

# views.py
def finance_export_excel(request):
    selected_region = request.GET.get("region")
    selected_month_val = request.GET.get("month")

    if not selected_region:
        return HttpResponse("Bölgə seçilməyib", status=400)

    today = date.today()
    current_year = today.year
    current_month = today.month

    # "2024-11" formatını parçalayırıq
    if selected_month_val:
        try:
            year, month = selected_month_val.split("-")
            selected_year = int(year)
            selected_month = int(month)
        except Exception:
            selected_year = current_year
            selected_month = current_month
    else:
        # Heç bir ay seçilməyibsə, default olaraq keçən ayı göstər
        if current_month == 1:
            selected_month = 12
            selected_year = current_year - 1
        else:
            selected_month = current_month - 1
            selected_year = current_year
        selected_month_val = f"{selected_year:04d}-{selected_month:02d}"

    try:
        selected_region = int(selected_region)
    except ValueError:
        return HttpResponse("Yanlış məlumat", status=400)

    doctors = Doctors.objects.filter(bolge_id=selected_region).order_by("id")
    export_data = []

    for doctor in doctors:
        payments = Payment_doctor.objects.filter(
            doctor=doctor,
            area_id=selected_region,
            date__month=selected_month,
            date__year=selected_year
        ).values("payment_type").annotate(total=Sum("pay"))

        avans = Decimal("0.00")
        investisiya = Decimal("0.00")

        for p in payments:
            if p["payment_type"] == "Avans":
                avans = p["total"] or Decimal("0.00")
            elif p["payment_type"] == "İnvest":
                investisiya = p["total"] or Decimal("0.00")

        # Əvvəlki borc sahəsi (None olarsa 0.00)
        previous_debt = doctor.previous_debt or Decimal("0.00")

        if avans > 0 or investisiya > 0 or previous_debt != 0:
            export_data.append({
                "doctor_name": doctor.ad,
                "previous_debt": previous_debt,
                "avans": avans,
                "investisiya": investisiya,
                "bolge": doctor.bolge.region_name if doctor.bolge else ""
            })

    # Excel yaradılması
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Report"

    headers = ["Bölgə", "Həkim", "Əvvəlki borc", "Avans", "İnvestisiya"]
    ws.append(headers)

    for item in export_data:
        ws.append([
            item["bolge"],
            item["doctor_name"],
            float(item["previous_debt"]),
            float(item["avans"]),
            float(item["investisiya"])
        ])

    region_name = export_data[0]["bolge"] if export_data else "Report"
    filename = f"{region_name} maliyəsi {selected_month}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    wb.save(response)
    return response



def create_razilasma(request):
    regions = Region.objects.all().order_by("region_name")
    selected_region = request.GET.get("region") or None
    doctors = Doctors.objects.filter(region_id=selected_region).order_by("id") if selected_region else []

    if request.method == "POST":
        region_id = request.POST.get("region")
        date = request.POST.get("date")

        if not region_id or not date:
            messages.error(request, "Zəhmət olmasa bütün sahələri doldurun.")
            return render(request, "crud/add-razılaşma.html", {
                "regions": regions,
                "doctors": doctors,
                "selected_region": region_id,
                "selected_date": date,
            })

        # Həkimlərin razılaşma sayını yenilə
        for key, value in request.POST.items():
            if key.startswith("razilasma_"):
                doctor_id = key.split("_")[1]
                if value.strip() == "":
                    continue

                try:
                    count = int(value)
                except ValueError:
                    count = 0

                if count == 0:
                    continue

                try:
                    doctor = Doctors.objects.get(id=doctor_id)
                except Doctors.DoesNotExist:
                    continue

                # Burada həkimin razılaşma sayını saxlayırsan
                doctor.razılaşma += count  
                doctor.save()

        messages.success(request, "Razılaşma uğurla əlavə edildi.")
        return redirect("razilasma")

    return render(request, "crud/add-razılaşma.html", {
        "regions": regions,
        "doctors": doctors,
        "selected_region": selected_region,
    })


def ajax_doctors_by_region(request):
    region_id = request.GET.get("region_id")
    month_str = request.GET.get("month")  # "2024-11" kimi gəlir

    if not region_id:
        return JsonResponse({"doctors": []})

    # Ay və ili müəyyən et
    today = date.today()
    current_month = today.month
    current_year = today.year

    if month_str:
        try:
            year, month = month_str.split("-")
            selected_year = int(year)
            selected_month = int(month)
        except (ValueError, AttributeError):
            selected_year = current_year
            selected_month = current_month
    else:
        # fallback: cari ay və il
        selected_year = current_year
        selected_month = current_month

    # Həkimləri və onların ödəniş məlumatlarını al
    doctors = Doctors.objects.filter(bolge=region_id).order_by("id")

    doctor_list = []

    for doctor in doctors:
        # Bu həkim üçün seçilmiş ayın Avans və İnvestisiya məlumatlarını hesabla
        payments = Payment_doctor.objects.filter(
            doctor=doctor,
            area_id=region_id,
            date__month=selected_month,
            date__year=selected_year,
        ).values("payment_type").annotate(total=Sum("pay"))

        # Default dəyərlər
        avans = Decimal("0.00")
        investisiya = Decimal("0.00")

        # Ödəniş növlərinə görə cəmlə
        for payment in payments:
            if payment["payment_type"] == "Avans":
                avans = payment["total"] or Decimal("0.00")
            elif payment["payment_type"] == "İnvest":
                investisiya = payment["total"] or Decimal("0.00")

        doctor_list.append(
            {
                "id": doctor.id,
                "ad": doctor.ad,
                "previous_debt": float(doctor.previous_debt) if doctor.previous_debt else 0.0,
                "avans": float(avans),
                "investisiya": float(investisiya),
            }
        )

    return JsonResponse({"doctors": doctor_list})

def data_list(request):
    region = Region.objects.all()
    drug = Medical.objects.all().order_by('id')

    context = {
        "region": region,
        "drug": drug
    }
    return render(request, "data-list.html", context)

def ajax_region_data(request):
    region_id = request.GET.get("region_id")
    date_range = request.GET.get("date_range")
    name_filter = request.GET.get("name_filter")
    month = request.GET.get("month")
    search = request.GET.get("search")  # ƏLAVƏ EDİLDİ
    page = request.GET.get("page", 1)
    per_page = 30

    try:
        region_id = int(region_id)
    except (TypeError, ValueError):
        return JsonResponse({"results": []})

    # Əsas queryset – arxivlənmiş həkimlər göstərilməsin
    doctors = Doctors.objects.filter(bolge=region_id, is_active=True).order_by("id")

    # AXTARIŞ FUNKSİONALLIĞI - ƏLAVƏ EDİLDİ
    if search:
        doctors = doctors.filter(
            Q(ad__icontains=search) |
            Q(barkod__icontains=search) |
            Q(bolge__region_name__icontains=search)
        )


    # Name filter
    if name_filter == 'with_dannisi':
        doctors = doctors.filter(
            Q(ad__icontains='dannı') | Q(ad__icontains='dannisi') | Q(ad__icontains='dannısı')
        )
    elif name_filter == 'without_dannisi':
        doctors = doctors.exclude(
            Q(ad__icontains='dannı') | Q(ad__icontains='dannisi') | Q(ad__icontains='dannısı')
        )

    # Tarix aralığı və ay filteri
    month_start = month_end = None
    dr_start = dr_end = None
    
    # Əvvəlcə date_range-dən ili əldə et
    year_from_range = None
    date_range = request.GET.get("date_range", "")
    if date_range and " - " in date_range:
        try:
            start_str, end_str = date_range.split(" - ")
            dr_start = datetime.strptime(start_str.strip(), '%Y-%m-%d').date()
            dr_end = datetime.strptime(end_str.strip(), '%Y-%m-%d').date()
            year_from_range = dr_start.year  # Tarix aralığından ili götür
        except ValueError:
            pass

    # Ay filtri üçün ili təyin et
    if month:
        try:
            month_int = int(month)
            # Əgər tarix aralığı varsa ondan ili istifadə et, yoxsa cari ili
            current_year = year_from_range if year_from_range else datetime.now().year
            month_start = datetime(current_year, month_int, 1).date()
            if month_int == 12:
                month_end = datetime(current_year + 1, 1, 1).date()
            else:
                month_end = datetime(current_year, month_int + 1, 1).date()
        except ValueError:
            pass

    # Final start_date və end_date
    if month_start and month_end and dr_start and dr_end:
        # Hər ikisi varsa, kəsişən hissəni tap
        start_date = max(month_start, dr_start)
        end_date = min(month_end, dr_end)
        # Əgər tarix aralığı səhvdirsə, yalnız date_range istifadə et
        if start_date > end_date:
            start_date = dr_start
            end_date = dr_end
    elif month_start and month_end:
        start_date = month_start
        end_date = month_end
    elif dr_start and dr_end:
        start_date = dr_start
        end_date = dr_end
    else:
        start_date = end_date = None

    result = []
    all_medical_drugs = Medical.objects.all().order_by('id')

    for doctor in doctors:
        # RecipeDrug queryset
        all_drugs = RecipeDrug.objects.filter(recipe__dr=doctor, recipe__region=region_id)

        if start_date and end_date:
            all_drugs = all_drugs.filter(recipe__date__gte=start_date, recipe__date__lte=end_date)

        # Hər dərman üzrə cəmi say
        drugs_agg = all_drugs.values('drug__med_name').annotate(total_count=Sum('number'))

        drugs = []
        total = 0
        for med_drug in all_medical_drugs:
            found_drug = next((d for d in drugs_agg if d['drug__med_name'] == med_drug.med_name), None)
            count = found_drug['total_count'] if found_drug else 0
            drugs.append({"name": med_drug.med_name, "count": count})
            total += count

        # Filter: without_dannisi və total=0
        if name_filter == "without_dannisi" and total == 0:
            continue

        # Son ödəniş
        last_payment = doctor.odenisler.order_by('-date').first()

        if last_payment:
            odeme_type = last_payment.payment_type.lower()
            if odeme_type == "avans":
                odeme_class = "text-primary"
            elif odeme_type == "investisiya":
                odeme_class = "text-warning"
            elif odeme_type == "geriqaytarma":
                odeme_class = "text-danger"
            else:
                odeme_class = "text-success"

            odeme_amount = float(last_payment.pay)
            odeme_date = last_payment.date.strftime('%Y-%m-%d')
            odeme_type_json = last_payment.payment_type.lower()
        else:
            odeme_class = "text-success"
            odeme_amount = 0
            odeme_date = None
            odeme_type_json = ""

        odeme = {
            "amount": odeme_amount,
            "type": odeme_type_json,
            "class": odeme_class,
            "date": odeme_date
        }

        result.append({
            "bolge": doctor.bolge.region_name,
            "doctor": doctor.ad,
            "doctor_id": doctor.id,
            "barcode": doctor.barkod,
            "odeme": odeme,
            "borc": float(doctor.borc),
            "drugs": drugs,
            "total": float(total)
        })

    # Pagination
    paginator = Paginator(result, per_page)
    try:
        current_page = paginator.page(page)
        paginated_results = list(current_page.object_list)
    except:
        current_page = paginator.page(1)
        paginated_results = list(current_page.object_list)

    return JsonResponse({
        "results": paginated_results,
        "total_pages": paginator.num_pages,
        "current_page": current_page.number,
        "has_previous": current_page.has_previous(),
        "has_next": current_page.has_next(),
        "total_results": len(result),
        "debug_search": search  # Debug üçün
    })

def export_region_excel(request):
    region_id = request.GET.get('region_id')
    date_range = request.GET.get("date_range")
    name_filter = request.GET.get('name_filter')
    search_term = request.GET.get("search")
    month = request.GET.get("month")

    if not region_id:
        return HttpResponse("Bölgə seçilməyib.", status=400)

    try:
        region = Region.objects.get(id=region_id)
    except Region.DoesNotExist:
        return HttpResponse("Bölgə tapılmadı.", status=404)

    # 🔹 Əsas queryset – arxivlənmiş həkimlər göstərilməsin
    doctors = Doctors.objects.filter(bolge=region, is_active=True).select_related('bolge').prefetch_related('odenisler')

    # 🔹 Axtarış filteri
    if search_term:
        doctors = doctors.filter(Q(ad__icontains=search_term) | Q(barkod__icontains=search_term))

    # 🔹 Dannisi filteri və filter aktivlik flag
    if name_filter == 'with_dannisi':
        doctors = doctors.filter(Q(ad__icontains='dannısı') | Q(ad__icontains='dannisi'))
        filter_active = True
    elif name_filter == 'without_dannisi':
        doctors = doctors.exclude(Q(ad__icontains='dannısı') | Q(ad__icontains='dannisi'))
        filter_active = True
    else:
        filter_active = False

    # 🔹 Tarix intervalı
    start_date = end_date = None
    if date_range:
        try:
            if "to" in date_range:
                start_str, end_str = date_range.split("to")
            elif " - " in date_range:
                start_str, end_str = date_range.split(" - ")
            start_date = datetime.strptime(start_str.strip(), '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str.strip(), '%Y-%m-%d').date()
        except ValueError:
            pass

    # 🔹 Ay filteri
    month_int = None
    if month:
        try:
            month_int = int(month)
        except ValueError:
            month_int = None

    doctor_ids = doctors.values_list('id', flat=True)

    # RecipeDrug aggregation + tarix filteri
    counts_qs = RecipeDrug.objects.filter(
        recipe__dr__in=doctor_ids,
        recipe__region=region
    )

    if start_date and end_date:
        counts_qs = counts_qs.filter(recipe__date__range=(start_date, end_date))
    elif month_int:
        counts_qs = counts_qs.filter(recipe__date__month=month_int)

    counts_qs = counts_qs.values('recipe__dr', 'drug_id').annotate(total=Sum('number'))

    # 🔹 total>0 olan həkimləri tapmaq yalnız filter aktivdirsə
    if filter_active:
        valid_doctor_ids = set(row['recipe__dr'] for row in counts_qs if (row['total'] or 0) > 0)
        doctors = doctors.filter(id__in=valid_doctor_ids)
    else:
        valid_doctor_ids = set(doctors.values_list('id', flat=True))

    # Həkim -> dərman -> sayı mapping
    doctor_drug_counts = defaultdict(dict)
    doctor_total_counts = defaultdict(int)
    for row in counts_qs:
        dr_id = row['recipe__dr']
        drug_id = row['drug_id']
        total = row['total'] or 0
        if dr_id in valid_doctor_ids:
            doctor_drug_counts[dr_id][drug_id] = total
            doctor_total_counts[dr_id] += total

    # 🔹 Dərmanlar siyahısı
    drugs = list(Medical.objects.all().order_by('id'))

    # 📊 Excel yaradılması
    wb = Workbook()
    ws = wb.active
    ws.title = f"{region.region_name} - Filterli"

    headers = ["№", "Bölgə", "Həkim adı", "Son Ödəniş"] + [d.med_name for d in drugs] + ["Total"]
    ws.append(headers)

    bold_font = Font(bold=True, color="060411")
    header_fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
    thin = Side(style='thin', color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="bottom")

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)

    ws.freeze_panes = "A2"
    
    # Sütun enliklərini sıxlaşdır
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 12  # Bölgə
    ws.column_dimensions['C'].width = 20  # Həkim adı
    ws.column_dimensions['D'].width = 18  # Son Ödəniş
    
    # Dərman sütunları - dar et (E sütunundan başlayır)
    for i in range(len(drugs)):
        col_letter = get_column_letter(5 + i)  # 5 = E sütunu
        ws.column_dimensions[col_letter].width = 4
    
    # Total sütunu
    total_col_letter = get_column_letter(5 + len(drugs))
    ws.column_dimensions[total_col_letter].width = 6

    idx = 1
    for doctor in doctors:
        total = doctor_total_counts.get(doctor.id, 0)
        last_payment = doctor.odenisler.order_by('-date').first()

        if last_payment:
            odeme = f"₼{float(last_payment.pay):.2f} - {last_payment.date.strftime('%Y-%m-%d')}"
        elif doctor.geriqaytarma > 0:
            odeme = f"₼{float(doctor.geriqaytarma):.2f}"
        elif doctor.investisiya > 0:
            odeme = f"₼{float(doctor.investisiya):.2f}"
        elif doctor.avans > 0:
            odeme = f"₼{float(doctor.avans):.2f}"
        else:
            odeme = "-"

        row = [
            idx,
            doctor.bolge.region_name,
            doctor.ad,
            odeme,
        ]

        counts_map = doctor_drug_counts.get(doctor.id, {})
        for drug in drugs:
            row.append(counts_map.get(drug.id, 0))
        row.append(total)

        ws.append(row)
        
        # Hər xanaya border və mərkəzləşdirmə əlavə et
        current_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            
            # Total sütunu bold olsun əgər 0-dan böyükdürsə
            if col == len(headers) and cell.value and cell.value > 0:
                cell.font = Font(bold=True)

        idx += 1

    # 📈 Cəmlər
    # Sütun strukturu: A=№, B=Bölgə, C=Həkim adı, D=Son Ödəniş, E-... =Dərmanlar, Son=Total
    start_drug_col = 5  # E sütunu (dərmanlar buradan başlayır)
    num_drugs = len(drugs)
    total_col_idx = start_drug_col + num_drugs  # Total sütunu
    
    # Hər dərman üçün cəm
    drug_totals = [0] * num_drugs
    overall_total = 0

    # Data sətirlərini oxuyub cəmləri hesabla
    for row_idx in range(2, ws.max_row + 1):
        # Dərman sütunları (E-dən başlayır)
        for i, drug in enumerate(drugs):
            col_idx = start_drug_col + i
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            try:
                drug_totals[i] += int(cell_value or 0)
            except:
                pass
        
        # Total sütunu
        try:
            total_cell_value = ws.cell(row=row_idx, column=total_col_idx).value
            overall_total += int(total_cell_value or 0)
        except:
            pass

    # Cəmi sətirini yaz
    total_row_idx = ws.max_row + 1
    
    # Boş xanaları doldur və "Cəmi" yaz
    for col in range(1, 4):
        cell = ws.cell(row=total_row_idx, column=col, value="")
        cell.border = thin_border
        cell.alignment = center_align
    
    cemi_cell = ws.cell(row=total_row_idx, column=4, value="Cəmi")
    cemi_cell.font = Font(bold=True)
    cemi_cell.border = thin_border
    cemi_cell.alignment = center_align

    # Hər dərman üçün cəmi (E sütunundan başlayır)
    for i, total in enumerate(drug_totals):
        cell = ws.cell(row=total_row_idx, column=start_drug_col + i, value=total)
        cell.border = thin_border
        cell.alignment = center_align
        if total > 0:
            cell.font = Font(bold=True)

    # Ümumi total xanası
    overall_cell = ws.cell(row=total_row_idx, column=total_col_idx, value=overall_total)
    overall_cell.border = thin_border
    overall_cell.alignment = center_align
    overall_cell.font = Font(bold=True if overall_total > 0 else False)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{region.region_name} Qeydiyyatı.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"
    wb.save(response)
    return response

def get_region(request):
    
    reg = Region.objects.all().order_by('id')
    context = {
        "reg": reg
    }
    return render(request, "test.html", context)

# Borcalacaq Hesablama
from django.db.models import Max


def region_report(request, region_id):
    """Yungul sehife - yalniz formu render edir, hec bir agir SQL yoxdur."""
    region = get_object_or_404(Region, id=region_id)
    now = datetime.now()
    current_year = now.year
    current_month = now.month
    report_scope_label = "Şəhər" if (region.region_type or "") == "Şəhər" else "Bölgə"
    context = {
        "region": region,
        "report_scope_label": report_scope_label,
        "aylar": range(1, 13),
        "iller": [2025, 2026, 2027],
        "current_year": current_year,
        "current_month": current_month,
    }
    return render(request, "test_2.html", context)


def region_report_data_ajax(request, region_id):
    """AJAX endpoint - yalniz Hesabla dugmesine basanda cagrilir."""
    month = request.GET.get("month")
    year = request.GET.get("year")
    current_year = datetime.now().year

    region = get_object_or_404(Region, id=region_id)
    dermanlar = list(Medical.objects.all().order_by("-id"))
    hekimler = Doctors.objects.filter(bolge_id=region_id, is_active=True).order_by("ad")

    region_recipe_drugs = RecipeDrug.objects.filter(recipe__region_id=region_id)
    sales = Sale.objects.filter(region_id=region_id)

    try:
        ay = int(month) if month else None
    except ValueError:
        ay = None
    try:
        il = int(year) if year else current_year
    except ValueError:
        il = current_year

    if ay:
        region_recipe_drugs = region_recipe_drugs.filter(recipe__date__month=ay)
        sales = sales.filter(sale_date__month=ay)
    if il:
        region_recipe_drugs = region_recipe_drugs.filter(recipe__date__year=il)
        sales = sales.filter(sale_date__year=il)

    dereceler = {"VIP": 1.00, "I": 0.90, "II": 0.65, "III": 0.40}

    # Butun resept saylarini bir sorgu ile cek
    drug_counts = (
        region_recipe_drugs
        .values("recipe__dr_id", "drug_id")
        .annotate(total=Sum("number"))
    )
    drug_map = {}
    for row in drug_counts:
        drug_map[(row["recipe__dr_id"], row["drug_id"])] = row["total"]

    report_data = []
    for hekim in hekimler:
        faktor = Decimal(str(dereceler.get(hekim.derece, 0)))
        d_say = []
        d_faizli = []
        toplam = 0
        faizli_toplam = Decimal("0.00")
        for derman in dermanlar:
            say = drug_map.get((hekim.id, derman.id), 0)
            faizli = (Decimal(say) * faktor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            d_say.append(say)
            d_faizli.append(faizli)
            toplam += say
            faizli_toplam += faizli
        report_data.append({
            "hekim": hekim,
            "dermanlar": d_say,
            "faizli_dermanlar": d_faizli,
            "toplam": toplam,
            "faizli_toplam": faizli_toplam,
        })

    derman_toplamlari = [sum(r["dermanlar"][i] for r in report_data) for i in range(len(dermanlar))]
    faizli_toplamlari = [round(float(sum(r["faizli_dermanlar"][i] for r in report_data)), 2) for i in range(len(dermanlar))]
    toplam_hekim_say = sum(r["toplam"] for r in report_data)
    toplam_faizli_say = round(float(sum(r["faizli_toplam"] for r in report_data)), 2)

    satis_map = {row["drug_id"]: row["s"] for row in sales.values("drug_id").annotate(s=Sum("quantity"))}
    effektivlik_faizleri = []
    for i, derman in enumerate(dermanlar):
        satis = satis_map.get(derman.id, 0) or 0
        fp = faizli_toplamlari[i]
        eff = (Decimal(satis) / Decimal(str(fp))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if fp else Decimal("0.00")
        effektivlik_faizleri.append(eff)

    for row in report_data:
        eff_d = []
        eff_t = Decimal("0.00")
        kom_d = []
        kom_t = Decimal("0.00")
        for i, faizli in enumerate(row["faizli_dermanlar"]):
            vurulmus = (faizli * effektivlik_faizleri[i]).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            eff_d.append(vurulmus)
            eff_t += vurulmus
            kom_faiz = dermanlar[i].komissiya or Decimal("0")
            miqdar = (vurulmus * kom_faiz).quantize(Decimal("0.01"))
            kom_d.append(miqdar)
            kom_t += miqdar
        row["effektivlikli_dermanlar"] = eff_d
        row["effektivlikli_toplam"] = eff_t
        row["komissiya_miqdarlari"] = kom_d
        row["umumi_komissiya"] = kom_t
        # DB-ye yaz — yalniz Hesabla basanda
        hekim = row["hekim"]
        hekim.hesablanan_miqdar = eff_t
        hekim.hekimden_silinen = kom_t
        hekim.save(update_fields=["hesablanan_miqdar", "hekimden_silinen"])

    effektivlikli_toplamlari = [
        round(float(sum(r["effektivlikli_dermanlar"][i] for r in report_data)), 2)
        for i in range(len(dermanlar))
    ]
    toplam_effektivlikli_say = round(sum(effektivlikli_toplamlari), 2)
    komissiya_toplamlari = [
        float(sum(r["komissiya_miqdarlari"][i] for r in report_data).quantize(Decimal("0.01")))
        for i in range(len(dermanlar))
    ]
    toplam_komissiya = float(sum(r["umumi_komissiya"] for r in report_data).quantize(Decimal("0.01")))

    def flt(v):
        return float(v)

    json_hekimler = []
    for row in report_data:
        json_hekimler.append({
            "ad": row["hekim"].ad,
            "derece": row["hekim"].derece,
            "dermanlar": row["dermanlar"],
            "faizli_dermanlar": [flt(x) for x in row["faizli_dermanlar"]],
            "toplam": row["toplam"],
            "faizli_toplam": flt(row["faizli_toplam"]),
            "effektivlikli_dermanlar": [flt(x) for x in row["effektivlikli_dermanlar"]],
            "effektivlikli_toplam": flt(row["effektivlikli_toplam"]),
            "komissiya_miqdarlari": [flt(x) for x in row["komissiya_miqdarlari"]],
            "umumi_komissiya": flt(row["umumi_komissiya"]),
        })

    return JsonResponse({
        "dermanlar": [d.med_name for d in dermanlar],
        "hekimler": json_hekimler,
        "derman_toplamlari": derman_toplamlari,
        "faizli_toplamlari": faizli_toplamlari,
        "toplam_hekim_say": toplam_hekim_say,
        "toplam_faizli_say": toplam_faizli_say,
        "effektivlik_faizleri": [flt(x) for x in effektivlik_faizleri],
        "effektivlikli_toplamlari": effektivlikli_toplamlari,
        "toplam_effektivlikli_say": toplam_effektivlikli_say,
        "komissiya_toplamlari": komissiya_toplamlari,
        "toplam_komissiya": toplam_komissiya,
        "ay": ay,
        "il": il,
    }, json_dumps_params={"ensure_ascii": False})
