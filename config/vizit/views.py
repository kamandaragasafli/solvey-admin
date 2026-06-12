from datetime import datetime as dt
from urllib.parse import urlencode

from django.contrib import messages
from django.db import IntegrityError
from django.db.models import Count, F, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.db import transaction
from datetime import date
from django.contrib.auth.decorators import login_required


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from doctors.models import Doctors
from medicine.models import Medical
from regions.models import City, Hospital, Region

from .models import Istifadeci, Vizit, VizitPreparat, AptekVizit, AptekVizitPreparat
from .utils import (
    menecer_rehber_required,
    rehber_required,
    vizit_giris_yoxla,
    vizit_login_required,
    vizit_session_temizle,
    vizit_session_yaz,
)

MUNASIBAT_SECIMLERI = [c[0] for c in Vizit.MUNASIBAT_CHOICES]

ROL_BASLIQLARI = {
    Istifadeci.ROL_NUMAYENDE: 'Tibbi Nümayəndə',
    Istifadeci.ROL_MENECER: 'Menecer',
    Istifadeci.ROL_REHBER: 'Diviziya Rəhbər',
}

EXCEL_PREP_ORDER = [
    'Solseda', 'Opsidol', 'Zemovar', 'Soltrop', 'Fensavin', 'Soltep', 'Litasol',
    'Prostazolin', 'Provital', 'Heptrazol', 'Vitomer Kids', 'Fesola', 'Kartovey',
    'Betasol', 'Genosfer', 'Serrasol', 'Kalvey', 'Vitomer D3', 'Levastronq',
    'Opeblok', 'Ropsol', 'Painstop', 'Moxivista',
]


def _admin_redirect(tab):    return redirect(f"{reverse('vizit:admin_panel')}?tab={tab}")


def _ixtisas_secimleri():
    return [{'id': kod, 'kod': kod, 'ad': ad} for kod, ad in Doctors.İXTİSAS_SECIMI]


def _kateqoriya_secimleri():
    return [{'id': kod, 'ad': ad} for kod, ad in Doctors.KATEQORIYA_SECIMI]


def _klinika_tap(city):
    return (
        Hospital.objects.filter(city=city).first()
        or Hospital.objects.filter(region_net=city.region).first()
    )


def _bolgeler_for_user(user_rol, user_bolge_ids):
    # Rəhbər və Diviziya Rəhbəri hər şeyi görür
    if user_rol in (Istifadeci.ROL_REHBER, Istifadeci.ROL_DIVIZIYA_REHB):
        return Region.objects.order_by('region_name')
    
    # Nümayəndə/Menecer üçün filter tətbiq edirik
    return Region.objects.filter(pk__in=user_bolge_ids).order_by('region_name')

def _api_bolge_icazesi(request, bolge_id):
    """
    Rəhbər və Diviziya rəhbərinə hər yerə giriş icazəsi verilir.
    Digərləri yalnız öz siyahısında olan bölgəyə baxa bilər.
    """
    user_rol = request.session.get('rol')
    user_bolge_ids = request.session.get('bolge_ids', []) # İndi siyahı gəlir

    # Rəhbər və Diviziya rəhbəri üçün tam giriş
    if user_rol in (Istifadeci.ROL_REHBER, Istifadeci.ROL_DIVIZIYA_REHB):
        return True

    # Nümayəndə və Menecer üçün yoxlanış
    try:
        return int(bolge_id) in [int(b_id) for b_id in user_bolge_ids]
    except (TypeError, ValueError):
        return False

def _api_rayon_icazesi(request, rayon_id):
    city = City.objects.filter(pk=rayon_id).values_list('region_id', flat=True).first()
    if city is None:
        return False
    return _api_bolge_icazesi(request, city)


def _filtered_vizit_qs(request):
    user_id = request.session['istifadeci_id']
    user_rol = request.session.get('rol')
    user_bolge_id = request.session.get('bolge_id')
    bugun = timezone.localdate()

    filter_bolge = request.GET.get('bolge_id', '')
    if not filter_bolge and user_rol in (Istifadeci.ROL_NUMAYENDE, Istifadeci.ROL_MENECER) and user_bolge_id:
        filter_bolge = str(user_bolge_id)

    filter_rayon = request.GET.get('rayon_id', '')
    filter_tarix_bas = request.GET.get('tarix_bas') or bugun.isoformat()
    filter_tarix_son = request.GET.get('tarix_son') or bugun.isoformat()
    filter_munasibat = request.GET.get('munasibat', '')

    filter_bolge_id = int(filter_bolge) if filter_bolge else None
    filter_rayon_id = int(filter_rayon) if filter_rayon else None

    qs = Vizit.objects.filter(
        tarix__gte=filter_tarix_bas,
        tarix__lte=filter_tarix_son,
    ).select_related('hekim', 'rayon', 'bolge', 'istifadeci').prefetch_related('preparatlar__preparat')

    if user_rol == Istifadeci.ROL_NUMAYENDE:
        qs = qs.filter(istifadeci_id=user_id)
    if filter_bolge_id:
        qs = qs.filter(bolge_id=filter_bolge_id)
    if filter_rayon_id:
        qs = qs.filter(rayon_id=filter_rayon_id)
    if filter_munasibat:
        qs = qs.filter(munasibat=filter_munasibat)

    return {
        'user_id': user_id,
        'user_rol': user_rol,
        'user_bolge_id': user_bolge_id,
        'filter_bolge': filter_bolge,
        'filter_rayon': filter_rayon,
        'filter_tarix_bas': filter_tarix_bas,
        'filter_tarix_son': filter_tarix_son,
        'filter_munasibat': filter_munasibat,
        'filter_bolge_id': filter_bolge_id,
        'filter_rayon_id': filter_rayon_id,
        'bugun': bugun,
        'qs': qs,
    }


def _format_tarix_aralig(tarix_bas, tarix_son):
    try:
        bas = dt.strptime(tarix_bas, '%Y-%m-%d')
        son = dt.strptime(tarix_son, '%Y-%m-%d')
        araliq = bas.strftime('%d.%m.%Y')
        if tarix_bas != tarix_son:
            araliq += f' — {son.strftime("%d.%m.%Y")}'
        return araliq
    except ValueError:
        return tarix_bas


def _excel_prep_sira():
    db_names = set(Medical.objects.filter(status=True).values_list('med_name', flat=True))
    sira = [name for name in EXCEL_PREP_ORDER if name in db_names]
    extras = sorted(db_names - set(EXCEL_PREP_ORDER))
    return sira + extras


def _excel_export_rows(vizitler_qs):
    rows = []
    for vizit in vizitler_qs.order_by('tarix', 'vaxt'):
        rows.append({
            'hekim': vizit.hekim.ad,
            'ixtisas_kod': vizit.hekim.ixtisas,
            'kateqoriya': vizit.hekim.kategoriya,
            'rayon': vizit.rayon.get_city_name_display() if vizit.rayon else '',
            'munasibat': vizit.munasibat,
            'preps_list': [vp.preparat.med_name for vp in vizit.preparatlar.all()],
        })
    return rows


def login_view(request):
    if request.session.get('istifadeci_id'):
        if request.session.get('rol') == Istifadeci.ROL_REHBER:
            return redirect('vizit:admin_panel')
        return redirect('vizit:index')

    xeta = ''
    login_deyeri = ''

    if request.method == 'POST':
        login = request.POST.get('login', '').strip()
        sifre = request.POST.get('sifre', '').strip()
        login_deyeri = login

        if login and sifre:
            istifadeci = Istifadeci.authenticate(login, sifre)
            if istifadeci:
                vizit_session_yaz(request, istifadeci)
                if istifadeci.rol == Istifadeci.ROL_REHBER:
                    return redirect('vizit:admin_panel')
                return redirect('vizit:index')
            xeta = 'Login və ya şifrə yanlışdır!'
        else:
            xeta = 'Bütün sahələri doldurun!'

    return render(
        request,
        'vizit/login.html',
        {'xeta': xeta, 'login_deyeri': login_deyeri},
    )


def logout_view(request):
    vizit_session_temizle(request)
    return redirect('vizit:login')


@rehber_required
def admin_panel_view(request):
    tab = request.GET.get('tab', 'istifadeciler')

    if request.method == 'POST':
        # 1. İSTİFADƏÇİ ƏLAVƏ ETMƏ
        if 'add_user' in request.POST:
            login = request.POST.get('login', '').strip()
            sifre = request.POST.get('sifre', '').strip()
            ad = request.POST.get('ad', '').strip()
            rol = request.POST.get('rol', Istifadeci.ROL_NUMAYENDE)
            
            # Seçilmiş bölgələri alırıq
            bolge_ids = request.POST.getlist('bolge_ids')
            valid_bolge_ids = [int(b_id) for b_id in bolge_ids if b_id]

            # Rol əsaslı məhdudiyyətlər
            if rol == Istifadeci.ROL_MENECER and len(valid_bolge_ids) > 3:
                messages.error(request, '❌ Menecer rolu üçün maksimum 3 bölgə seçilə bilər.')
                return _admin_redirect(tab)
            elif rol == Istifadeci.ROL_DIVIZIYA_REHB and len(valid_bolge_ids) > 5:
                messages.error(request, '❌ Diviziya Rəhbəri rolu üçün maksimum 3 bölgə seçilə bilər.')
                return _admin_redirect(tab)

            if login and sifre and ad:
                try:
                    istifadeci = Istifadeci(login=login, ad=ad, rol=rol, aktiv=True)
                    istifadeci.set_password(sifre)
                    istifadeci.save()
                    
                    if valid_bolge_ids:
                        istifadeci.bolgeler.set(valid_bolge_ids)
                    
                    messages.success(request, '✅ İstifadəçi uğurla əlavə edildi.')
                except IntegrityError:
                    messages.error(request, '❌ Xəta: login artıq mövcuddur.')
            else:
                messages.error(request, '❌ Bütün sahələri doldurun.')
            return _admin_redirect(tab)

        # 2. HƏKİM ƏLAVƏ ETMƏ
        elif 'add_hekim' in request.POST:
            ad_soyad = request.POST.get('ad_soyad', '').strip()
            rayon_id = request.POST.get('rayon_id', '')
            ixtisas = request.POST.get('ixtisas_id', '').strip() or 'TE'
            kategoriya = request.POST.get('kateqoriya_id', '').strip() or 'B'

            if ad_soyad and rayon_id:
                city = City.objects.select_related('region').filter(pk=int(rayon_id)).first()
                if not city:
                    messages.error(request, '❌ Rayon tapılmadı.')
                else:
                    klinika = _klinika_tap(city)
                    if not klinika:
                        messages.error(request, '❌ Bu rayon/bölgə üçün xəstəxana tapılmadı.')
                    else:
                        Doctors.objects.create(
                            ad=ad_soyad[:100],
                            bolge=city.region,
                            city=city,
                            klinika=klinika,
                            ixtisas=ixtisas,
                            kategoriya=kategoriya,
                            is_active=True,
                        )
                        messages.success(request, '✅ Həkim əlavə edildi.')
            else:
                messages.error(request, '❌ Ad/soyad və rayon mütləqdir.')
            return _admin_redirect(tab)

    # 3. İSTİFADƏÇİ SİLMƏ (GET metodu ilə)
    if 'del_user' in request.GET:
        Istifadeci.objects.filter(pk=int(request.GET['del_user'])).delete()
        messages.success(request, '✅ İstifadəçi silindi.')
        return _admin_redirect('istifadeciler')

    # 4. SƏHİFƏNİ YÜKLƏMƏ
    return render(
        request,
        'vizit/admin_panel.html',
        {
            'tab': tab,
            'istifadeciler': Istifadeci.objects.prefetch_related('bolgeler').order_by('rol', 'ad'),
            'bolgeler': Region.objects.order_by('region_name'),
            'rayonlar': City.objects.select_related('region').order_by('region__region_name', 'city_name'),
            'ixtisaslar': _ixtisas_secimleri(),
            'kateqoriyalar': _kateqoriya_secimleri(),
            'hekimler': (
                Doctors.objects.filter(is_active=True)
                .select_related('city', 'bolge')
                .annotate(
                    rayon_ad=Coalesce(F('city__city_name'), Value('—')),
                    ix_kod=F('ixtisas'),
                    kat=F('kategoriya'),
                )
                .order_by('city__city_name', 'ad')
            ),
        },
    )
@vizit_login_required
def yeni_vizit_view(request):
    user_id = request.session.get('istifadeci_id')
    user_rol = request.session.get('rol')
    user_bolge_ids = request.session.get('bolge_ids', [])
    bugun = timezone.localdate()

    # POST sorğusu: Viziti qeyd et
    if request.method == 'POST' and 'vizit_bagla' in request.POST:
        hekim_id = request.POST.get('hekim_id')
        bolge_id = request.POST.get('bolge_id')
        munasibat = request.POST.get('munasibat', '')
        preparatlar = request.POST.getlist('preparatlar[]')

        # Seçimi sessiyada yadda saxlayırıq ki, növbəti dəfə avtomatik seçilsin
        if bolge_id:
            request.session['son_bolge_id'] = bolge_id

        if not hekim_id or not bolge_id or not preparatlar:
            messages.error(request, '❌ Bütün sahələri düzgün doldurun.')
            return redirect('vizit:index')

        try:
            with transaction.atomic():
                vizit = Vizit.objects.create(
                    istifadeci_id=user_id,
                    hekim_id=int(hekim_id),
                    bolge_id=int(bolge_id),
                    munasibat=munasibat,
                    tarix=bugun,
                    vaxt=timezone.localtime().time().replace(second=0, microsecond=0),
                )
                
                VizitPreparat.objects.bulk_create([
                    VizitPreparat(vizit=vizit, preparat_id=int(pid)) for pid in preparatlar
                ])
            
            messages.success(request, '✅ Vizit uğurla qeydə alındı!')
        except Exception as e:
            messages.error(request, f'❌ Xəta baş verdi: {e}')
        
        return redirect('vizit:index')

    # GET sorğusu və ya Səhifənin açılması
    # Vizitləri filtrələmək: Rəhbərlər hamını, digərləri yalnız özünü görsün
    vizitler_query = Vizit.objects.filter(tarix=bugun)
    
    if user_rol not in [Istifadeci.ROL_REHBER, Istifadeci.ROL_DIVIZIYA_REHB]:
        vizitler_query = vizitler_query.filter(istifadeci_id=user_id)

    # Son seçilən bölgəni sessiyadan oxuyuruq
    selected_bolge_id = request.session.get('son_bolge_id')

    return render(request, 'vizit/create-vizit.html', {
        'bolgeler': _bolgeler_for_user(user_rol, user_bolge_ids),
        'selected_bolge_id': int(selected_bolge_id) if selected_bolge_id else None,
        'preparatlar_siyahisi': Medical.objects.filter(status=True).order_by('med_name'),
        'bugun_vizitler': vizitler_query.order_by('-id'),
        'user_rol': user_rol,
        'bugun_tarix': bugun,
    })


def _rayonlar_list(request, bolge_id):
    if not bolge_id or not _api_bolge_icazesi(request, bolge_id):
        return []
    return [
        {'id': c.id, 'ad': c.get_city_name_display()}
        for c in City.objects.filter(region_id=bolge_id).order_by('city_name')
    ]


def _hekimler_list(request, bolge_id=None, rayon_id=None):
    if bolge_id:
        if not _api_bolge_icazesi(request, bolge_id):
            return []
        qs = Doctors.objects.filter(bolge_id=bolge_id, is_active=True)
    elif rayon_id:
        if not _api_rayon_icazesi(request, rayon_id):
            return []
        qs = Doctors.objects.filter(city_id=rayon_id, is_active=True)
    else:
        return []
    return [
        {
            'id': d.id,
            'ad_soyad': d.ad,
            'ixtisas_kod': d.ixtisas,
            'kateqoriya': d.kategoriya,
            'derece': d.derece,
        }
        for d in qs.order_by('ad')
    ]


@vizit_login_required
def get_rayonlar_api(request):
    return JsonResponse({'rayonlar': _rayonlar_list(request, request.GET.get('bolge_id'))})


@vizit_login_required
def get_hekimler_api(request):
    return JsonResponse({
        'hekimler': _hekimler_list(
            request,
            bolge_id=request.GET.get('bolge_id'),
            rayon_id=request.GET.get('rayon_id'),
        )
    })


@vizit_login_required
def ajax_compat_view(request):
    """Köhnə vizit/ajax.php ilə eyni ?action= parametrləri (JSON massiv)."""
    action = request.GET.get('action', '')
    if action == 'rayonlar':
        return JsonResponse(
            _rayonlar_list(request, request.GET.get('bolge_id')),
            safe=False,
        )
    if action == 'hekimler':
        return JsonResponse(
            _hekimler_list(
                request,
                bolge_id=request.GET.get('bolge_id'),
                rayon_id=request.GET.get('rayon_id'),
            ),
            safe=False,
        )
    return JsonResponse([], safe=False)

@vizit_login_required
def hesabat_view(request):
    f = _filtered_vizit_qs(request)
    user_rol = f['user_rol']
    user_bolge_id = f['user_bolge_id']
    vizitler_qs = f['qs']

    # --- TƏHLÜKƏSİZLİK YOXLAMASI: None xətasını aradan qaldırırıq ---
    # Əgər user_bolge_id tək bir dəyərdirsə siyahıya çeviririk, None-dursa boş siyahı edirik
    if user_bolge_id is None:
        bolge_list = []
    elif isinstance(user_bolge_id, (list, tuple)):
        bolge_list = user_bolge_id
    else:
        bolge_list = [user_bolge_id]

    # Bölgələri və rayonları təyin edirik
    bolgeler = _bolgeler_for_user(user_rol, bolge_list)
    
    rayonlar = (
        City.objects.filter(region_id=f['filter_bolge_id']).order_by('city_name')
        if f['filter_bolge_id']
        else City.objects.none()
    )

    # Vizitləri sıralayırıq
    vizitler = vizitler_qs.order_by('-tarix', '-vaxt')
    total_vizit = vizitler.count()

    # Preparat statistikası
    try:
        prep_stat = [
            {'ad': row['preparat__med_name'], 'c': row['c']}
            for row in (
                VizitPreparat.objects.filter(vizit__in=vizitler_qs.values('pk'))
                .values('preparat__med_name')
                .annotate(c=Count('id'))
                .order_by('-c')[:8]
            )
        ]
    except Exception:
        prep_stat = []

    return render(
        request,
        'vizit/hesabat.html',
        {
            'bolgeler': bolgeler,
            'rayonlar': rayonlar,
            'vizitler': vizitler,
            'total_vizit': total_vizit,
            'prep_stat': prep_stat,
            'munasibatler': MUNASIBAT_SECIMLERI,
            'user_rol': user_rol,
            'user_bolge_id': user_bolge_id,
            'filter_bolge': f['filter_bolge'],
            'filter_rayon': f['filter_rayon'],
            'filter_tarix_bas': f['filter_tarix_bas'],
            'filter_tarix_son': f['filter_tarix_son'],
            'filter_munasibat': f['filter_munasibat'],
            'excel_url_params': urlencode(request.GET),
        },
    )


@vizit_login_required
def excel_export_view(request):
    f = _filtered_vizit_qs(request)
    prep_sira = _excel_prep_sira()
    vizitler = _excel_export_rows(f['qs'])

    tarix_aralig = _format_tarix_aralig(f['filter_tarix_bas'], f['filter_tarix_son'])
    rol_basliq = ROL_BASLIQLARI.get(f['user_rol'], 'Vizit')

    clean_tarix = tarix_aralig.replace(' ', '_').replace('.', '_').replace('—', '-')
    file_name = f'Vizit_Hesabat_{clean_tarix}.xls'

    response = render(
        request,
        'vizit/export_excel.html',
        {
            'rol_basliq': rol_basliq,
            'tarix_aralig': tarix_aralig,
            'prep_sira': prep_sira,
            'vizitler': vizitler,
            'colspan_count': 6 + len(prep_sira),
        },
    )
    response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-8'
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    response['Cache-Control'] = 'max-age=0'
    response.content = b'\xef\xbb\xbf' + response.content
    return response


def _bolge_stat_data(bolge_id, tarix_bas, tarix_son):
    vizitler = Vizit.objects.filter(
        bolge_id=bolge_id,
        tarix__gte=tarix_bas,
        tarix__lte=tarix_son,
    ).select_related('istifadeci', 'rayon')

    by_user = {}
    for vizit in vizitler:
        uid = vizit.istifadeci_id
        if uid not in by_user:
            by_user[uid] = {
                'numayende': vizit.istifadeci.ad,
                'vizit_sayi': 0,
                'rayonlar': set(),
            }
        by_user[uid]['vizit_sayi'] += 1
        if vizit.rayon:
            by_user[uid]['rayonlar'].add(vizit.rayon.get_city_name_display())

    numayende_stat = sorted(
        [
            {
                'numayende': row['numayende'],
                'vizit_sayi': row['vizit_sayi'],
                'rayonlar': ', '.join(sorted(row['rayonlar'])),
            }
            for row in by_user.values()
        ],
        key=lambda row: -row['vizit_sayi'],
    )

    rayon_rows = (
        vizitler.values('rayon_id')
        .annotate(c=Count('id'))
        .order_by('-c')
    )
    rayon_ids = [row['rayon_id'] for row in rayon_rows if row['rayon_id']]
    city_map = {
        city.pk: city.get_city_name_display()
        for city in City.objects.filter(pk__in=rayon_ids)
    }
    rayon_stat = [
        {'rayon': city_map.get(row['rayon_id'], '—'), 'c': row['c']}
        for row in rayon_rows
    ]

    return numayende_stat, rayon_stat

@menecer_rehber_required
def bolge_stat_view(request):
    user_rol = request.session.get('rol')
    # SESSİYADAN İD-LƏRİ SİYAHI KİMİ ALIN
    user_bolge_ids = request.session.get('bolge_ids', []) 
    
    today = timezone.localdate()
    default_tarix_bas = today.replace(day=1).isoformat()
    default_tarix_son = today.isoformat()

    filter_bolge = request.GET.get('bolge_id', '')
    
    # Əgər menecerdirsə və filter seçilməyibsə, sessiyadakı ilk bölgəni seçirik
    if not filter_bolge and user_rol == Istifadeci.ROL_MENECER and user_bolge_ids:
        filter_bolge = str(user_bolge_ids[0])

    filter_tarix_bas = request.GET.get('tarix_bas') or default_tarix_bas
    filter_tarix_son = request.GET.get('tarix_son') or default_tarix_son

    # --- TƏHLÜKƏSİZLİK YOXLAMASI ---
    # _bolgeler_for_user mütləq siyahı gözləyir
    bolgeler = _bolgeler_for_user(user_rol, user_bolge_ids if isinstance(user_bolge_ids, list) else [])

    numayende_stat = []
    rayon_stat = []
    
    if filter_bolge:
        try:
            numayende_stat, rayon_stat = _bolge_stat_data(
                int(filter_bolge), filter_tarix_bas, filter_tarix_son
            )
        except (ValueError, TypeError):
            messages.error(request, "❌ Seçilmiş bölgə məlumatları tapılmadı.")

    return render(
        request,
        'vizit/bolge_stat.html',
        {
            'bolgeler': bolgeler,
            'numayende_stat': numayende_stat,
            'rayon_stat': rayon_stat,
            'user_rol': user_rol,
            'filter_bolge': filter_bolge,
            'filter_tarix_bas': filter_tarix_bas,
            'filter_tarix_son': filter_tarix_son,
        },
    )

@vizit_login_required
def aptek_vizit_view(request):
    dermanlar = Medical.objects.filter(status=True).order_by('med_name')
    return render(request, 'vizit/aptek-vizit.html', {
        'dermanlar': dermanlar,
    })


@vizit_login_required
def statistika(request):
    
    user_id = request.session.get('istifadeci_id')
    user_rol = request.session.get('rol')
    user_bolge_ids = request.session.get('bolge_ids', [])
    
    bolge_id = request.GET.get('bolge_id') or (user_bolge_ids[0] if user_rol != Istifadeci.ROL_REHBER else None)
    tarix_bas = request.GET.get('tarix_bas') or date.today().replace(day=1).isoformat()
    tarix_son = request.GET.get('tarix_son') or date.today().isoformat()

    context = {'bolge_id': bolge_id, 'tarix_bas': tarix_bas, 'tarix_son': tarix_son}
    
    if user_rol == Istifadeci.ROL_REHBER:
        context['bolgeler'] = Region.objects.all().order_by('region_name')
    
    if bolge_id:
        # Nümayəndə statistikası
        context['num_stat'] = AptekVizit.objects.filter(bolge_id=bolge_id, tarix__range=[tarix_bas, tarix_son]) \
            .values('user__ad') \
            .annotate(c=Count('id')) \
            .order_by('-c')

        # Rayon statistikası
        context['rayon_stat'] = AptekVizit.objects.filter(bolge_id=bolge_id, tarix__range=[tarix_bas, tarix_son]) \
            .values('rayon__city_name') \
            .annotate(c=Count('id')) \
            .order_by('-c')

        # Top apteklər
        context['aptek_stat'] = AptekVizit.objects.filter(bolge_id=bolge_id, tarix__range=[tarix_bas, tarix_son]) \
            .values('aptek_ad') \
            .annotate(c=Count('aptek_ad')) \
            .order_by('-c')[:10]

    return render(request, 'vizit/aptek-stats.html', context)

@login_required
def ajax_rayonlar(request):
    bolge_id = request.GET.get('bolge_id')
    if not bolge_id:
        return JsonResponse([], safe=False)
    
    rayonlar = City.objects.filter(
        region_id=bolge_id
    ).order_by('city_name').values('id', 'city_name')
    
    data = [{'id': r['id'], 'ad': r['city_name']} for r in rayonlar]
    return JsonResponse(data, safe=False)

@login_required
def yeni_aptek_vizit(request):
    user_id = request.session.get('istifadeci_id')
    user_rol = request.session.get('rol')
    user_bolge_ids = request.session.get('bolge_ids', [])

    # Bakı bölgə ID-lərini təyin et
    baki_bolge_ids = list(
        Region.objects.filter(region_type='Bakı')
        .values_list('id', flat=True)
    )

    if request.method == 'POST':
        try:
            with transaction.atomic():
                bolge_id      = request.POST.get('bolge_id')
                rayon_id      = request.POST.get('rayon_id') or None
                aptek_ad      = request.POST.get('aptek_ad', '').strip()
                aptek_nomre   = request.POST.get('aptek_nomre', '').strip()
                ref_veziyyeti = request.POST.get('ref_veziyyeti', '').strip()
                aptek_iscisi  = request.POST.get('aptek_iscisi', '').strip()
                qeyd          = request.POST.get('qeyd', '').strip()

                # Validasiya
                if not bolge_id:
                    messages.error(request, "Bölgə seçilməyib!")
                    return redirect('vizit:yeni_aptek_vizit')
                
                if not aptek_ad:
                    messages.error(request, "Aptekin adı daxil edilməyib!")
                    return redirect('vizit:yeni_aptek_vizit')
                
                if not ref_veziyyeti:
                    messages.error(request, "Rəf vəziyyəti seçilməyib!")
                    return redirect('vizit:yeni_aptek_vizit')

                # Bakı yoxlaması
                bolge = Region.objects.filter(id=bolge_id).first()
                is_baki = bolge and bolge.region_type == 'Bakı'
                
                if not is_baki and not rayon_id:
                    messages.error(request, "Rayon seçilməyib!")
                    return redirect('vizit:yeni_aptek_vizit')
                
                if is_baki:
                    rayon_id = None

                # AptekVizit yarat
                vizit = AptekVizit.objects.create(
                    user_id=user_id,
                    bolge_id=bolge_id,
                    rayon_id=rayon_id,
                    aptek_ad=aptek_ad,
                    aptek_nomre=aptek_nomre or None,
                    tarix=timezone.now().date(),
                    vaxt=timezone.now().time(),
                    qeyd=qeyd or None,
                )

                # Preparatları saxla
                sorusulub_ids = request.POST.getlist('sorusulub')
                satilib_ids   = request.POST.getlist('satilib')
                yoxdur_ids    = request.POST.getlist('yoxdur')

                all_ids = set(sorusulub_ids + satilib_ids + yoxdur_ids)

                objs = []
                for pid in all_ids:
                    if not pid:
                        continue
                    objs.append(AptekVizitPreparat(
                        aptek_vizit=vizit,
                        preparat_id=int(pid),
                        sorusulub=pid in sorusulub_ids,
                        satilib=pid in satilib_ids,
                        movcuddur=pid not in yoxdur_ids,
                        ref_vez=ref_veziyyeti,
                        aptek_iscisi=aptek_iscisi or None,
                        qeyd=qeyd or None,
                    ))

                AptekVizitPreparat.objects.bulk_create(objs)
                
                messages.success(request, "✅ Aptek viziti uğurla qeydə alındı!")
                return redirect('vizit:yeni_aptek_vizit')

        except Exception as e:
            messages.error(request, f"Xəta baş verdi: {str(e)}")
            return redirect('vizit:yeni_aptek_vizit')

    # ✅ GET - Bölgələri rol və user-ə görə filtrələ
    bolgeler = _bolgeler_for_user(user_rol, user_bolge_ids)
    
    # ✅ Yalnız aktiv dərmanları göstər
    preparatlar = Medical.objects.filter(status=True).order_by('med_name')
    
    # ✅ Bugünkü vizitləri filtrələ
    bugun_vizitler = AptekVizit.objects.filter(
        tarix=date.today()
    ).select_related('rayon', 'bolge', 'user').order_by('-id')

    # ✅ Rəhbər deyilsə, yalnız öz vizitlərini görsün
    if user_rol not in [Istifadeci.ROL_REHBER, Istifadeci.ROL_DIVIZIYA_REHB]:
        bugun_vizitler = bugun_vizitler.filter(user_id=user_id)

    return render(request, 'vizit/aptek-vizit.html', {
        'bolgeler': bolgeler,              # ✅ Filtrlənmiş bölgələr
        'preparatlar': preparatlar,        # ✅ Aktiv dərmanlar
        'bugun_vizitler': bugun_vizitler,  # ✅ Filtrlənmiş vizitlər
        'today': date.today(),
        'baki_bolge_ids': baki_bolge_ids,
    })
def export_to_excel(request):
    # 1. Filtrləri qəbul edirik
    bolge_id = request.GET.get('bolge_id')
    tarix_bas = request.GET.get('tarix_bas') or date.today().replace(day=1).isoformat()
    tarix_son = request.GET.get('tarix_son') or date.today().isoformat()
    
    # 2. Workbook yaradılması və şəbəkə xətləri
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aptek Vizit Blanki"
    ws.views.sheetView[0].showGridLines = True
    
    # Dizayn və Rəng Palitrası (Şəkildəki kimi)
    HEADER_BLUE = "205478"     
    MAIN_TEAL = "1ABC9C"       
    fill_title = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    fill_header = PatternFill(start_color=MAIN_TEAL, end_color=MAIN_TEAL, fill_type="solid")
    fill_meta = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_meta = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_data = Font(name="Segoe UI", size=10)
    
    # ✓ və ✗ üçün rəngli fontlar
    font_check = Font(name="Segoe UI", size=11, bold=True, color="1B5E20") 
    font_cross = Font(name="Segoe UI", size=11, bold=True, color="B71C1C") 
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )
    
    # 3. Filtrlərə görə Ana Vizitləri select_related ilə sürətli çəkirik
    vizit_queryset = AptekVizit.objects.select_related('user', 'rayon', 'bolge').all()
    if bolge_id:
        vizit_queryset = vizit_queryset.filter(bolge_id=bolge_id)
    if tarix_bas and tarix_son:
        vizit_queryset = vizit_queryset.filter(tarix__range=[tarix_bas, tarix_son])
        
    vizitler = vizit_queryset.order_by('tarix', 'vaxt')
    
    # 4. ÜST BANNER (Row 1)
    ws.merge_cells("A1:CA1")
    ws["A1"] = f"SOLVEY PHARMA — APTEK VİZİT BLANKI | {tarix_bas} - {tarix_son}"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # 5. META MƏLUMATLAR (Row 3)
    bolge_adi = "Bütün bölgələr"
    if bolge_id:
        try:
            bolge_adi = Region.objects.get(id=bolge_id).region_name
        except:
            pass
            
    ws.merge_cells("A3:B3")
    ws["A3"] = f"Bölgə: {bolge_adi}"
    ws["A3"].font = font_meta
    ws["A3"].fill = fill_meta
    ws["A3"].border = thin_border
    
    ws.merge_cells("C3:E3")
    ws["C3"] = f"Tarix: {tarix_bas} / {tarix_son}"
    ws["C3"].font = font_meta
    ws["C3"].fill = fill_meta
    ws["C3"].border = thin_border
    ws.row_dimensions[3].height = 22
    
    # 6. ƏSAS SOL BAŞLIQLAR (Row 4 və Row 5 Birləşir)
    base_headers = ["#", "Aptekin adı və nömrəsi", "Rayon", "Rəfdəki vəziyyət", "Aptek işçisi", "Vaxt", "Nümayəndə"]
    for i, h in enumerate(base_headers):
        col = get_column_letter(i + 1)
        ws.merge_cells(f"{col}4:{col}5")
        ws[f"{col}4"] = h
        ws[f"{col}4"].font = font_header
        ws[f"{col}4"].fill = fill_header
        ws[f"{col}4"].alignment = align_center
        ws[f"{col}4"].border = thin_border
        
    # 7. DƏRMAN BAŞLIQLARI (Dinamik - Medical Modelindən)
    preparatlar = Medical.objects.all().order_by('med_name')
    current_col = 8 # H sütunundan başlayır (8-ci sütun)
    
    preparat_col_mapping = {} 
    for p in preparatlar:
        start_col = get_column_letter(current_col)
        end_col = get_column_letter(current_col + 2)
        
        # Dərmanın Adı (Üst Sətir)
        ws.merge_cells(f"{start_col}4:{end_col}4")
        ws[f"{start_col}4"] = p.med_name
        ws[f"{start_col}4"].font = font_header
        ws[f"{start_col}4"].fill = fill_header
        ws[f"{start_col}4"].alignment = align_center
        
        # Alt Başlıqlar
        sub_headers = ["Soruşulan", "Satılıb", "Mövcud"]
        for j, sub in enumerate(sub_headers):
            sub_col = get_column_letter(current_col + j)
            ws[f"{sub_col}5"] = sub
            ws[f"{sub_col}5"].font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
            ws[f"{sub_col}5"].fill = fill_header
            ws[f"{sub_col}5"].alignment = align_center
            
        preparat_col_mapping[p.id] = current_col
        current_col += 3
        
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 22
    
    # Başlıqlara border verilməsi
    for r in [4, 5]:
        for c in range(1, current_col):
            ws.cell(row=r, column=c).border = thin_border
            
    # 8. SƏTİRLƏRİN REAL DATA İLƏ DOLDURULMASI
    row_num = 6
    for idx, v in enumerate(vizitler, start=1):
        
        # Bu vizitə bağlı olan bütün preparat qeydlərini (`related_name='preparatlar'`) çəkirik
        vizit_preparatlari = v.preparatlar.all()
        
        # Excel-dəki "Rəfdəki vəziyyət" və "Aptek işçisi" sütunları üçün dataları götürürük
        # (Vizitə aid ilk dərmanın qeydini və ya tapılan ilk boş olmayan dəyəri götürür)
        ref_veziyyeti_degeri = ""
        aptek_iscisi_degeri = ""
        for vp in vizit_preparatlari:
            if vp.ref_vez:
                ref_veziyyeti_degeri = vp.ref_vez
            if vp.aptek_iscisi:
                aptek_iscisi_degeri = vp.aptek_iscisi
            if ref_veziyyeti_degeri and aptek_iscisi_degeri:
                break

        # Sətir məlumatlarının yazılması
        ws.cell(row=row_num, column=1, value=idx).alignment = align_center
        
        aptek_tam_ad = f"{v.aptek_ad} {v.aptek_nomre or ''}".strip()
        ws.cell(row=row_num, column=2, value=aptek_tam_ad).alignment = align_left
        
        rayon_ad = v.rayon.city_name if v.rayon else ""
        ws.cell(row=row_num, column=3, value=rayon_ad).alignment = align_center
        
        # Sizin istədiyiniz model dəyişəni bura oturtuldu: `v.ref_vez` əvəzinə `ref_veziyyeti_degeri`
        ws.cell(row=row_num, column=4, value=ref_veziyyeti_degeri).alignment = align_center
        ws.cell(row=row_num, column=5, value=aptek_iscisi_degeri).alignment = align_left
        
        # Tarix və Vaxtın birləşdirilməsi formatı (Məs: 02.06.2026 23:40)
        tarix_str = v.tarix.strftime("%d.%m.%Y") if v.tarix else ""
        vaxt_str = v.vaxt.strftime("%H:%M") if v.vaxt else ""
        vaxt_tam = f"{tarix_str} {vaxt_str}".strip()
        ws.cell(row=row_num, column=6, value=vaxt_tam).alignment = align_center
        
        numayende_ad = v.user.ad if v.user else ""
        ws.cell(row=row_num, column=7, value=numayende_ad).alignment = align_left
        
        # Sol tərəfə stil vermək
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font = font_data
            ws.cell(row=row_num, column=c).border = thin_border
            
        # ── MATRİSİN (DƏRMAN STATUSLARININ) DOLDURULMASI ──
        # Vizit preparatlarını sürətli axtarış üçün dictionary halına salırıq
        vp_dict = {vp.preparat_id: vp for vp in vizit_preparatlari}
        
        for p_id, col_start in preparat_col_mapping.items():
            s_cell = ws.cell(row=row_num, column=col_start)
            sat_cell = ws.cell(row=row_num, column=col_start + 1)
            m_cell = ws.cell(row=row_num, column=col_start + 2)
            
            # Hər üç hüceyrə üçün ilkin border və nizamlamanı veririk
            for cell in [s_cell, sat_cell, m_cell]:
                cell.alignment = align_center
                cell.border = thin_border
            
            # Əgər bu vizitdə bu dərmana aid qeyd (sətir) varsa:
            if p_id in vp_dict:
                vp_obj = vp_dict[p_id]
                
                # 1. Soruşulubsa "✓" yaz
                if vp_obj.sorusulub:
                    s_cell.value = "✓"
                    s_cell.font = font_check
                    
                # 2. Satılıbsa "✓" yaz
                if vp_obj.satilib:
                    sat_cell.value = "✓"
                    sat_cell.font = font_check
                    
                # 3. Mövcuddursa "✓", yoxsa "✗" yaz
                if vp_obj.movcuddur:
                    m_cell.value = "✓"
                    m_cell.font = font_check
                else:
                    m_cell.value = "✗"
                    m_cell.font = font_cross
            else:
                # Əgər vizit zamanı bu dərman haqqında heç bir qeyd daxil edilməyibsə, 
                # "Mövcuddur" bölməsini standart olaraq boş və ya şəkildəki kimi nəzarətsiz buraxırıq.
                pass
                
        ws.row_dimensions[row_num].height = 24
        row_num += 1
        
    # 9. Sütun ölçülərinin avtomatik nizamlanması
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column <= 7:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        else:
            ws.column_dimensions[col_letter].width = 10 
            
    # H6 sütunundan paneli dondururuq
    ws.freeze_panes = "H6"
    
    # 10. Brauzerə ötürmə
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Solvey_Aptek_Hesabat_{tarix_bas}.xlsx"'
    wb.save(response)
    return response